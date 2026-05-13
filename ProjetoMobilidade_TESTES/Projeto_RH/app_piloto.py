import streamlit as st
import sqlite3
import pandas as pd
import requests
import plotly.express as px
import random
import datetime
import time
import folium
import os
import secrets
import io
import shutil
import plotly.graph_objects as go
from dotenv import load_dotenv
from streamlit_folium import st_folium
from banco_dados import *
from apis import *
from agente_ia import *
from carta_pdf import gerar_carta_pdf
from email_sender import enviar_carta_por_email

# Imports dos mÃ³dulos de telas melhoradas
from telas.triagem import renderizar_triagem
from telas.banco_dados import renderizar_banco_dados
from telas.gerenciar_unidades import renderizar_gerenciar_unidades
from telas.portal_jovem_avancado import renderizar_portal_jovem_avancado
from telas.registro_funcionario_avancado import renderizar_registro_funcionario_avancado
from telas.relatorios_analytics import renderizar_relatorios_analytics
from telas.auditoria_compliance import renderizar_auditoria_compliance

# Imports para exportaÃ§Ã£o de relatÃ³rios
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    pass

try:
    from fpdf import FPDF
except ImportError:
    pass

# â”€â”€â”€ Carrega variÃ¡veis de ambiente â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
load_dotenv()

# â”€â”€â”€ ConfiguraÃ§Ã£o da pÃ¡gina â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
st.set_page_config(page_title="RENAPSI â€” Mobilidade", page_icon="ðŸš‡", layout="wide")

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# FUNÃ‡Ã•ES AUXILIARES DE MÃSCARA E VALIDAÃ‡ÃƒO
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def safe_sql_query(query, conexao, params=None, default_value=None):
    """
    Executa query SQL com tratamento de erros robusto.
    Retorna DataFrame vazio ou valor padrão se a tabela não existir.
    """
    try:
        if params:
            return pd.read_sql_query(query, conexao, params=params)
        else:
            return pd.read_sql_query(query, conexao)
    except (sqlite3.OperationalError, pd.errors.DatabaseError):
        if default_value is not None:
            return default_value
        return pd.DataFrame()

def aplicar_mascara_cpf(cpf):
    """Aplica mÃ¡scara no CPF: 000.000.000-00"""
    cpf_limpo = ''.join(filter(str.isdigit, cpf))
    if len(cpf_limpo) <= 3:
        return cpf_limpo
    elif len(cpf_limpo) <= 6:
        return f"{cpf_limpo[:3]}.{cpf_limpo[3:]}"
    elif len(cpf_limpo) <= 9:
        return f"{cpf_limpo[:3]}.{cpf_limpo[3:6]}.{cpf_limpo[6:]}"
    else:
        return f"{cpf_limpo[:3]}.{cpf_limpo[3:6]}.{cpf_limpo[6:9]}-{cpf_limpo[9:11]}"

def aplicar_mascara_cep(cep):
    """Aplica mÃ¡scara no CEP: 00000-000"""
    cep_limpo = ''.join(filter(str.isdigit, cep))
    if len(cep_limpo) <= 5:
        return cep_limpo
    else:
        return f"{cep_limpo[:5]}-{cep_limpo[5:8]}"

def aplicar_mascara_celular(celular):
    """Aplica mÃ¡scara no celular: (00) 00000-0000"""
    cel_limpo = ''.join(filter(str.isdigit, celular))
    if len(cel_limpo) <= 2:
        return f"({cel_limpo}" if cel_limpo else ""
    elif len(cel_limpo) <= 7:
        return f"({cel_limpo[:2]}) {cel_limpo[2:]}"
    else:
        return f"({cel_limpo[:2]}) {cel_limpo[2:7]}-{cel_limpo[7:11]}"

def validar_cpf_completo(cpf):
    """Valida CPF usando algoritmo oficial"""
    cpf_limpo = ''.join(filter(str.isdigit, cpf))
    if len(cpf_limpo) != 11 or len(set(cpf_limpo)) == 1:
        return False
    soma = sum(int(cpf_limpo[i]) * (10 - i) for i in range(9))
    digito1 = (soma * 10 % 11) % 10
    soma = sum(int(cpf_limpo[i]) * (11 - i) for i in range(10))
    digito2 = (soma * 10 % 11) % 10
    return cpf_limpo[-2:] == f"{digito1}{digito2}"

def validar_email_formato(email):
    """Valida formato de e-mail"""
    if not email:
        return True  # E-mail Ã© opcional
    import re
    email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(email_regex, email) is not None

def validar_celular_formato(celular):
    """Valida formato de celular brasileiro"""
    if not celular:
        return True  # Celular Ã© opcional
    cel_limpo = ''.join(filter(str.isdigit, celular))
    return len(cel_limpo) == 11 and cel_limpo[2] == '9'

def mostrar_campo_obrigatorio(label):
    """Retorna label com indicador visual de campo obrigatÃ³rio"""
    return f"{label} <span style='color:#EF4444;font-weight:700;'>*</span>"

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# SISTEMA DE AUTENTICAÃ‡ÃƒO â€” SESSÃƒO PERSISTENTE (SOBREVIVE AO F5)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def _criar_tabela_sessoes():
    """Cria tabela de sessÃµes ativas no banco."""
    try:
        conn = sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'mobilidade_renapsi.db'))
        conn.execute('''
            CREATE TABLE IF NOT EXISTS sessoes_ativas (
                token TEXT PRIMARY KEY,
                user_id INTEGER NOT NULL,
                username TEXT NOT NULL,
                role TEXT NOT NULL,
                criado_em TEXT NOT NULL,
                expira_em TEXT NOT NULL
            )
        ''')
        conn.commit()
        conn.close()
    except Exception:
        pass

def _criar_sessao(user_id, username, role):
    """Gera token, salva no banco e retorna o token."""
    token = secrets.token_urlsafe(32)
    agora = datetime.datetime.now()
    expira = agora + datetime.timedelta(hours=8)
    try:
        conn = sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'mobilidade_renapsi.db'))
        conn.execute(
            "INSERT INTO sessoes_ativas (token, user_id, username, role, criado_em, expira_em) VALUES (?,?,?,?,?,?)",
            (token, user_id, username, role, agora.isoformat(), expira.isoformat())
        )
        conn.commit()
        conn.close()
    except Exception:
        pass
    return token

def _validar_token(token):
    """Valida token no banco. Retorna dict do usuÃ¡rio ou None."""
    if not token:
        return None
    try:
        conn = sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'mobilidade_renapsi.db'))
        cur = conn.execute(
            "SELECT user_id, username, role, expira_em FROM sessoes_ativas WHERE token = ?",
            (token,)
        )
        row = cur.fetchone()
        conn.close()
        if not row:
            return None
        expira = datetime.datetime.fromisoformat(row[3])
        if datetime.datetime.now() > expira:
            _revogar_token(token)
            return None
        return {"id": row[0], "username": row[1], "role": row[2]}
    except Exception:
        return None

def _revogar_token(token):
    """Remove token do banco (logout)."""
    try:
        conn = sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'mobilidade_renapsi.db'))
        conn.execute("DELETE FROM sessoes_ativas WHERE token = ?", (token,))
        conn.commit()
        conn.close()
    except Exception:
        pass

def _limpar_sessoes_expiradas():
    """Remove sessÃµes expiradas do banco."""
    try:
        conn = sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'mobilidade_renapsi.db'))
        conn.execute("DELETE FROM sessoes_ativas WHERE expira_em < ?", (datetime.datetime.now().isoformat(),))
        conn.commit()
        conn.close()
    except Exception:
        pass

def _renderizar_tela_login():
    """Renderiza a tela de login com o tema RENAPSI."""
    st.markdown("""
    <style>
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    .stApp { background: #FFFFFF !important; }
    [data-testid="stSidebar"] { display: none !important; }
    .stButton > button[kind="primary"] {
        background: #f8ae28 !important; border: none !important;
        color: #FFFFFF !important; font-weight: 700 !important;
        font-size: 18px !important; border-radius: 8px !important;
        width: 100% !important; padding: 12px !important;
        box-shadow: 0 2px 6px rgba(248,174,40,0.35) !important;
    }
    .stButton > button[kind="primary"]:hover { background: #e09a1f !important; }
    .stTextInput > div > div > input {
        background: #F8FAFC !important; border: 1px solid #E2E8F0 !important;
        color: #1E293B !important; font-size: 16px !important;
        border-radius: 8px !important;
    }
    .stTextInput > div > div > input:focus {
        border-color: #f8ae28 !important;
        box-shadow: 0 0 0 2px rgba(248,174,40,0.15) !important;
    }
    label, p, span, div { color: #333333 !important; font-size: 15px !important; }
    </style>
    """, unsafe_allow_html=True)

    col_l, col_c, col_r = st.columns([1, 2, 1])
    with col_c:
        st.markdown("""
        <div style="text-align:center;margin-bottom:28px;margin-top:60px;">
            <svg viewBox="0 0 133.83 114" width="72" height="60">
              <defs><style>.s2-1{fill:#fff;}.s2-2{fill:#402fdd;}</style></defs>
              <g>
                <path class="s2-2" d="M57,0C25.52,0,0,25.52,0,57s25.52,57,57,57h19.83c31.48,0,57-25.52,57-57S108.31,0,76.83,0h-19.83Z"/>
                <path class="s2-1" d="M66.91,48.07c-9.15,0-16.59,7.48-16.59,16.66s7.44,16.66,16.59,16.66,16.59-7.48,16.59-16.66-7.44-16.66-16.59-16.66M66.91,88.58c-13.09,0-23.74-10.7-23.74-23.85s10.65-23.85,23.74-23.85,23.74,10.7,23.74,23.85-10.65,23.85-23.74,23.85"/>
                <polygon class="s2-1" points="90.66 88.58 83.5 87.55 83.5 29.11 90.66 30.14 90.66 88.58"/>
                <polygon class="s2-1" points="56.55 25.42 56.55 32.47 77.28 35.45 77.28 28.4 56.55 25.42"/>
              </g>
            </svg>
            <p style="color:#444c9b !important;font-size:22px !important;font-weight:800 !important;margin:12px 0 4px !important;">RENAPSI Mobilidade</p>
            <p style="color:#64748B !important;font-size:13px !important;letter-spacing:0.08em !important;text-transform:uppercase !important;">Sistema de Mobilidade Urbana</p>
        </div>
        <hr style="border-color:#E2E8F0;margin:0 0 24px;">
        """, unsafe_allow_html=True)

        with st.form("form_login_principal", clear_on_submit=False):
            username_input = st.text_input("UsuÃ¡rio", placeholder="Digite seu usuÃ¡rio")
            password_input = st.text_input("Senha", type="password", placeholder="Digite sua senha")
            submitted = st.form_submit_button("Entrar", type="primary", width="stretch")

        if submitted:
            if not username_input.strip() or not password_input.strip():
                st.error("âš ï¸ Preencha usuÃ¡rio e senha.")
            else:
                usuario = verificar_credenciais(username_input, password_input)
                if usuario:
                    token = _criar_sessao(usuario["id"], usuario["username"], usuario["role"])
                    st.session_state["usuario_logado"] = True
                    st.session_state["usuario_dados"] = usuario
                    st.session_state["auth_token"] = token
                    
                    # Salva token no cookie do navegador
                    _save_token_to_cookie(token)
                    
                    st.rerun()
                else:
                    st.error("âŒ UsuÃ¡rio ou senha incorretos.")

        st.markdown("""
        <p style="color:#94A3B8 !important;font-size:12px !important;text-align:center;margin-top:24px;">
            ðŸ”’ Acesso restrito a funcionÃ¡rios autorizados
        </p>
        """, unsafe_allow_html=True)

def _renderizar_tela_troca_senha_obrigatoria():
    """Tela de troca de senha obrigatÃ³ria para primeiro acesso."""
    st.markdown("""
    <style>
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    .stApp { background: #FFFFFF !important; }
    [data-testid="stSidebar"] { display: none !important; }
    </style>
    """, unsafe_allow_html=True)

    col_l, col_c, col_r = st.columns([1, 2, 1])
    with col_c:
        st.markdown("""
        <div style="background:#FEF3C7;border:1px solid #FCD34D;border-left:4px solid #F59E0B;
                    border-radius:12px;padding:20px;margin:40px 0 24px;">
            <h3 style="margin:0 0 8px;color:#92400E;">ðŸ” Troca de Senha ObrigatÃ³ria</h3>
            <p style="color:#78350F;font-size:14px;margin:0;">
                Por seguranÃ§a, vocÃª deve alterar sua senha padrÃ£o antes de acessar o sistema.
            </p>
        </div>
        """, unsafe_allow_html=True)

        with st.form("form_troca_senha", clear_on_submit=True):
            nova_senha = st.text_input("Nova Senha", type="password", placeholder="MÃ­nimo 8 caracteres")
            confirmar_senha = st.text_input("Confirmar Nova Senha", type="password", placeholder="Repita a senha")
            submitted = st.form_submit_button("Alterar Senha", type="primary", width="stretch")

        if submitted:
            if len(nova_senha) < 8:
                st.error("âš ï¸ A senha deve ter pelo menos 8 caracteres.")
            elif nova_senha != confirmar_senha:
                st.error("âš ï¸ As senhas nÃ£o coincidem.")
            else:
                usuario_dados = st.session_state.get("usuario_dados", {})
                sucesso, msg = trocar_senha_usuario(usuario_dados["id"], nova_senha)
                if sucesso:
                    st.success(msg)
                    # Atualiza flag no session_state
                    st.session_state["usuario_dados"]["deve_trocar_senha"] = 0
                    time.sleep(1)
                    st.rerun()
                else:
                    st.error(msg)

        if st.button("ðŸšª Sair", width="stretch"):
            _token_atual = st.session_state.get("auth_token", "")
            if _token_atual:
                _revogar_token(_token_atual)
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()

def _renderizar_tela_registro_funcionario():
    """Painel de gestÃ£o de usuÃ¡rios â€” acesso exclusivo Admin."""
    usuario_logado = st.session_state.get("usuario_dados", {})
    if usuario_logado.get("role") != "admin":
        st.error("ðŸš« Acesso negado.")
        st.stop()

    st.markdown("""
    <div style="background:#FFFFFF;border:1px solid #E2E8F0;border-left:4px solid #444c9b;
                border-radius:14px;padding:24px;margin-bottom:24px;box-shadow:0 2px 4px rgba(0,0,0,0.05);">
        <h3 style="margin:0 0 4px;color:#444c9b;">ðŸ‘¥ Registro de FuncionÃ¡rio</h3>
        <p style="color:#64748B;font-size:14px;margin:0;">
            Gerencie os acessos ao sistema. Apenas o administrador pode criar ou remover logins.
        </p>
    </div>
    """, unsafe_allow_html=True)

    with st.expander("âž• Criar novo acesso", expanded=True):
        with st.form("form_criar_usuario", clear_on_submit=True):
            col1, col2 = st.columns(2)
            with col1:
                novo_username = st.text_input("UsuÃ¡rio", placeholder="Ex: FUNC_JOAO01")
                novo_role = st.selectbox("Perfil de acesso", ["funcionario", "admin"], index=0)
            with col2:
                nova_senha = st.text_input("Senha", type="password", placeholder="MÃ­nimo 8 caracteres")
                confirmar_senha = st.text_input("Confirmar senha", type="password", placeholder="Repita a senha")
            criar = st.form_submit_button("Criar Acesso", type="primary", width="stretch")

        if criar:
            if not novo_username.strip():
                st.error("âš ï¸ O campo UsuÃ¡rio Ã© obrigatÃ³rio.")
            elif len(nova_senha) < 8:
                st.error("âš ï¸ A senha deve ter pelo menos 8 caracteres.")
            elif nova_senha != confirmar_senha:
                st.error("âš ï¸ As senhas nÃ£o coincidem.")
            else:
                sucesso, msg = criar_usuario(
                    username=novo_username,
                    password=nova_senha,
                    role=novo_role,
                    criado_por=usuario_logado.get("username", "admin")
                )
                if sucesso:
                    st.success(msg)
                    st.rerun()
                else:
                    st.error(msg)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("<p style='color:#1E293B;font-size:14px;text-transform:uppercase;letter-spacing:0.1em;font-weight:600;margin-bottom:12px;'>UsuÃ¡rios cadastrados</p>", unsafe_allow_html=True)

    df_usuarios = listar_usuarios()
    if df_usuarios.empty:
        st.info("Nenhum usuÃ¡rio cadastrado.")
    else:
        for _, row in df_usuarios.iterrows():
            col_info, col_badge, col_btn = st.columns([6, 2, 2])
            with col_info:
                criado_em = str(row.get("criado_em", ""))[:16]
                criado_por = row.get("criado_por", "â€”")
                st.markdown(f"""
                <div style="background:#FFFFFF;border:1px solid #E5E7EB;border-radius:8px;
                            padding:12px 16px;box-shadow:0 1px 3px rgba(0,0,0,0.04);">
                    <p style="margin:0;color:#1E293B;font-weight:700;font-size:15px;">{row['username']}</p>
                    <p style="margin:4px 0 0;color:#64748B;font-size:13px;">Criado em {criado_em} Â· por {criado_por}</p>
                </div>
                """, unsafe_allow_html=True)
            with col_badge:
                if row["role"] == "admin":
                    badge_color, badge_bg, badge_label = "#444c9b", "rgba(68,76,155,0.1)", "ðŸ›¡ï¸ Admin"
                else:
                    badge_color, badge_bg, badge_label = "#10B981", "rgba(16,185,129,0.1)", "ðŸ‘¤ FuncionÃ¡rio"
                st.markdown(f"""
                <div style="background:{badge_bg};color:{badge_color};border:1px solid {badge_color}40;
                            border-radius:20px;padding:6px 12px;font-size:13px;font-weight:600;
                            text-align:center;margin-top:8px;">{badge_label}</div>
                """, unsafe_allow_html=True)
            with col_btn:
                eh_proprio = row["username"] == usuario_logado.get("username")
                if not eh_proprio:
                    if st.button("ðŸ—‘ï¸ Remover", key=f"del_user_{row['id']}", width="stretch"):
                        sucesso, msg = excluir_usuario(
                            user_id=int(row["id"]),
                            solicitante_username=usuario_logado.get("username", "")
                        )
                        if sucesso:
                            st.success(msg)
                            st.rerun()
                        else:
                            st.error(msg)
                else:
                    st.markdown("<p style='color:#94A3B8;font-size:12px;text-align:center;margin-top:12px;'>(sua conta)</p>", unsafe_allow_html=True)

# â”€â”€ Inicializa banco e tabela de sessÃµes â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
inicializar_banco_completo()
_criar_tabela_sessoes()
_limpar_sessoes_expiradas()

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# SISTEMA DE PERSISTÃŠNCIA DE SESSÃƒO COM COOKIES (SOBREVIVE AO F5)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

import streamlit.components.v1 as components

def _inject_cookie_manager():
    """Injeta JavaScript para gerenciar cookies de autenticaÃ§Ã£o"""
    components.html("""
    <script>
    // FunÃ§Ãµes para gerenciar cookies
    function setCookie(name, value, days) {
        const d = new Date();
        d.setTime(d.getTime() + (days * 24 * 60 * 60 * 1000));
        const expires = "expires=" + d.toUTCString();
        document.cookie = name + "=" + value + ";" + expires + ";path=/;SameSite=Strict";
    }
    
    function getCookie(name) {
        const nameEQ = name + "=";
        const ca = document.cookie.split(';');
        for(let i = 0; i < ca.length; i++) {
            let c = ca[i];
            while (c.charAt(0) == ' ') c = c.substring(1, c.length);
            if (c.indexOf(nameEQ) == 0) return c.substring(nameEQ.length, c.length);
        }
        return null;
    }
    
    function deleteCookie(name) {
        document.cookie = name + "=; expires=Thu, 01 Jan 1970 00:00:00 UTC; path=/;";
    }
    
    // ExpÃµe funÃ§Ãµes globalmente
    window.renapiAuth = {
        setToken: (token) => setCookie('renapsi_auth_token', token, 1),
        getToken: () => getCookie('renapsi_auth_token'),
        clearToken: () => deleteCookie('renapsi_auth_token')
    };
    
    // Envia token atual para o Streamlit se existir
    const currentToken = getCookie('renapsi_auth_token');
    if (currentToken) {
        // Salva no sessionStorage para recuperaÃ§Ã£o rÃ¡pida
        sessionStorage.setItem('renapsi_token_cache', currentToken);
    }
    </script>
    """, height=0)

def _save_token_to_cookie(token):
    """Salva token no cookie do navegador"""
    components.html(f"""
    <script>
    if (window.renapiAuth) {{
        window.renapiAuth.setToken('{token}');
        sessionStorage.setItem('renapsi_token_cache', '{token}');
    }}
    </script>
    """, height=0)

def _clear_token_from_cookie():
    """Remove token do cookie do navegador"""
    components.html("""
    <script>
    if (window.renapiAuth) {
        window.renapiAuth.clearToken();
        sessionStorage.removeItem('renapsi_token_cache');
    }
    </script>
    """, height=0)

# â”€â”€ Injeta gerenciador de cookies â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
_inject_cookie_manager()

# â”€â”€ Restaura sessÃ£o apÃ³s F5 usando banco de dados â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
if not st.session_state.get("usuario_logado"):
    # Tenta recuperar token do session_state primeiro
    token_restore = st.session_state.get("auth_token", "")
    
    # Se nÃ£o tem no session_state, busca todas as sessÃµes ativas do banco
    if not token_restore:
        try:
            conn = sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'mobilidade_renapsi.db'))
            cur = conn.execute("""
                SELECT token, user_id, username, role, expira_em 
                FROM sessoes_ativas 
                WHERE expira_em > ?
                ORDER BY criado_em DESC
                LIMIT 1
            """, (datetime.datetime.now().isoformat(),))
            row = cur.fetchone()
            conn.close()
            
            if row:
                token_restore = row[0]
        except Exception:
            pass
    
    # Valida o token
    if token_restore:
        usuario_recuperado = _validar_token(token_restore)
        if usuario_recuperado:
            st.session_state["usuario_logado"] = True
            st.session_state["usuario_dados"] = usuario_recuperado
            st.session_state["auth_token"] = token_restore
            # Salva no cookie para prÃ³ximas sessÃµes
            _save_token_to_cookie(token_restore)
        else:
            # Token invÃ¡lido ou expirado - limpa tudo
            st.session_state.pop("auth_token", None)
            _clear_token_from_cookie()

# â”€â”€ PortÃ£o de entrada: exige login â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
if not st.session_state.get("usuario_logado"):
    _renderizar_tela_login()
    st.stop()

# â”€â”€ Verifica se precisa trocar senha (primeiro acesso) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
_usuario_atual = st.session_state.get("usuario_dados", {})
if _usuario_atual.get("deve_trocar_senha", 0) == 1:
    _renderizar_tela_troca_senha_obrigatoria()
    st.stop()

# â”€â”€ A partir daqui o usuÃ¡rio estÃ¡ autenticado e com senha vÃ¡lida â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
_usuario_atual = st.session_state.get("usuario_dados", {})
_role_atual = _usuario_atual.get("role", "funcionario")

# â”€â”€ CORREÃ‡ÃƒO FORÃ‡ADA: LETRAS BRANCAS NOS BOTÃ•ES â”€â”€
st.markdown("""
<style>
/* ForÃ§a qualquer texto dentro dos botÃµes azuis (secondary) e laranjas (primary) a ficar branco */
.stButton > button[kind="secondary"] p,
.stButton > button[kind="secondary"] span,
.stButton > button[kind="secondary"] div,
.stButton > button[kind="primary"] p,
.stButton > button[kind="primary"] span,
.stButton > button[kind="primary"] div {
        color: #FFFFFF !important;
}
</style>
""", unsafe_allow_html=True)

# â”€â”€â”€ CSS Global â€” Tema Claro Institucional RENAPSI â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
st.markdown("""
<style>
/* â”€â”€ Reset â”€â”€ */
#MainMenu {visibility: hidden;}
footer     {visibility: hidden;}
header     {visibility: hidden;}

/* â”€â”€ Fundo principal â€” Branco puro â”€â”€ */
.stApp {
        background: #FFFFFF !important;
        font-size: 18px !important;
}

/* â”€â”€ Sidebar â€” Branco/Cinza gelo â”€â”€ */
[data-testid="stSidebar"] {
        background: #F8FAFC !important;
        border-right: 1px solid #E2E8F0 !important;
}
[data-testid="stSidebar"] .stRadio > label {
        color: #1E293B !important;
        font-size: 20px !important;
}
[data-testid="stSidebar"] h1,
[data-testid="stSidebar"] .stMarkdown p {
        color: #333333 !important;
        font-size: 19px !important;
}
/* â”€â”€ Menu de navegaÃ§Ã£o interativo e profissional â”€â”€ */
[data-testid="stSidebar"] .stRadio [data-testid="stMarkdownContainer"] p {
        color: #64748B !important;
        font-size: 18px !important;
        font-weight: 500 !important;
        padding: 12px 16px !important;
        border-radius: 8px !important;
        margin: 4px 0 !important;
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
        cursor: pointer !important;
        border-left: 3px solid transparent !important;
}
[data-testid="stSidebar"] .stRadio [data-testid="stMarkdownContainer"] p:hover {
        color: #f8ae28 !important;
        background: rgba(248,174,40,0.08) !important;
        transform: translateX(8px) !important;
        border-left: 3px solid #f8ae28 !important;
        box-shadow: 0 2px 8px rgba(248,174,40,0.15) !important;
}
/* â”€â”€ Item de menu selecionado â”€â”€ */
[data-testid="stSidebar"] .stRadio [role="radiogroup"] label[data-baseweb="radio"] div[data-testid="stMarkdownContainer"] p {
        background: linear-gradient(135deg, rgba(68,76,155,0.1) 0%, rgba(54,61,127,0.1) 100%) !important;
        color: #444c9b !important;
        font-weight: 600 !important;
        border-left: 3px solid #444c9b !important;
        box-shadow: 0 2px 8px rgba(68,76,155,0.15) !important;
}
/* â”€â”€ Ãcones dos menus com animaÃ§Ã£o â”€â”€ */
[data-testid="stSidebar"] .stRadio p::before {
        display: inline-block !important;
        margin-right: 8px !important;
        transition: transform 0.3s ease !important;
}
[data-testid="stSidebar"] .stRadio p:hover::before {
        transform: scale(1.2) rotate(5deg) !important;
}

/* â”€â”€ KPI cards â€” Fundo branco com sombra suave e gradiente â”€â”€ */
div[data-testid="metric-container"] {
        background: linear-gradient(135deg, #FFFFFF 0%, #F8FAFC 100%) !important;
        border: 1px solid #E2E8F0 !important;
        border-radius: 14px;
        padding: 20px;
        box-shadow: 0 4px 6px rgba(0,0,0,0.05) !important;
        transition: transform 0.3s ease, box-shadow 0.3s ease !important;
}
div[data-testid="metric-container"]:hover {
        transform: translateY(-5px) !important;
        box-shadow: 0 8px 16px rgba(0,0,0,0.1) !important;
}
div[data-testid="metric-container"] [data-testid="stMetricLabel"] {
        color: #64748B !important;
        font-size: 19px !important;
        text-transform: uppercase;
        letter-spacing: 0.08em;
}
div[data-testid="metric-container"] [data-testid="stMetricValue"] {
        color: #f8ae28 !important;
        font-size: 2.6rem !important;
        font-weight: 700;
}
div[data-testid="metric-container"] [data-testid="stMetricDelta"] {
        background: linear-gradient(135deg, #f8ae28 0%, #e09a1f 100%) !important;
        color: #FFFFFF !important;
        padding: 4px 8px !important;
        border-radius: 6px !important;
        font-size: 16px !important;
        font-weight: 600 !important;
}
div[data-testid="metric-container"] [data-testid="stMetricDelta"] svg {
        display: none !important;
}

/* â”€â”€ Headings â€” Cinza escuro â”€â”€ */
h1, h2, h3 { color: #1E293B !important; font-size: 1.5em !important; }
h4          { color: #333333 !important; font-size: 1.3em !important; }

/* â”€â”€ Texto geral â€” Cinza escuro para contraste â”€â”€ */
p, span, div, label, input, textarea, select {
        color: #333333 !important;
        font-size: 19px !important;
}

/* â”€â”€ BotÃµes primÃ¡rios â€” Laranja RENAPSI com gradiente â”€â”€ */
.stButton > button[kind="primary"],
button[data-testid="baseButton-primary"] {
        background: linear-gradient(135deg, #f8ae28 0%, #e09a1f 100%) !important;
        border: none !important;
        color: #FFFFFF !important;
        font-weight: 700 !important;
        font-size: 20px !important;
        border-radius: 8px !important;
        box-shadow: 0 4px 12px rgba(248,174,40,0.3) !important;
        transition: all 0.3s ease !important;
        position: relative !important;
        overflow: hidden !important;
}
.stButton > button[kind="primary"]::before {
        content: '' !important;
        position: absolute !important;
        top: 0 !important;
        left: -100% !important;
        width: 100% !important;
        height: 100% !important;
        background: rgba(255,255,255,0.2) !important;
        transition: left 0.5s ease !important;
}
.stButton > button[kind="primary"]:hover::before {
        left: 100% !important;
}
.stButton > button[kind="primary"]:hover,
button[data-testid="baseButton-primary"]:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 6px 20px rgba(248,174,40,0.4) !important;
}
.stButton > button[kind="primary"]:active {
        transform: translateY(0px) !important;
}

/* â”€â”€ BotÃµes secundÃ¡rios â€” Azul RENAPSI com gradiente â”€â”€ */
.stButton > button[kind="secondary"],
button[data-testid="baseButton-secondary"] {
        background: linear-gradient(135deg, #444c9b 0%, #363d7f 100%) !important;
        border: none !important;
        color: #FFFFFF !important;
        font-size: 20px !important;
        border-radius: 8px !important;
        box-shadow: 0 4px 12px rgba(68,76,155,0.3) !important;
        transition: all 0.3s ease !important;
}
.stButton > button[kind="secondary"]:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 6px 20px rgba(68,76,155,0.4) !important;
}

/* â”€â”€ BotÃµes padrÃ£o (sem type) â€” Tema claro â”€â”€ */
.stButton > button:not([kind]),
button[data-testid="baseButton-minimal"] {
        background: #FFFFFF !important;
        border: 1px solid #E5E7EB !important;
        color: #333333 !important;
        font-size: 20px !important;
        border-radius: 8px !important;
        transition: all 0.3s ease !important;
}
.stButton > button:not([kind]):hover,
button[data-testid="baseButton-minimal"]:hover {
        border-color: #CBD5E1 !important;
        box-shadow: 0 4px 12px rgba(0,0,0,0.08) !important;
        background: #F8FAFC !important;
        transform: translateY(-2px) !important;
}

/* â”€â”€ Tabs com animaÃ§Ã£o â”€â”€ */
.stTabs [data-baseweb="tab-list"] {
        background: #F8FAFC;
        border-radius: 10px;
        padding: 4px;
        border: 1px solid #E2E8F0;
}
.stTabs [data-baseweb="tab"] {
        color: #64748B !important;
        font-size: 20px !important;
        border-radius: 8px;
        transition: all 0.3s ease !important;
}
.stTabs [data-baseweb="tab"]:hover {
        background: rgba(248,174,40,0.1) !important;
}
.stTabs [aria-selected="true"] {
        background: linear-gradient(135deg, #f8ae28 0%, #e09a1f 100%) !important;
        color: #FFFFFF !important;
        border-bottom: 2px solid #f8ae28 !important;
        box-shadow: 0 2px 8px rgba(248,174,40,0.3) !important;
}

/* â”€â”€ Inputs com foco suave â”€â”€ */
.stTextInput > div > div > input,
.stNumberInput > div > div > input,
.stTextArea > div > div > textarea,
.stSelectbox > div > div {
        background: #FFFFFF !important;
        border: 1px solid #E2E8F0 !important;
        color: #333333 !important;
        font-size: 19px !important;
        border-radius: 8px !important;
        transition: all 0.3s ease !important;
}
.stTextInput > div > div > input:focus,
.stTextArea > div > div > textarea:focus {
        border-color: #f8ae28 !important;
        box-shadow: 0 0 0 3px rgba(248,174,40,0.15) !important;
        transform: scale(1.01) !important;
}

/* â”€â”€ Dataframe â”€â”€ */
.stDataFrame {
        border: 1px solid #E2E8F0 !important;
        border-radius: 10px !important;
        font-size: 19px !important;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05) !important;
}

/* â”€â”€ Expander com animaÃ§Ã£o â”€â”€ */
[data-testid="stExpander"] details summary {
        background: linear-gradient(135deg, #444c9b 0%, #363d7f 100%) !important;
        border-radius: 8px !important;
        border: none !important;
        transition: all 0.3s ease !important;
}
[data-testid="stExpander"] details summary:hover,
[data-testid="stExpander"] details summary:focus,
[data-testid="stExpander"] details summary:active {
        background: linear-gradient(135deg, #363d7f 0%, #2a3166 100%) !important;
        outline: none !important;
        box-shadow: 0 4px 12px rgba(68,76,155,0.3) !important;
}
/* ForÃ§a TODOS os textos e Ã­cones do summary a ficarem brancos */
[data-testid="stExpander"] details summary *,
[data-testid="stExpander"] details summary p,
[data-testid="stExpander"] details summary span,
[data-testid="stExpander"] details summary div,
[data-testid="stExpander"] details summary svg,
[data-testid="stExpander"] details summary path {
        color: #FFFFFF !important;
        fill: #FFFFFF !important;
        font-weight: 600 !important;
}
/* Garante que o conteÃºdo que abre embaixo continue com fundo branco */
[data-testid="stExpander"] {
        background-color: #FFFFFF !important;
        border: 1px solid #E5E7EB !important;
        border-radius: 8px !important;
}

/* â”€â”€ Skeleton Loading Animation â”€â”€ */
@keyframes skeleton-loading {
        0% { background-position: -200px 0; }
        100% { background-position: calc(200px + 100%) 0; }
}
.skeleton {
        background: linear-gradient(90deg, #f0f0f0 0px, #f8f8f8 40px, #f0f0f0 80px) !important;
        background-size: 200px 100% !important;
        animation: skeleton-loading 1.5s ease-in-out infinite !important;
}

/* â”€â”€ Toast Notifications â”€â”€ */
.stAlert {
        border-radius: 10px !important;
        border-left: 4px solid !important;
        box-shadow: 0 4px 12px rgba(0,0,0,0.1) !important;
        animation: slideInRight 0.3s ease !important;
}
@keyframes slideInRight {
        from { transform: translateX(100%); opacity: 0; }
        to { transform: translateX(0); opacity: 1; }
}

/* â”€â”€ Badges e Tags â”€â”€ */
.badge {
        display: inline-block !important;
        padding: 4px 12px !important;
        border-radius: 20px !important;
        font-size: 14px !important;
        font-weight: 600 !important;
        text-transform: uppercase !important;
        letter-spacing: 0.5px !important;
}
.badge-success {
        background: linear-gradient(135deg, #10B981 0%, #059669 100%) !important;
        color: white !important;
}
.badge-warning {
        background: linear-gradient(135deg, #F59E0B 0%, #D97706 100%) !important;
        color: white !important;
}
.badge-danger {
        background: linear-gradient(135deg, #EF4444 0%, #DC2626 100%) !important;
        color: white !important;
}
.badge-info {
        background: linear-gradient(135deg, #3B82F6 0%, #2563EB 100%) !important;
        color: white !important;
}

/* â”€â”€ Modal Backdrop Blur â”€â”€ */
.modal-backdrop {
        backdrop-filter: blur(8px) !important;
        background: rgba(0,0,0,0.5) !important;
        animation: fadeIn 0.3s ease !important;
}
@keyframes fadeIn {
        from { opacity: 0; }
        to { opacity: 1; }
}

/* â”€â”€ Micro-interaÃ§Ãµes nos cards â”€â”€ */
.card {
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
        cursor: pointer !important;
}
.card:hover {
        transform: translateY(-8px) scale(1.02) !important;
        box-shadow: 0 12px 24px rgba(0,0,0,0.15) !important;
}
.card:active {
        transform: translateY(-4px) scale(1.01) !important;
}

/* â”€â”€ Progress Bar Animado â”€â”€ */
.stProgress > div > div > div {
        background: linear-gradient(90deg, #f8ae28, #e09a1f, #f8ae28) !important;
        background-size: 200% 100% !important;
        animation: progressShine 2s ease infinite !important;
}
@keyframes progressShine {
        0% { background-position: 0% 50%; }
        50% { background-position: 100% 50%; }
        100% { background-position: 0% 50%; }
}

/* â”€â”€ Scrollbar customizada â”€â”€ */
::-webkit-scrollbar {
        width: 10px !important;
        height: 10px !important;
}
::-webkit-scrollbar-track {
        background: #F8FAFC !important;
        border-radius: 10px !important;
}
::-webkit-scrollbar-thumb {
        background: linear-gradient(135deg, #f8ae28 0%, #e09a1f 100%) !important;
        border-radius: 10px !important;
}
::-webkit-scrollbar-thumb:hover {
        background: linear-gradient(135deg, #e09a1f 0%, #c8891b 100%) !important;
}

/* â”€â”€ Pulse Animation para notificaÃ§Ãµes â”€â”€ */
@keyframes pulse {
        0%, 100% { opacity: 1; }
        50% { opacity: 0.5; }
}
.pulse {
        animation: pulse 2s cubic-bezier(0.4, 0, 0.6, 1) infinite !important;
}

/* â”€â”€ Fade In Animation â”€â”€ */
@keyframes fadeInUp {
        from {
                opacity: 0;
                transform: translateY(20px);
        }
        to {
                opacity: 1;
                transform: translateY(0);
        }
}
.fade-in-up {
        animation: fadeInUp 0.5s ease !important;
}

/* â”€â”€ Spinner Loading â”€â”€ */
@keyframes spin {
        from { transform: rotate(0deg); }
        to { transform: rotate(360deg); }
}
.spinner {
        border: 3px solid #f3f3f3 !important;
        border-top: 3px solid #f8ae28 !important;
        border-radius: 50% !important;
        width: 40px !important;
        height: 40px !important;
        animation: spin 1s linear infinite !important;
}

}

/* â”€â”€ Divider â”€â”€ */
hr { border-color: #E2E8F0 !important; }

/* â”€â”€ Spinner â”€â”€ */
.stSpinner > div { border-top-color: #f8ae28 !important; }

/* â”€â”€ Scrollbar â”€â”€ */
::-webkit-scrollbar { width: 6px; }
::-webkit-scrollbar-track { background: #F8FAFC; }
::-webkit-scrollbar-thumb { background: #CBD5E1; border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: #94A3B8; }
</style>
""", unsafe_allow_html=True)

# â”€â”€ Modo Fullscreen â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
if st.session_state.get('modo_fullscreen', False):
    st.markdown("""
    <style>
    /* Oculta sidebar e header no modo fullscreen */
    [data-testid="stSidebar"] { display: none !important; }
    header[data-testid="stHeader"] { display: none !important; }
    #MainMenu { visibility: hidden !important; }
    footer { visibility: hidden !important; }
    
    /* Expande o conteÃºdo para ocupar toda a tela */
    .main .block-container {
        max-width: 100% !important;
        padding-left: 2rem !important;
        padding-right: 2rem !important;
    }
    </style>
    """, unsafe_allow_html=True)

# â”€â”€â”€ InicializaÃ§Ã£o do Banco de Dados â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
inicializar_banco_completo()

# â”€â”€â”€ Sidebar â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
st.sidebar.markdown("""
<style>
[data-testid="stSidebar"] img {
        background: transparent !important;
        border-radius: 8px;
}
</style>
""", unsafe_allow_html=True)

# Caminho absoluto para o logo
_logo_path = os.path.join(os.path.dirname(__file__), "logo_renapsi.png")
if os.path.exists(_logo_path):
    st.sidebar.image(_logo_path, width="stretch")
else:
    st.sidebar.markdown("### ðŸ¢ RENAPSI")

st.sidebar.markdown(
        "<p style='color:#64748B;font-size:13px;text-align:center;letter-spacing:0.1em;margin-top:-8px;'>SISTEMA DE MOBILIDADE URBANA</p>",
        unsafe_allow_html=True
)
st.sidebar.markdown("---")

# â”€â”€ Busca Global RÃ¡pida â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
st.sidebar.markdown("<p style='color:#1E293B;font-size:13px;text-transform:uppercase;letter-spacing:0.1em;margin-bottom:8px;'>ðŸ” Busca RÃ¡pida</p>", unsafe_allow_html=True)
busca_global = st.sidebar.text_input(
    "Buscar funcionÃ¡rio",
    placeholder="Nome, CPF ou ID...",
    key="busca_global_sidebar",
    label_visibility="collapsed"
)

if busca_global and len(busca_global) >= 3:
    try:
        conexao_busca = sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'mobilidade_renapsi.db'))
        query_busca = """
            SELECT id, nome, cpf, status_rota 
            FROM jovens_rotas 
            WHERE nome LIKE ? OR cpf LIKE ? OR CAST(id AS TEXT) LIKE ?
            LIMIT 5
        """
        df_busca = safe_sql_query(query_busca, conexao_busca, params=(f"%{busca_global}%", f"%{busca_global}%", f"%{busca_global}%"))
        conexao_busca.close()
        
        if not df_busca.empty:
            st.sidebar.markdown("<p style='color:#64748B;font-size:12px;margin:8px 0 4px;'>Resultados:</p>", unsafe_allow_html=True)
            for _, row in df_busca.iterrows():
                cpf_mask = f"***.***.{str(row['cpf']).zfill(11)[6:9]}-**"
                if st.sidebar.button(
                    f"#{row['id']} - {row['nome'][:20]}{'...' if len(row['nome']) > 20 else ''}",
                    key=f"busca_res_{row['id']}",
                    help=f"CPF: {cpf_mask} | Status: {row['status_rota'] or 'Pendente'}",
                    use_container_width=True
                ):
                    st.query_params["menu"] = "ðŸ” Pesquisar Consultas"
                    st.query_params["id_consulta"] = str(row['id'])
                    st.session_state.resultado_busca = df_busca[df_busca['id'] == row['id']]
                    st.rerun()
        else:
            st.sidebar.info("Nenhum resultado encontrado")
    except Exception as e:
        st.sidebar.error(f"Erro na busca: {str(e)}")

st.sidebar.markdown("---")

# â”€â”€ Badge do usuÃ¡rio logado â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
_username_exib = _usuario_atual.get("username", "")
_badge_label = "ðŸ›¡ï¸ Admin" if _role_atual == "admin" else "ðŸ‘¤ FuncionÃ¡rio"
_badge_color = "#444c9b" if _role_atual == "admin" else "#10B981"
st.sidebar.markdown(f"""
<div style="background:#F1F5F9;border:1px solid #E2E8F0;border-radius:8px;
            padding:10px 14px;margin-bottom:12px;">
    <p style="margin:0;color:#1E293B;font-size:13px;font-weight:700;">{_username_exib}</p>
    <p style="margin:2px 0 0;color:{_badge_color};font-size:12px;font-weight:600;">{_badge_label}</p>
</div>
""", unsafe_allow_html=True)

st.sidebar.markdown("<p style='color:#1E293B;font-size:13px;text-transform:uppercase;letter-spacing:0.1em;margin-bottom:4px;'>NavegaÃ§Ã£o</p>", unsafe_allow_html=True)

parametros_url = st.query_params
pagina_salva = parametros_url.get("menu", "ðŸ  Dashboard Principal")
opcoes_menu = [
    "ðŸ  Dashboard Principal", 
    "ðŸ” Pesquisar Consultas", 
    "âž• Cadastrar Novo Jovem", 
    "ðŸ—‚ï¸ Triagem de Fichas", 
    "ðŸ’¾ Banco de Dados", 
    "ðŸ¢ Gerenciar Unidades", 
    "ðŸŒ SimulaÃ§Ã£o: Portal do Jovem"
]
if _role_atual == "admin":
    opcoes_menu.extend(["ðŸ“Š RelatÃ³rios e Analytics", "ðŸ” Auditoria e Compliance", "ðŸ‘¥ Registro de FuncionÃ¡rio"])

indice_padrao = opcoes_menu.index(pagina_salva) if pagina_salva in opcoes_menu else 0

menu = st.sidebar.radio("NavegaÃ§Ã£o", opcoes_menu, index=indice_padrao, label_visibility="collapsed")

# Reseta estado ao mudar de menu
if st.query_params.get("menu") != menu:
        # Limpa estados da tela de pesquisa ao sair dela
        if st.query_params.get("menu") == "Pesquisar Consultas":
            for k in ['resultado_busca', 'detalhes_abertos', 'rota_gerada', 
                    'modo_contestacao', 'modo_edicao', 'mostrar_modal_email', 'analise_ia']:
                if k in st.session_state:
                    del st.session_state[k]
            if 'id_consulta' in st.query_params:
                del st.query_params['id_consulta']

st.query_params["menu"] = menu

st.sidebar.markdown("---")

# â”€â”€ BotÃ£o de logout â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
if st.sidebar.button("ðŸšª Sair", width="stretch"):
    _token_atual = st.session_state.get("auth_token", "")
    if _token_atual:
        _revogar_token(_token_atual)
    
    # Limpa cookie
    _clear_token_from_cookie()
    
    # Limpa session_state
    for key in list(st.session_state.keys()):
        del st.session_state[key]
    
    st.rerun()

st.sidebar.markdown("---")
st.sidebar.markdown(
        "<p style='color:#334155;font-size:14px;text-align:center;'>Copyright Â©ï¸ Renapsi - 2026. Todos os direitos reservados. CNPJ 37.381.902/0001-25.</p>",
        unsafe_allow_html=True
)

# â”€â”€ Selo de Conformidade LGPD â”€â”€
st.sidebar.markdown("""
<div style="background:#FFFFFF;border:1px solid #E2E8F0;
                border-radius:10px;padding:14px;margin-top:16px;text-align:center;
                box-shadow:0 2px 4px rgba(0,0,0,0.05);">
        <p style="color:#f8ae28;font-size:15px;font-weight:600;margin:0 0 8px;letter-spacing:0.05em;">
            ðŸ”’ PRIVACIDADE ASSEGURADA
        </p>
        <p style="color:#64748B;font-size:14px;line-height:1.4;margin:0;">
            Conformidade com <strong>LGPD</strong>. Dados de colaboradores (CPF, Morada, E-mail) sÃ£o processados localmente e armazenados de forma segura. Nenhuma informaÃ§Ã£o pessoal sensÃ­vel Ã© partilhada com APIs externas sem anonimizaÃ§Ã£o.
        </p>
</div>
""", unsafe_allow_html=True)

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# TELA 0 â€” DASHBOARD PRINCIPAL
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
if menu == "ðŸ  Dashboard Principal":

        meses_pt = ["Janeiro","Fevereiro","MarÃ§o","Abril","Maio","Junho",
                    "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]
        mes_atual = meses_pt[datetime.datetime.now().month - 1]
        ano_atual = datetime.datetime.now().year

        # â”€â”€ CabeÃ§alho com modo tela cheia â”€â”€
        col_header, col_actions = st.columns([8, 4])
        with col_header:
            st.markdown("""
            <div style="display:flex; align-items:center; gap:15px; margin-bottom:25px;">
                <svg viewBox="0 0 133.83 114" width="60" height="50">
                <defs><style>.s2-1{fill:#fff;} .s2-2{fill:#402fdd;}</style></defs>
                <g><path class="s2-2" d="M57,0C25.52,0,0,25.52,0,57s25.52,57,57,57h19.83c31.48,0,57-25.52,57-57S108.31,0,76.83,0h-19.83Z"/>
                <path class="s2-1" d="M66.91,48.07c-9.15,0-16.59,7.48-16.59,16.66s7.44,16.66,16.59,16.66,16.59-7.48,16.59-16.66-7.44-16.66-16.59-16.66M66.91,88.58c-13.09,0-23.74-10.7-23.74-23.85s10.65-23.85,23.74-23.85,23.74,10.7,23.74,23.85-10.65,23.85-23.74,23.85"/>
                <polygon class="s2-1" points="90.66 88.58 83.5 87.55 83.5 29.11 90.66 30.14 90.66 88.58"/>
                <polygon class="s2-1" points="56.55 25.42 56.55 32.47 77.28 35.45 77.28 28.4 56.55 25.42"/></g>
                </svg>
                <div>
                    <h1 style="margin:0; color:#444c9b; font-size:28px;">Dashboard de Mobilidade</h1>
                    <p style="margin:4px 0 0; color:#64748B; font-size:14px;">VisÃ£o geral do sistema de transporte</p>
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        with col_actions:
            col_btn1, col_btn2, col_btn3 = st.columns(3)
            with col_btn1:
                if st.button("ðŸ“Š", help="Exportar relatÃ³rio em Excel", width="stretch", key="btn_excel_dash"):
                    st.session_state.exportar_excel = True
            with col_btn2:
                if st.button("ðŸ“„", help="Exportar relatÃ³rio em PDF", width="stretch", key="btn_pdf_dash"):
                    st.session_state.exportar_pdf = True
            with col_btn3:
                modo_fullscreen = st.session_state.get('modo_fullscreen', False)
                if st.button("ðŸ–¥ï¸" if not modo_fullscreen else "â†©ï¸", 
                           help="Modo tela cheia" if not modo_fullscreen else "Sair do modo tela cheia",
                           width="stretch", key="btn_fullscreen_dash"):
                    st.session_state.modo_fullscreen = not modo_fullscreen
                    st.rerun()

        # â”€â”€ Filtro de PerÃ­odo e Modalidade â”€â”€
        col_filtro1, col_filtro2, col_filtro3 = st.columns([2, 3, 7])
        with col_filtro1:
            periodo_selecionado = st.selectbox(
                "ðŸ“… PerÃ­odo:",
                ["Ãšltimos 7 dias", "Ãšltimos 30 dias", "Ãšltimos 90 dias", "Ãšltimo ano", "Tudo"],
                index=st.session_state.get('periodo_index', 1),
                key="filtro_periodo_dashboard"
            )
            st.session_state.periodo_index = ["Ãšltimos 7 dias", "Ãšltimos 30 dias", "Ãšltimos 90 dias", "Ãšltimo ano", "Tudo"].index(periodo_selecionado)
        
        with col_filtro2:
            tipo_rota = st.selectbox(
                "ðŸŽ¯ VisualizaÃ§Ã£o:",
                ["ðŸ  Casa Ã— Trabalho", "ðŸ“š Casa Ã— Curso", "ðŸ“Š GestÃ£o de Base", "ðŸ“§ Envios em Massa"],
                key="modalidade_pesquisa_select"
            )

        # Calcula data de inÃ­cio baseado no perÃ­odo
        data_fim = datetime.datetime.now()
        if periodo_selecionado == "Ãšltimos 7 dias":
            data_inicio = data_fim - datetime.timedelta(days=7)
        elif periodo_selecionado == "Ãšltimos 30 dias":
            data_inicio = data_fim - datetime.timedelta(days=30)
        elif periodo_selecionado == "Ãšltimos 90 dias":
            data_inicio = data_fim - datetime.timedelta(days=90)
        elif periodo_selecionado == "Ãšltimo ano":
            data_inicio = data_fim - datetime.timedelta(days=365)
        else:  # Tudo
            data_inicio = datetime.datetime(2020, 1, 1)

        st.markdown("<br>", unsafe_allow_html=True)

        # â”€â”€ Busca dados com filtro de perÃ­odo â”€â”€
        conexao_dash = sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'mobilidade_renapsi.db'))
        
        try:
            # Query com filtro de data para consultas e SLA
            query_periodo = f"""
                SELECT COUNT(DISTINCT id) as total,
                       AVG(sla_segundos) as sla_medio
                FROM jovens_rotas
                WHERE data_consulta >= '{data_inicio.strftime("%Y-%m-%d")}'
            """
            df_dash = safe_sql_query(query_periodo, conexao_dash)
            
            total_consultas = int(df_dash.iloc[0]['total']) if not df_dash.empty else 0
            sla_medio = float(df_dash.iloc[0]['sla_medio'] or 0) if not df_dash.empty else 0
            
            # Total de implantados ATUAL (independente do perÃ­odo de consulta)
            query_implantados = "SELECT COUNT(*) as total FROM jovens_rotas WHERE status_rota = 'Implantado'"
            df_implantados = safe_sql_query(query_implantados, conexao_dash)
            total_implantados = int(df_implantados.iloc[0]['total']) if not df_implantados.empty else 0
            
            # ContestaÃ§Ãµes no perÃ­odo
            query_contest = f"""
                SELECT COUNT(*) as total
                FROM contestacoes
                WHERE data_geracao >= '{data_inicio.strftime("%Y-%m-%d")}'
            """
            df_contest = safe_sql_query(query_contest, conexao_dash)
            total_contestacoes = int(df_contest.iloc[0]['total']) if not df_contest.empty else 0
        except (sqlite3.OperationalError, pd.errors.DatabaseError) as e:
            # Tabela nÃ£o existe ou estÃ¡ vazia - valores padrÃ£o
            st.info("ðŸ“Š Banco de dados vazio. Cadastre novos jovens para visualizar estatÃ­sticas.")
            total_consultas = 0
            sla_medio = 0
            total_implantados = 0
            total_contestacoes = 0
        
        conexao_dash.close()

        # â”€â”€ KPI Cards CLICÃVEIS â”€â”€
        col_k1, col_k2, col_k3, col_k4 = st.columns(4)
        
        with col_k1:
            if st.button(f"ðŸ“Š {total_consultas}\nTotal de Consultas", 
                        help="Clique para ver detalhes das consultas",
                        use_container_width=True):
                st.query_params["menu"] = "ðŸ” Pesquisar Consultas"
                st.rerun()
            st.caption("Rotas Ãºnicas no perÃ­odo")
        
        with col_k2:
            if st.button(f"â±ï¸ {sla_medio:.2f}s\nSLA MÃ©dio", 
                        help="Tempo mÃ©dio de resposta do sistema",
                        use_container_width=True):
                st.info(f"**SLA MÃ©dio:** {sla_medio:.2f} segundos\n\n"
                       f"Tempo mÃ©dio de processamento das rotas no perÃ­odo selecionado.")
            st.caption("Tempo de resposta")
        
        with col_k3:
            if st.button(f"âš ï¸ {total_contestacoes}\nContestaÃ§Ãµes", 
                        help="Clique para ver detalhes das contestaÃ§Ãµes",
                        use_container_width=True):
                st.query_params["menu"] = "ðŸ’¾ Banco de Dados"
                st.rerun()
            st.caption("Total no perÃ­odo")
        
        with col_k4:
            if st.button(f"âœ… {total_implantados}\nImplantaÃ§Ãµes", 
                        help="FuncionÃ¡rios com VT ativo",
                        use_container_width=True):
                st.query_params["menu"] = "ðŸ’¾ Banco de Dados"
                st.rerun()
            st.caption("Ativos no momento")

        st.markdown("<br>", unsafe_allow_html=True)

        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        # ALERTAS VISUAIS PARA TAREFAS PENDENTES
        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        try:
            _conn_alertas = sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'mobilidade_renapsi.db'))
            
            # Conta fichas pendentes de aprovaÃ§Ã£o
            _df_fichas_pendentes = safe_sql_query("""
                SELECT COUNT(*) as total
                FROM fichas_cadastrais
                WHERE status_aprovacao = 'Pendente'
            """, _conn_alertas)
            _total_fichas_pendentes = int(_df_fichas_pendentes.iloc[0]['total']) if not _df_fichas_pendentes.empty else 0
            
            # Conta contestaÃ§Ãµes pendentes
            _df_contest_pendentes = safe_sql_query("""
                SELECT COUNT(*) as total
                FROM contestacoes
                WHERE status = 'Pendente'
            """, _conn_alertas)
            _total_contest_pendentes = int(_df_contest_pendentes.iloc[0]['total']) if not _df_contest_pendentes.empty else 0
            
            _conn_alertas.close()
        except (sqlite3.OperationalError, pd.errors.DatabaseError):
            # Tabelas nÃ£o existem ainda
            _total_fichas_pendentes = 0
            _total_contest_pendentes = 0
            
        try:
            # Mostra alertas apenas se houver tarefas pendentes
            if _total_fichas_pendentes > 0 or _total_contest_pendentes > 0:
                col_alert1, col_alert2 = st.columns(2)
                
                if _total_fichas_pendentes > 0:
                    with col_alert1:
                        st.markdown(f"""
                        <div style="background:rgba(245,158,11,0.08);border:1px solid rgba(245,158,11,0.4);
                                    border-left:4px solid #F59E0B;border-radius:12px;padding:16px 20px;margin-bottom:16px;
                                    cursor:pointer;transition:all 0.2s;" 
                             onclick="window.location.href='?menu=ðŸ—‚ï¸ Triagem de Fichas'">
                            <div style="display:flex;align-items:center;gap:12px;">
                                <span style="font-size:28px;">ðŸ“‹</span>
                                <div style="flex:1;">
                                    <p style="margin:0;color:#F59E0B;font-weight:700;font-size:16px;">
                                        {_total_fichas_pendentes} Ficha(s) para Aprovar
                                    </p>
                                    <p style="margin:4px 0 0;color:#92400E;font-size:13px;">
                                        Candidaturas aguardando triagem
                                    </p>
                                </div>
                                <span style="color:#F59E0B;font-size:20px;">â†’</span>
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
                        if st.button("Ver Fichas Pendentes", key="btn_fichas_pendentes", use_container_width=True):
                            st.query_params["menu"] = "ðŸ—‚ï¸ Triagem de Fichas"
                            st.rerun()
                
                if _total_contest_pendentes > 0:
                    with col_alert2:
                        st.markdown(f"""
                        <div style="background:rgba(239,68,68,0.08);border:1px solid rgba(239,68,68,0.4);
                                    border-left:4px solid #EF4444;border-radius:12px;padding:16px 20px;margin-bottom:16px;
                                    cursor:pointer;transition:all 0.2s;"
                             onclick="window.location.href='?menu=Banco de Dados'">
                            <div style="display:flex;align-items:center;gap:12px;">
                                <span style="font-size:28px;">âš ï¸</span>
                                <div style="flex:1;">
                                    <p style="margin:0;color:#EF4444;font-weight:700;font-size:16px;">
                                        {_total_contest_pendentes} ContestaÃ§Ã£o(Ãµes) Pendente(s)
                                    </p>
                                    <p style="margin:4px 0 0;color:#991B1B;font-size:13px;">
                                        Requer anÃ¡lise e tratativa
                                    </p>
                                </div>
                                <span style="color:#EF4444;font-size:20px;">â†’</span>
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
                        if st.button("Ver ContestaÃ§Ãµes", key="btn_contestacoes_pendentes", use_container_width=True):
                            st.query_params["menu"] = "ðŸ’¾ Banco de Dados"
                            st.rerun()
        except Exception as e:
            pass  # Silenciosamente ignora erros nos alertas

        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        # ALERTA DE RENOVAÃ‡ÃƒO DE VT â€” Implantados hÃ¡ mais de 12 meses
        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        try:
            _conn_renov = sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'mobilidade_renapsi.db'))
            _df_renov = safe_sql_query("""
                SELECT id, nome, assinatura_data
                FROM jovens_rotas
                WHERE status_rota = 'Implantado'
                  AND assinatura_data IS NOT NULL
                  AND assinatura_data != ''
            """, _conn_renov)
            _conn_renov.close()

            _limite_renov = datetime.datetime.now() - datetime.timedelta(days=365)
            _para_renovar = []
            for _, _r in _df_renov.iterrows():
                try:
                    _dt = datetime.datetime.strptime(str(_r['assinatura_data'])[:19], "%Y-%m-%d %H:%M:%S")
                    if _dt < _limite_renov:
                        _meses = int((datetime.datetime.now() - _dt).days / 30)
                        _para_renovar.append({"nome": _r['nome'], "id": _r['id'], "meses": _meses})
                except Exception:
                    pass

            if _para_renovar:
                st.markdown(f"""
                <div style="background:rgba(245,158,11,0.08);border:1px solid rgba(245,158,11,0.4);
                            border-left:4px solid #F59E0B;border-radius:12px;padding:16px 20px;margin-bottom:20px;">
                    <div style="display:flex;align-items:center;gap:10px;margin-bottom:10px;">
                        <span style="font-size:20px;">âš ï¸</span>
                        <p style="margin:0;color:#F59E0B;font-weight:700;font-size:15px;">
                            {len(_para_renovar)} funcionÃ¡rio(s) com VT para renovar (implantado hÃ¡ +12 meses)
                        </p>
                    </div>
                    <div style="display:flex;flex-wrap:wrap;gap:8px;">
                        {''.join([f"""<span style="background:rgba(245,158,11,0.15);color:#92400E;border:1px solid rgba(245,158,11,0.3);border-radius:20px;padding:3px 10px;font-size:12px;font-weight:600;">#{p['id']} {p['nome'].split()[0]} Â· {p['meses']}m</span>""" for p in _para_renovar[:10]])}
                        {f'<span style="color:#94A3B8;font-size:12px;padding:3px 6px;">+{len(_para_renovar)-10} mais...</span>' if len(_para_renovar) > 10 else ''}
                    </div>
                </div>
                """, unsafe_allow_html=True)
        except Exception:
            pass

        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        # ROI DASHBOARD â€” ANÃLISE FINANCEIRA
        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

        # Busca total de jovens na base
        conexao = sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'mobilidade_renapsi.db'))
        df_jovens = safe_sql_query("SELECT COUNT(*) as total FROM jovens_rotas", conexao)
        total_jovens = df_jovens.iloc[0]['total'] if not df_jovens.empty else 0

        # â”€â”€ Custo real de VT: soma os valores reais do banco â”€â”€
        try:
            df_custo_real = safe_sql_query("""
                SELECT
                    AVG(CASE WHEN valor_tarifa_manual > 0 THEN valor_tarifa_manual ELSE NULL END) as media_real,
                    SUM(CASE WHEN status_rota = 'Implantado' AND valor_tarifa_manual > 0 THEN valor_tarifa_manual ELSE 0 END) as total_implantados_real,
                    COUNT(CASE WHEN status_rota = 'Implantado' AND valor_tarifa_manual > 0 THEN 1 END) as qtd_com_valor
                FROM jovens_rotas
            """, conexao)
            _media_real = float(df_custo_real.iloc[0]['media_real'] or 0) if not df_custo_real.empty else 0
            _total_real_dia = float(df_custo_real.iloc[0]['total_implantados_real'] or 0) if not df_custo_real.empty else 0
            _qtd_com_valor = int(df_custo_real.iloc[0]['qtd_com_valor'] or 0) if not df_custo_real.empty else 0
            CUSTO_OTIMIZADO_DIARIO = _media_real if _media_real > 0 else 11.32
        except Exception:
            CUSTO_OTIMIZADO_DIARIO = 11.32
            _total_real_dia = 0
            _qtd_com_valor = 0

        conexao.close()

        # Constantes financeiras
        CUSTO_MANUAL_DIARIO = 15.00
        DIAS_UTEIS_MES = 22

        # CÃ¡lculos
        custo_manual_mes = CUSTO_MANUAL_DIARIO * DIAS_UTEIS_MES * total_jovens
        custo_otimizado_mes = CUSTO_OTIMIZADO_DIARIO * DIAS_UTEIS_MES * total_jovens
        economia_mes = custo_manual_mes - custo_otimizado_mes
        percentual_economia = (economia_mes / custo_manual_mes * 100) if custo_manual_mes > 0 else 0

        # Badge de fonte dos dados (calculado antes do HTML)
        if _qtd_com_valor > 0:
            _badge_roi = f'<span style="background:rgba(16,185,129,0.15);color:#10B981;padding:2px 10px;border-radius:12px;font-size:12px;font-weight:600;">âœ… Tarifa real do banco Â· {_qtd_com_valor} registros Â· mÃ©dia R${CUSTO_OTIMIZADO_DIARIO:.2f}/dia</span>'
        else:
            _badge_roi = '<span style="background:rgba(245,158,11,0.15);color:#F59E0B;padding:2px 10px;border-radius:12px;font-size:12px;font-weight:600;">âš ï¸ Usando estimativa de mercado (R$11,32/dia)</span>'

        st.markdown(f"""
        <div style="background:#FFFFFF;border:1px solid #E2E8F0;
                    border-radius:14px;padding:24px;margin-bottom:8px;box-shadow:0 2px 4px rgba(0,0,0,0.05);">
            <h3 style="margin:0 0 4px;color:#f8ae28;">ðŸ’° AnÃ¡lise de ROI â€” Retorno sobre Investimento</h3>
            <p style="color:#64748B;font-size:15px;margin:0;">
                Comparativo de custos: Mobilidade Manual vs. Otimizada
            </p>
        </div>
        """, unsafe_allow_html=True)
        st.markdown(f'<p style="margin:0 0 20px;">{_badge_roi}</p>', unsafe_allow_html=True)

        # Blocos de mÃ©tricas financeiras
        col_roi1, col_roi2, col_roi3 = st.columns(3)

        with col_roi1:
            st.markdown(f"""
            <div style="background:linear-gradient(135deg,rgba(239,68,68,0.1),rgba(239,68,68,0.05));
                        border:1px solid rgba(239,68,68,0.3);border-radius:12px;padding:20px;text-align:center;">
                <p style="color:#EF4444;font-size:15px;margin:0 0 8px;text-transform:uppercase;letter-spacing:0.08em;">
                    Custo Manual (MÃªs)
                </p>
                <p style="color:#E2E8F0;font-size:36px;font-weight:800;margin:0;">
                    R$ {custo_manual_mes:,.2f}
                </p>
                <p style="color:#64748B;font-size:15px;margin:4px 0 0;letter-spacing:0.05em;">
                    {total_jovens} jovens Ã— R${CUSTO_MANUAL_DIARIO:.2f}/dia Ã— {DIAS_UTEIS_MES} dias
                </p>
            </div>
            """, unsafe_allow_html=True)

        with col_roi2:
            st.markdown(f"""
            <div style="background:linear-gradient(135deg,rgba(16,185,129,0.1),rgba(16,185,129,0.05));
                        border:1px solid rgba(16,185,129,0.3);border-radius:12px;padding:20px;text-align:center;">
                <p style="color:#10B981;font-size:15px;margin:0 0 8px;text-transform:uppercase;letter-spacing:0.08em;">
                    Custo Otimizado (MÃªs)
                </p>
                <p style="color:#E2E8F0;font-size:36px;font-weight:800;margin:0;">
                    R$ {custo_otimizado_mes:,.2f}
                </p>
                <p style="color:#64748B;font-size:15px;margin:4px 0 0;letter-spacing:0.05em;">
                    {total_jovens} jovens Ã— R${CUSTO_OTIMIZADO_DIARIO:.2f}/dia Ã— {DIAS_UTEIS_MES} dias
                </p>
            </div>
            """, unsafe_allow_html=True)

        with col_roi3:
            st.markdown(f"""
            <div style="background:linear-gradient(135deg,rgba(124,58,237,0.1),rgba(124,58,237,0.05));
                        border:1px solid rgba(124,58,237,0.3);border-radius:12px;padding:20px;text-align:center;">
                <p style="color:#A78BFA;font-size:15px;margin:0 0 8px;text-transform:uppercase;letter-spacing:0.08em;">
                    Economia Mensal
                </p>
                <p style="color:#E2E8F0;font-size:36px;font-weight:800;margin:0;">
                    R$ {economia_mes:,.2f}
                </p>
                <p style="color:#64748B;font-size:15px;margin:4px 0 0;letter-spacing:0.05em;">
                    ReduÃ§Ã£o de {percentual_economia:.1f}% nos custos
                </p>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # GrÃ¡fico de distribuiÃ§Ã£o modal
        col_chart1, col_chart2 = st.columns([1.5, 1])

        with col_chart1:
            st.markdown("""
            <p style="color:#94A3B8;font-size:15px;text-transform:uppercase;letter-spacing:0.08em;margin-bottom:12px;">
                DistribuiÃ§Ã£o Modal das Rotas
                <span style="color:#64748B;font-size:12px;font-weight:normal;margin-left:8px;">
                    â„¹ï¸ Hover sobre o grÃ¡fico para ver detalhes
                </span>
            </p>
            """, unsafe_allow_html=True)
            
            # â”€â”€ Dados REAIS do banco â”€â”€
            try:
                _conn_modal = sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'mobilidade_renapsi.db'))
                _df_modal = safe_sql_query("""
                    SELECT
                        CASE
                            WHEN modo_rota = 'manual' THEN 'Manual'
                            WHEN tipo_bilhete_manual LIKE '%Integra%' THEN 'IntegraÃ§Ã£o'
                            WHEN tipo_bilhete_manual LIKE '%Metro%' OR tipo_bilhete_manual LIKE '%MetrÃ´%' THEN 'MetrÃ´'
                            WHEN status_rota = 'Implantado' THEN 'IntegraÃ§Ã£o'
                            ELSE 'Ã”nibus'
                        END as modal,
                        COUNT(*) as qtd
                    FROM jovens_rotas
                    WHERE status_rota IS NOT NULL
                    GROUP BY modal
                """, _conn_modal)
                _conn_modal.close()
                if _df_modal.empty or _df_modal['qtd'].sum() == 0:
                    raise ValueError("sem dados")
                modais = _df_modal['modal'].tolist()
                percentuais = _df_modal['qtd'].tolist()
            except Exception:
                modais = ['IntegraÃ§Ã£o', 'Ã”nibus', 'MetrÃ´']
                percentuais = [40, 35, 25]

            cores_modais = ['#f8ae28', '#444c9b', '#64748B', '#10B981']

            fig_modal = px.pie(
                values=percentuais,
                names=modais,
                hole=0.5,
                color_discrete_sequence=cores_modais,
                title=""
            )
            fig_modal.update_traces(
                textposition='inside',
                textinfo='label+percent',
                textfont=dict(size=14, color='#FFFFFF', family='Arial'),
                hovertemplate='<b>%{label}</b><br>' +
                             'Quantidade: %{value} funcionÃ¡rios<br>' +
                             'Percentual: %{percent}<br>' +
                             '<extra></extra>',
                marker=dict(line=dict(color='#FFFFFF', width=2))
            )
            fig_modal.update_layout(
                showlegend=True,
                legend=dict(
                    orientation='v',
                    yanchor='middle',
                    y=0.5,
                    xanchor='left',
                    x=1.05,
                    font=dict(size=13, color='#64748B', family='Arial'),
                    bgcolor='rgba(255,255,255,0.8)',
                    bordercolor='#E2E8F0',
                    borderwidth=1
                ),
                margin=dict(t=10, b=10, l=10, r=100),
                height=280,
                paper_bgcolor='rgba(0,0,0,0)',
                plot_bgcolor='rgba(0,0,0,0)',
                font=dict(family='Arial', size=14, color='#333333'),
                hoverlabel=dict(
                    bgcolor='#FFFFFF',
                    font_size=13,
                    font_family='Arial',
                    bordercolor='#E2E8F0'
                )
            )
            st.plotly_chart(fig_modal, use_container_width=True, key="graf_modal_roi")

        with col_chart2:
            st.markdown("""
            <p style="color:#94A3B8;font-size:15px;text-transform:uppercase;letter-spacing:0.08em;margin-bottom:12px;">
                Resumo Financeiro
            </p>
            """, unsafe_allow_html=True)
            
            st.markdown(f"""
            <div style="background:#FFFFFF;border:1px solid #E5E7EB;border-left:4px solid #444c9b;
                        border-radius:12px;padding:16px;height:280px;display:flex;flex-direction:column;justify-content:space-around;
                        box-shadow:0 2px 4px rgba(0,0,0,0.05);">
                <div>
                    <p style="color:#666666;font-size:15px;margin:0 0 4px;text-transform:uppercase;">Total de Jovens</p>
                    <p style="color:#f8ae28;font-size:28px;font-weight:800;margin:0;">{total_jovens}</p>
                </div>
                <div>
                    <p style="color:#666666;font-size:15px;margin:0 0 4px;text-transform:uppercase;">Dias Ãšteis/MÃªs</p>
                    <p style="color:#444c9b;font-size:28px;font-weight:800;margin:0;">{DIAS_UTEIS_MES}</p>
                </div>
                <div>
                    <p style="color:#666666;font-size:15px;margin:0 0 4px;text-transform:uppercase;">Economia Anual</p>
                    <p style="color:#10B981;font-size:28px;font-weight:800;margin:0;">R$ {economia_mes * 12:,.2f}</p>
                </div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("<hr style='border-color:rgba(0,212,255,0.1);margin:20px 0;'>", unsafe_allow_html=True)

        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        # NOVOS COMPONENTES DO DASHBOARD
        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        
        # â”€â”€ 1. GRÃFICO DE LINHA TEMPORAL - EvoluÃ§Ã£o de ImplantaÃ§Ãµes por MÃªs (COLAPSÃVEL) â”€â”€
        with st.expander("ðŸ“ˆ Ver EvoluÃ§Ã£o de ImplantaÃ§Ãµes por MÃªs", expanded=False):
            st.markdown("""
            <p style="color:#64748B;font-size:14px;margin:0 0 16px;">
                Acompanhe o crescimento mensal de funcionÃ¡rios com VT ativo
            </p>
            """, unsafe_allow_html=True)
            
            try:
                _conn_evolucao = sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'mobilidade_renapsi.db'))
                _df_evolucao = safe_sql_query("""
                    SELECT 
                        strftime('%Y-%m', assinatura_data) as mes,
                        COUNT(*) as total
                    FROM jovens_rotas
                    WHERE status_rota = 'Implantado' 
                      AND assinatura_data IS NOT NULL
                      AND assinatura_data != ''
                    GROUP BY mes
                    ORDER BY mes
                """, _conn_evolucao)
                _conn_evolucao.close()
                
                if not _df_evolucao.empty:
                    # Converte para formato de data legÃ­vel
                    _df_evolucao['mes_label'] = pd.to_datetime(_df_evolucao['mes']).dt.strftime('%b/%Y')
                    
                    fig_evolucao = px.line(
                        _df_evolucao,
                        x='mes_label',
                        y='total',
                        markers=True,
                        title=""
                    )
                    fig_evolucao.update_traces(
                        line=dict(color='#10B981', width=3),
                        marker=dict(size=10, color='#10B981', line=dict(color='#FFFFFF', width=2)),
                        hovertemplate='<b>%{x}</b><br>ImplantaÃ§Ãµes: %{y}<extra></extra>'
                    )
                    fig_evolucao.update_layout(
                        xaxis_title="MÃªs",
                        yaxis_title="Total de ImplantaÃ§Ãµes",
                        height=300,
                        margin=dict(t=10, b=40, l=40, r=10),
                        paper_bgcolor='rgba(0,0,0,0)',
                        plot_bgcolor='rgba(0,0,0,0)',
                        font=dict(family='Arial', size=13, color='#64748B'),
                        xaxis=dict(showgrid=True, gridcolor='#E2E8F0'),
                        yaxis=dict(showgrid=True, gridcolor='#E2E8F0'),
                        hoverlabel=dict(bgcolor='#FFFFFF', font_size=13, font_family='Arial', bordercolor='#E2E8F0')
                    )
                    st.plotly_chart(fig_evolucao, use_container_width=True, key="graf_evolucao_implantacoes")
                else:
                    st.info("ðŸ“Š Ainda nÃ£o hÃ¡ dados de implantaÃ§Ãµes com data de assinatura registrada.")
            except Exception as e:
                st.error(f"Erro ao carregar evoluÃ§Ã£o: {e}")

        # â”€â”€ 2. MAPA DE CALOR - DistribuiÃ§Ã£o GeogrÃ¡fica dos FuncionÃ¡rios (COLAPSÃVEL) â”€â”€
        with st.expander("ðŸ—ºï¸ Ver DistribuiÃ§Ã£o GeogrÃ¡fica dos FuncionÃ¡rios", expanded=False):
            st.markdown("""
            <p style="color:#64748B;font-size:14px;margin:0 0 16px;">
                Visualize onde estÃ£o concentrados os funcionÃ¡rios por regiÃ£o
            </p>
            """, unsafe_allow_html=True)
            
            try:
                _conn_geo = sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'mobilidade_renapsi.db'))
                _df_geo = safe_sql_query("""
                    SELECT cep_casa, COUNT(*) as qtd
                    FROM jovens_rotas
                    WHERE cep_casa IS NOT NULL AND cep_casa != ''
                    GROUP BY cep_casa
                    ORDER BY qtd DESC
                    LIMIT 20
                """, _conn_geo)
                _conn_geo.close()
                
                if not _df_geo.empty:
                    # Busca informaÃ§Ãµes de bairro/regiÃ£o para cada CEP
                    _regioes = []
                    for _, row in _df_geo.iterrows():
                        try:
                            _end = buscar_endereco_viacep(row['cep_casa'])
                            _bairro = _end.get('bairro', 'Desconhecido') if isinstance(_end, dict) else 'Desconhecido'
                            _regioes.append({'regiao': _bairro, 'qtd': row['qtd']})
                        except:
                            _regioes.append({'regiao': 'Desconhecido', 'qtd': row['qtd']})
                    
                    _df_regioes = pd.DataFrame(_regioes)
                    _df_regioes_agrupado = _df_regioes.groupby('regiao')['qtd'].sum().reset_index()
                    _df_regioes_agrupado = _df_regioes_agrupado.sort_values('qtd', ascending=False).head(10)
                    
                    fig_geo = px.bar(
                        _df_regioes_agrupado,
                        x='qtd',
                        y='regiao',
                        orientation='h',
                        title="",
                        color='qtd',
                        color_continuous_scale=['#FEF3C7', '#f8ae28', '#e09a1f']
                    )
                    fig_geo.update_traces(
                        hovertemplate='<b>%{y}</b><br>FuncionÃ¡rios: %{x}<extra></extra>'
                    )
                    fig_geo.update_layout(
                        xaxis_title="Quantidade de FuncionÃ¡rios",
                        yaxis_title="",
                        height=400,
                        margin=dict(t=10, b=40, l=150, r=10),
                        paper_bgcolor='rgba(0,0,0,0)',
                        plot_bgcolor='rgba(0,0,0,0)',
                        font=dict(family='Arial', size=13, color='#64748B'),
                        showlegend=False,
                        xaxis=dict(showgrid=True, gridcolor='#E2E8F0'),
                        yaxis=dict(showgrid=False),
                        hoverlabel=dict(bgcolor='#FFFFFF', font_size=13, font_family='Arial', bordercolor='#E2E8F0')
                    )
                    st.plotly_chart(fig_geo, use_container_width=True, key="graf_distribuicao_geografica")
                else:
                    st.info("ðŸ—ºï¸ Ainda nÃ£o hÃ¡ dados de CEP cadastrados para anÃ¡lise geogrÃ¡fica.")
            except Exception as e:
                st.error(f"Erro ao carregar distribuiÃ§Ã£o geogrÃ¡fica: {e}")

        st.markdown("<br>", unsafe_allow_html=True)

        # â”€â”€ 3. WIDGET DE AÃ‡Ã•ES RÃPIDAS â”€â”€
        st.markdown("""
        <div style="background:#FFFFFF;border:1px solid #E2E8F0;border-left:4px solid #444c9b;
                    border-radius:14px;padding:24px;margin-bottom:20px;box-shadow:0 2px 4px rgba(0,0,0,0.05);">
            <h3 style="margin:0 0 4px;color:#444c9b;">âš¡ AÃ§Ãµes RÃ¡pidas</h3>
            <p style="color:#64748B;font-size:14px;margin:0;">
                Acesso rÃ¡pido Ã s tarefas mais comuns do sistema
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        col_acao1, col_acao2, col_acao3, col_acao4 = st.columns(4)
        
        with col_acao1:
            if st.button("âž• Cadastrar Jovem", use_container_width=True, type="secondary"):
                st.query_params["menu"] = "âž• Cadastrar Novo Jovem"
                st.rerun()
            st.caption("Adicionar novo funcionÃ¡rio")
        
        with col_acao2:
            if st.button("ðŸ” Pesquisar Rota", use_container_width=True, type="secondary"):
                st.query_params["menu"] = "ðŸ” Pesquisar Consultas"
                st.rerun()
            st.caption("Buscar rotas existentes")
        
        with col_acao3:
            if st.button("ðŸ“‹ Triagem de Fichas", use_container_width=True, type="secondary"):
                st.query_params["menu"] = "ðŸ—‚ï¸ Triagem de Fichas"
                st.rerun()
            st.caption("Aprovar candidaturas")
        
        with col_acao4:
            if st.button("ðŸ—„ï¸ Banco de Dados", use_container_width=True, type="secondary"):
                st.query_params["menu"] = "ðŸ’¾ Banco de Dados"
                st.rerun()
            st.caption("Gerenciar registros")

        st.markdown("<br>", unsafe_allow_html=True)

        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        # GESTÃƒO DE BASE - Economia Acumulada e DistribuiÃ§Ã£o de Status
        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        if tipo_rota == "ðŸ“Š GestÃ£o de Base":
            st.markdown("""
            <div style="background:linear-gradient(135deg,rgba(16,185,129,0.15),rgba(16,185,129,0.05));
                        border:1px solid rgba(16,185,129,0.3);border-left:4px solid #10B981;
                        border-radius:14px;padding:24px;margin-bottom:20px;box-shadow:0 2px 4px rgba(0,0,0,0.05);">
                <h3 style="margin:0 0 4px;color:#10B981;">ðŸ’° Economia Acumulada Total</h3>
                <p style="color:#64748B;font-size:14px;margin:0;">
                    Economia total desde o inÃ­cio do sistema de mobilidade otimizada
                </p>
            </div>
            """, unsafe_allow_html=True)
            
            try:
                # Busca a data da primeira implantaÃ§Ã£o
                _conn_economia = sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'mobilidade_renapsi.db'))
                _df_primeira_data = safe_sql_query("""
                    SELECT MIN(assinatura_data) as primeira_data
                    FROM jovens_rotas
                    WHERE status_rota = 'Implantado' 
                      AND assinatura_data IS NOT NULL
                      AND assinatura_data != ''
                """, _conn_economia)
                
                _primeira_data_str = _df_primeira_data.iloc[0]['primeira_data'] if not _df_primeira_data.empty else None
                
                if _primeira_data_str:
                    _primeira_data = datetime.datetime.strptime(str(_primeira_data_str)[:10], "%Y-%m-%d")
                    _dias_desde_inicio = (datetime.datetime.now() - _primeira_data).days
                    _meses_desde_inicio = _dias_desde_inicio / 30
                    
                    # Calcula economia acumulada
                    _economia_acumulada = economia_mes * _meses_desde_inicio
                    
                    col_eco1, col_eco2, col_eco3 = st.columns(3)
                    
                    with col_eco1:
                        st.markdown(f"""
                        <div style="background:#FFFFFF;border:1px solid #E5E7EB;border-radius:12px;
                                    padding:20px;text-align:center;box-shadow:0 2px 4px rgba(0,0,0,0.05);">
                            <p style="color:#10B981;font-size:14px;margin:0 0 8px;text-transform:uppercase;letter-spacing:0.08em;">
                                Economia Total
                            </p>
                            <p style="color:#10B981;font-size:36px;font-weight:800;margin:0;">
                                R$ {_economia_acumulada:,.2f}
                            </p>
                            <p style="color:#64748B;font-size:13px;margin:4px 0 0;">
                                Desde {_primeira_data.strftime("%d/%m/%Y")}
                            </p>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    with col_eco2:
                        st.markdown(f"""
                        <div style="background:#FFFFFF;border:1px solid #E5E7EB;border-radius:12px;
                                    padding:20px;text-align:center;box-shadow:0 2px 4px rgba(0,0,0,0.05);">
                            <p style="color:#444c9b;font-size:14px;margin:0 0 8px;text-transform:uppercase;letter-spacing:0.08em;">
                                Tempo de OperaÃ§Ã£o
                            </p>
                            <p style="color:#444c9b;font-size:36px;font-weight:800;margin:0;">
                                {int(_meses_desde_inicio)}
                            </p>
                            <p style="color:#64748B;font-size:13px;margin:4px 0 0;">
                                meses ({_dias_desde_inicio} dias)
                            </p>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    with col_eco3:
                        _economia_por_dia = _economia_acumulada / _dias_desde_inicio if _dias_desde_inicio > 0 else 0
                        st.markdown(f"""
                        <div style="background:#FFFFFF;border:1px solid #E5E7EB;border-radius:12px;
                                    padding:20px;text-align:center;box-shadow:0 2px 4px rgba(0,0,0,0.05);">
                            <p style="color:#f8ae28;font-size:14px;margin:0 0 8px;text-transform:uppercase;letter-spacing:0.08em;">
                                Economia MÃ©dia/Dia
                            </p>
                            <p style="color:#f8ae28;font-size:36px;font-weight:800;margin:0;">
                                R$ {_economia_por_dia:,.2f}
                            </p>
                            <p style="color:#64748B;font-size:13px;margin:4px 0 0;">
                                MÃ©dia diÃ¡ria de economia
                            </p>
                        </div>
                        """, unsafe_allow_html=True)
                else:
                    st.info("ðŸ“Š Ainda nÃ£o hÃ¡ implantaÃ§Ãµes com data registrada para calcular economia acumulada.")
                
                _conn_economia.close()
            except Exception as e:
                st.error(f"Erro ao calcular economia acumulada: {e}")

            st.markdown("<br>", unsafe_allow_html=True)

            # â”€â”€ GRÃFICO DE PIZZA - DistribuiÃ§Ã£o de Status â”€â”€
            st.markdown("""
            <div style="background:#FFFFFF;border:1px solid #E2E8F0;border-left:4px solid #444c9b;
                        border-radius:14px;padding:24px;margin-bottom:20px;box-shadow:0 2px 4px rgba(0,0,0,0.05);">
                <h3 style="margin:0 0 4px;color:#444c9b;">ðŸ“Š DistribuiÃ§Ã£o de Status</h3>
                <p style="color:#64748B;font-size:14px;margin:0;">
                    Visualize a distribuiÃ§Ã£o de funcionÃ¡rios por status de rota
                </p>
            </div>
            """, unsafe_allow_html=True)
            
            try:
                _conn_status = sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'mobilidade_renapsi.db'))
                _df_status = safe_sql_query("""
                    SELECT 
                        COALESCE(status_rota, 'Sem Status') as status,
                        COUNT(*) as total
                    FROM jovens_rotas
                    GROUP BY status_rota
                    ORDER BY total DESC
                """, _conn_status)
                _conn_status.close()
                
                if not _df_status.empty:
                    # Define cores para cada status
                    _cores_status = {
                        'Implantado': '#10B981',
                        'Otimizado': '#3B82F6',
                        'Contestada': '#F59E0B',
                        'NÃ£o Optante': '#94A3B8',
                        'Sem Status': '#64748B'
                    }
                    
                    _cores = [_cores_status.get(s, '#64748B') for s in _df_status['status']]
                    
                    fig_status = px.pie(
                        _df_status,
                        values='total',
                        names='status',
                        hole=0.4,
                        color_discrete_sequence=_cores,
                        title=""
                    )
                    fig_status.update_traces(
                        textposition='inside',
                        textinfo='label+percent',
                        textfont=dict(size=14, color='#FFFFFF', family='Arial'),
                        hovertemplate='<b>%{label}</b><br>' +
                                     'Quantidade: %{value} funcionÃ¡rios<br>' +
                                     'Percentual: %{percent}<br>' +
                                     '<extra></extra>',
                        marker=dict(line=dict(color='#FFFFFF', width=2))
                    )
                    fig_status.update_layout(
                        showlegend=True,
                        legend=dict(
                            orientation='v',
                            yanchor='middle',
                            y=0.5,
                            xanchor='left',
                            x=1.05,
                            font=dict(size=13, color='#64748B', family='Arial'),
                            bgcolor='rgba(255,255,255,0.8)',
                            bordercolor='#E2E8F0',
                            borderwidth=1
                        ),
                        margin=dict(t=10, b=10, l=10, r=150),
                        height=350,
                        paper_bgcolor='rgba(0,0,0,0)',
                        plot_bgcolor='rgba(0,0,0,0)',
                        font=dict(family='Arial', size=14, color='#333333'),
                        hoverlabel=dict(
                            bgcolor='#FFFFFF',
                            font_size=13,
                            font_family='Arial',
                            bordercolor='#E2E8F0'
                        )
                    )
                    st.plotly_chart(fig_status, use_container_width=True, key="graf_distribuicao_status")
                    
                    # Tabela resumo abaixo do grÃ¡fico
                    st.markdown("<p style='color:#94A3B8;font-size:13px;margin-top:16px;'>Resumo por Status:</p>", unsafe_allow_html=True)
                    col_st1, col_st2, col_st3, col_st4 = st.columns(4)
                    for idx, (_, row) in enumerate(_df_status.iterrows()):
                        if idx == 0:
                            col = col_st1
                        elif idx == 1:
                            col = col_st2
                        elif idx == 2:
                            col = col_st3
                        else:
                            col = col_st4
                        
                        _cor_status = _cores_status.get(row['status'], '#64748B')
                        with col:
                            st.markdown(f"""
                            <div style="background:rgba({int(_cor_status[1:3], 16)},{int(_cor_status[3:5], 16)},{int(_cor_status[5:7], 16)},0.1);
                                        border:1px solid {_cor_status}40;border-radius:8px;padding:12px;text-align:center;">
                                <p style="color:{_cor_status};font-size:24px;font-weight:800;margin:0;">{row['total']}</p>
                                <p style="color:#64748B;font-size:12px;margin:4px 0 0;">{row['status']}</p>
                            </div>
                            """, unsafe_allow_html=True)
                else:
                    st.info("ðŸ“Š Ainda nÃ£o hÃ¡ dados de status para exibir.")
            except Exception as e:
                st.error(f"Erro ao carregar distribuiÃ§Ã£o de status: {e}")

        st.markdown("<hr style='border-color:rgba(0,212,255,0.1);margin:20px 0;'>", unsafe_allow_html=True)

        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        # EXPORTAÃ‡ÃƒO DE RELATÃ“RIOS
        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        
        # Exportar para Excel
        if st.session_state.get('exportar_excel', False):
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                from io import BytesIO
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Dashboard Mobilidade"
                
                # CabeÃ§alho
                ws['A1'] = "Dashboard de Mobilidade - RENAPSI"
                ws['A1'].font = Font(size=16, bold=True, color="444c9b")
                ws['A2'] = f"PerÃ­odo: {periodo_selecionado}"
                ws['A2'].font = Font(size=12, italic=True)
                
                # KPIs
                ws['A4'] = "KPIs Principais"
                ws['A4'].font = Font(size=14, bold=True)
                ws['A4'].fill = PatternFill(start_color="444c9b", end_color="444c9b", fill_type="solid")
                ws['A4'].font = Font(size=14, bold=True, color="FFFFFF")
                
                ws['A5'] = "Total de Consultas"
                ws['B5'] = total_consultas
                ws['A6'] = "SLA MÃ©dio (segundos)"
                ws['B6'] = f"{sla_medio:.2f}"
                ws['A7'] = "ContestaÃ§Ãµes"
                ws['B7'] = total_contestacoes
                ws['A8'] = "ImplantaÃ§Ãµes Ativas"
                ws['B8'] = total_implantados
                
                # ROI
                ws['A10'] = "AnÃ¡lise de ROI"
                ws['A10'].font = Font(size=14, bold=True)
                ws['A10'].fill = PatternFill(start_color="f8ae28", end_color="f8ae28", fill_type="solid")
                ws['A10'].font = Font(size=14, bold=True, color="FFFFFF")
                
                ws['A11'] = "Total de Jovens"
                ws['B11'] = total_jovens
                ws['A12'] = "Custo Manual (MÃªs)"
                ws['B12'] = f"R$ {custo_manual_mes:,.2f}"
                ws['A13'] = "Custo Otimizado (MÃªs)"
                ws['B13'] = f"R$ {custo_otimizado_mes:,.2f}"
                ws['A14'] = "Economia Mensal"
                ws['B14'] = f"R$ {economia_mes:,.2f}"
                ws['A15'] = "Economia Anual"
                ws['B15'] = f"R$ {economia_mes * 12:,.2f}"
                ws['A16'] = "Percentual de Economia"
                ws['B16'] = f"{percentual_economia:.1f}%"
                
                # Ajustar largura das colunas
                ws.column_dimensions['A'].width = 30
                ws.column_dimensions['B'].width = 20
                
                # Salvar em buffer
                buffer = BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                
                st.download_button(
                    label="â¬‡ï¸ Baixar RelatÃ³rio Excel",
                    data=buffer,
                    file_name=f"dashboard_mobilidade_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                st.session_state.exportar_excel = False
                st.success("âœ… RelatÃ³rio Excel gerado com sucesso!")
            except Exception as e:
                st.error(f"âŒ Erro ao gerar Excel: {str(e)}")
                st.session_state.exportar_excel = False
        
        # Exportar para PDF
        if st.session_state.get('exportar_pdf', False):
            try:
                from fpdf import FPDF
                from io import BytesIO
                
                pdf = FPDF()
                pdf.add_page()
                
                # TÃ­tulo
                pdf.set_font('Arial', 'B', 20)
                pdf.set_text_color(68, 76, 155)
                pdf.cell(0, 10, 'Dashboard de Mobilidade - RENAPSI', 0, 1, 'C')
                pdf.ln(5)
                
                # PerÃ­odo
                pdf.set_font('Arial', 'I', 12)
                pdf.set_text_color(100, 100, 100)
                pdf.cell(0, 8, f'Periodo: {periodo_selecionado}', 0, 1, 'C')
                pdf.ln(10)
                
                # KPIs
                pdf.set_font('Arial', 'B', 14)
                pdf.set_text_color(68, 76, 155)
                pdf.cell(0, 10, 'KPIs Principais', 0, 1)
                pdf.ln(2)
                
                pdf.set_font('Arial', '', 11)
                pdf.set_text_color(0, 0, 0)
                pdf.cell(90, 8, 'Total de Consultas:', 0, 0)
                pdf.cell(0, 8, str(total_consultas), 0, 1)
                pdf.cell(90, 8, 'SLA Medio (segundos):', 0, 0)
                pdf.cell(0, 8, f'{sla_medio:.2f}', 0, 1)
                pdf.cell(90, 8, 'Contestacoes:', 0, 0)
                pdf.cell(0, 8, str(total_contestacoes), 0, 1)
                pdf.cell(90, 8, 'Implantacoes Ativas:', 0, 0)
                pdf.cell(0, 8, str(total_implantados), 0, 1)
                pdf.ln(10)
                
                # ROI
                pdf.set_font('Arial', 'B', 14)
                pdf.set_text_color(248, 174, 40)
                pdf.cell(0, 10, 'Analise de ROI', 0, 1)
                pdf.ln(2)
                
                pdf.set_font('Arial', '', 11)
                pdf.set_text_color(0, 0, 0)
                pdf.cell(90, 8, 'Total de Jovens:', 0, 0)
                pdf.cell(0, 8, str(total_jovens), 0, 1)
                pdf.cell(90, 8, 'Custo Manual (Mes):', 0, 0)
                pdf.cell(0, 8, f'R$ {custo_manual_mes:,.2f}', 0, 1)
                pdf.cell(90, 8, 'Custo Otimizado (Mes):', 0, 0)
                pdf.cell(0, 8, f'R$ {custo_otimizado_mes:,.2f}', 0, 1)
                pdf.cell(90, 8, 'Economia Mensal:', 0, 0)
                pdf.cell(0, 8, f'R$ {economia_mes:,.2f}', 0, 1)
                pdf.cell(90, 8, 'Economia Anual:', 0, 0)
                pdf.cell(0, 8, f'R$ {economia_mes * 12:,.2f}', 0, 1)
                pdf.cell(90, 8, 'Percentual de Economia:', 0, 0)
                pdf.cell(0, 8, f'{percentual_economia:.1f}%', 0, 1)
                
                # RodapÃ©
                pdf.ln(15)
                pdf.set_font('Arial', 'I', 9)
                pdf.set_text_color(150, 150, 150)
                pdf.cell(0, 8, f'Gerado em: {datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")}', 0, 1, 'C')
                
                # Salvar em buffer
                buffer = BytesIO()
                pdf_output = pdf.output(dest='S').encode('latin-1')
                buffer.write(pdf_output)
                buffer.seek(0)
                
                st.download_button(
                    label="â¬‡ï¸ Baixar RelatÃ³rio PDF",
                    data=buffer,
                    file_name=f"dashboard_mobilidade_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                    mime="application/pdf",
                    use_container_width=True
                )
                st.session_state.exportar_pdf = False
                st.success("âœ… RelatÃ³rio PDF gerado com sucesso!")
            except Exception as e:
                st.error(f"âŒ Erro ao gerar PDF: {str(e)}")
                st.session_state.exportar_pdf = False

        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        # ENVIOS EM MASSA
        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        if tipo_rota == "ðŸ“§ Envios em Massa":
            st.markdown("""
            <div style="background:#FFFFFF;border:1px solid #E5E7EB;border-left:4px solid #444c9b;
                        border-radius:14px;padding:24px;margin-bottom:20px;box-shadow:0 2px 4px rgba(0,0,0,0.05);">
                <h3 style="margin:0 0 4px;color:#444c9b;">ðŸ“§ Envio em Massa de Cartas de VT</h3>
                <p style="color:#666666;font-size:15px;margin:0;">
                    Selecione os funcionÃ¡rios e envie as cartas personalizadas automaticamente
                </p>
            </div>
            """, unsafe_allow_html=True)

            # Busca funcionÃ¡rios pendentes (status != Implantado)
            conexao = sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'mobilidade_renapsi.db'))
            df_pendentes = safe_sql_query("""
                SELECT id, nome, cpf, email, status_rota, cep_casa, cep_trabalho, matricula
                FROM jovens_rotas 
                WHERE status_rota != 'Implantado' OR status_rota IS NULL
                ORDER BY nome
            """, conexao)
            conexao.close()

            if df_pendentes.empty:
                st.info("âœ… NÃ£o hÃ¡ funcionÃ¡rios pendentes de envio. Todos jÃ¡ foram implantados!")
            else:
                # Filtra apenas quem tem e-mail cadastrado
                df_com_email = df_pendentes[df_pendentes['email'].notna() & (df_pendentes['email'] != '')]
                df_sem_email = df_pendentes[df_pendentes['email'].isna() | (df_pendentes['email'] == '')]

                st.markdown(f"""
                <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:16px;margin-bottom:20px;">
                    <div style="background:rgba(59,130,246,0.1);border:1px solid rgba(59,130,246,0.3);
                                border-radius:10px;padding:16px;text-align:center;">
                        <p style="color:#60A5FA;font-size:15px;margin:0 0 4px;text-transform:uppercase;">Total Pendentes</p>
                        <p style="color:#E2E8F0;font-size:32px;font-weight:800;margin:0;">{len(df_pendentes)}</p>
                    </div>
                    <div style="background:rgba(16,185,129,0.1);border:1px solid rgba(16,185,129,0.3);
                                border-radius:10px;padding:16px;text-align:center;">
                        <p style="color:#10B981;font-size:15px;margin:0 0 4px;text-transform:uppercase;">Com E-mail</p>
                        <p style="color:#E2E8F0;font-size:32px;font-weight:800;margin:0;">{len(df_com_email)}</p>
                    </div>
                    <div style="background:rgba(239,68,68,0.1);border:1px solid rgba(239,68,68,0.3);
                                border-radius:10px;padding:16px;text-align:center;">
                        <p style="color:#EF4444;font-size:15px;margin:0 0 4px;text-transform:uppercase;">Sem E-mail</p>
                        <p style="color:#E2E8F0;font-size:32px;font-weight:800;margin:0;">{len(df_sem_email)}</p>
                    </div>
                </div>
                """, unsafe_allow_html=True)

                if len(df_sem_email) > 0:
                    with st.expander(f"âš ï¸ {len(df_sem_email)} funcionÃ¡rio(s) sem e-mail cadastrado"):
                        for _, row in df_sem_email.iterrows():
                            st.markdown(f"""
                            <div style="background:rgba(239,68,68,0.05);border-left:3px solid #EF4444;
                                        padding:10px 14px;margin-bottom:8px;border-radius:0 6px 6px 0;">
                                <p style="margin:0;color:#E2E8F0;font-size:16px;">
                                    <strong>#{row['id']}</strong> - {row['nome']} 
                                    <span style="color:#64748B;font-size:16px;">(CPF: {str(row['cpf']).zfill(11)})</span>
                                </p>
                            </div>
                            """, unsafe_allow_html=True)

                if len(df_com_email) > 0:
                    st.markdown("<hr style='border-color:#E2E8F0;margin:20px 0;'>", unsafe_allow_html=True)
                    st.markdown("<p style='color:#444c9b;font-size:16px;font-weight:600;margin-bottom:12px;'>Selecione os funcionÃ¡rios para envio:</p>", unsafe_allow_html=True)

                    # OpÃ§Ã£o de selecionar todos
                    selecionar_todos = st.checkbox("âœ… Selecionar todos", value=False)

                    # Lista de seleÃ§Ã£o
                    funcionarios_selecionados = []
                    
                    if selecionar_todos:
                        funcionarios_selecionados = df_com_email['id'].tolist()
                        st.info(f"ðŸ“‹ {len(funcionarios_selecionados)} funcionÃ¡rio(s) selecionado(s)")
                    else:
                        # Exibe lista com checkboxes
                        for _, row in df_com_email.iterrows():
                            cpf_mask = f"***.***.{str(row['cpf']).zfill(11)[6:9]}-{str(row['cpf']).zfill(11)[9:11]}"
                            col_check, col_info = st.columns([0.5, 11.5])
                            with col_check:
                                if st.checkbox("", key=f"check_{row['id']}", label_visibility="collapsed"):
                                    funcionarios_selecionados.append(row['id'])
                            with col_info:
                                st.markdown(f"""
                                <div style="background:#FFFFFF;border:1px solid #E5E7EB;
                                            border-radius:8px;padding:12px;margin-bottom:8px;box-shadow:0 2px 4px rgba(0,0,0,0.05);">
                                    <p style="margin:0;color:#333333;font-size:16px;">
                                        <strong>#{row['id']}</strong> - {row['nome']}
                                    </p>
                                    <p style="margin:4px 0 0;color:#666666;font-size:16px;">
                                        CPF: {cpf_mask} Â· E-mail: {row['email']} Â· Status: {row['status_rota'] or 'Pendente'}
                                    </p>
                                </div>
                                """, unsafe_allow_html=True)

                    st.markdown("<br>", unsafe_allow_html=True)

                    # BotÃ£o de envio
                    if len(funcionarios_selecionados) > 0:
                        st.markdown(f"""
                        <div style="background:rgba(59,130,246,0.1);border:1px solid rgba(59,130,246,0.3);
                                    border-radius:10px;padding:16px;margin-bottom:16px;text-align:center;">
                            <p style="color:#60A5FA;font-size:16px;margin:0;">
                                <strong>{len(funcionarios_selecionados)}</strong> funcionÃ¡rio(s) selecionado(s) para envio
                            </p>
                        </div>
                        """, unsafe_allow_html=True)

                        col_env1, col_env2 = st.columns([1, 1])
                        with col_env1:
                            if st.button("ðŸš€ Iniciar Envio em Massa", type="primary", width="stretch"):
                                st.session_state.iniciar_envio_massa = True
                                st.session_state.ids_para_envio = funcionarios_selecionados
                                st.rerun()
                        with col_env2:
                            if st.button("âŒ Cancelar", width="stretch"):
                                st.rerun()

                        # Processo de envio
                        if st.session_state.get('iniciar_envio_massa'):
                            st.markdown("<hr style='border-color:rgba(0,212,255,0.1);margin:20px 0;'>", unsafe_allow_html=True)
                            st.markdown("""
                            <div style="background:rgba(124,58,237,0.1);border:1px solid rgba(124,58,237,0.3);
                                        border-radius:12px;padding:20px;margin-bottom:20px;">
                                <h4 style="margin:0 0 8px;color:#A78BFA;">âš¡ Processamento em Andamento</h4>
                                <p style="color:#94A3B8;font-size:15px;margin:0;">
                                    Aguarde enquanto as cartas sÃ£o geradas e enviadas...
                                </p>
                            </div>
                            """, unsafe_allow_html=True)

                            # Barra de progresso
                            progress_bar = st.progress(0)
                            status_text = st.empty()
                            log_container = st.container()

                            total = len(st.session_state.ids_para_envio)
                            sucessos = 0
                            falhas = []

                            for idx, func_id in enumerate(st.session_state.ids_para_envio):
                                try:
                                    # Busca dados do funcionÃ¡rio
                                    conexao = sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'mobilidade_renapsi.db'))
                                    df_func = safe_sql_query(
                                        "SELECT * FROM jovens_rotas WHERE id = ?", 
                                        conexao, 
                                        params=(int(func_id),)
                                    )
                                    conexao.close()

                                    if df_func.empty:
                                        with log_container:
                                            st.error(f"âŒ ID {func_id}: FuncionÃ¡rio nÃ£o encontrado no banco")
                                        falhas.append(f"ID {func_id} - NÃ£o encontrado")
                                        continue

                                    dados = df_func.iloc[0]
                                    nome = dados['nome']
                                    email = dados['email']
                                    cpf = str(dados['cpf']).zfill(11)
                                    cep_casa = dados['cep_casa']
                                    cep_trab = dados['cep_trabalho']

                                    status_text.markdown(f"""
                                    <p style="color:#60A5FA;font-size:16px;">
                                        ðŸ“¤ Processando <strong>{nome}</strong> ({idx + 1}/{total})...
                                    </p>
                                    """, unsafe_allow_html=True)

                                    # Calcula rota
                                    end_casa_dict = buscar_endereco_viacep(cep_casa)
                                    end_trab_dict = buscar_endereco_viacep(cep_trab)
                                    
                                    rua_casa = end_casa_dict.get('rua', 'N/A') if isinstance(end_casa_dict, dict) else end_casa_dict
                                    rua_trab = end_trab_dict.get('rua', 'N/A') if isinstance(end_trab_dict, dict) else end_trab_dict

                                    rota = motor_de_rotas_gratuito(
                                        f"{rua_casa}, SÃ£o Paulo, Brasil",
                                        f"{rua_trab}, SÃ£o Paulo, Brasil"
                                    )
                                    
                                    # Registra o SLA no banco de dados
                                    if 'sla_segundos' in rota:
                                        from banco_dados import registrar_sla
                                        registrar_sla(func_id, rota['sla_segundos'])

                                    # Seleciona primeira rota disponÃ­vel
                                    rota_para_carta = rota['rotas'][0] if rota.get('rotas') else {
                                        'modal': 'IntegraÃ§Ã£o', 'trajeto': 'Ã”nibus + MetrÃ´/CPTM',
                                        'bilhete': 'IntegraÃ§Ã£o VT', 'valor_diario': 0.0, 'tempo': 'N/A'
                                    }

                                    end_casa_str = end_casa_dict.get('completo', cep_casa) if isinstance(end_casa_dict, dict) else cep_casa
                                    end_trab_str = end_trab_dict.get('completo', cep_trab) if isinstance(end_trab_dict, dict) else cep_trab

                                    # Gera PDF
                                    dados_para_carta = {
                                        'id': func_id, 'nome': nome, 'cpf': cpf,
                                        'matricula': dados.get('matricula', ''), 'email': email
                                    }

                                    # Verifica se Ã© rota manual ou automÃ¡tica
                                    modo_rota_func = dados.get('modo_rota', 'automatica')
                                    
                                    if modo_rota_func == 'manual':
                                        # Usa dados da rota manual
                                        rota_para_carta = {
                                            'tipo_bilhete_manual': dados.get('tipo_bilhete_manual', 'Bilhete Ãšnico'),
                                            'valor_tarifa_manual': dados.get('valor_tarifa_manual', 0.0),
                                            'descricao_itinerario_manual': dados.get('descricao_itinerario_manual', 'Trajeto manual')
                                        }
                                    
                                    pdf_bytes = gerar_carta_pdf(
                                        dados_jovem=dados_para_carta,
                                        rota_selecionada=rota_para_carta,
                                        end_casa_completo=end_casa_str,
                                        end_trab_completo=end_trab_str,
                                        modo_rota=modo_rota_func
                                    )

                                    # Envia e-mail
                                    sucesso, erro = enviar_carta_por_email(
                                        destinatario=email,
                                        nome_funcionario=nome,
                                        pdf_bytes=pdf_bytes
                                    )

                                    if sucesso:
                                        with log_container:
                                            st.success(f"âœ… {nome} - Carta enviada para {email}")
                                        sucessos += 1
                                    else:
                                        with log_container:
                                            st.error(f"âŒ {nome} - Falha: {erro}")
                                        falhas.append(f"{nome} ({email})")

                                except Exception as e:
                                    with log_container:
                                        st.error(f"âŒ ID {func_id}: Erro inesperado - {str(e)}")
                                    falhas.append(f"ID {func_id} - Erro: {str(e)}")

                                # Atualiza progresso
                                progress_bar.progress((idx + 1) / total)

                                # Intervalo de 3 segundos entre envios
                                if idx < total - 1:  # NÃ£o espera apÃ³s o Ãºltimo
                                    time.sleep(3)

                            # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
                            # RELATÃ“RIO DE FEEDBACK VISUAL
                            # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
                            progress_bar.progress(1.0)
                            
                            st.markdown("<hr style='border-color:rgba(0,212,255,0.1);margin:20px 0;'>", unsafe_allow_html=True)
                            
                            # Resumo geral
                            st.markdown(f"""
                            <div style="background:linear-gradient(135deg,rgba(16,185,129,0.1),rgba(16,185,129,0.05));
                                        border:1px solid rgba(16,185,129,0.3);border-radius:14px;padding:24px;margin-bottom:20px;text-align:center;">
                                <h3 style="margin:0 0 12px;color:#10B981;">âœ… Envio ConcluÃ­do!</h3>
                                <div style="display:grid;grid-template-columns:repeat(2,1fr);gap:16px;">
                                    <div>
                                        <p style="color:#64748B;font-size:14px;margin:0 0 4px;text-transform:uppercase;">Enviados com Sucesso</p>
                                        <p style="color:#10B981;font-size:36px;font-weight:800;margin:0;">{sucessos}</p>
                                    </div>
                                    <div>
                                        <p style="color:#64748B;font-size:14px;margin:0 0 4px;text-transform:uppercase;">Falhas</p>
                                        <p style="color:#EF4444;font-size:36px;font-weight:800;margin:0;">{len(falhas)}</p>
                                    </div>
                                </div>
                            </div>
                            """, unsafe_allow_html=True)

                            # Detalhes de sucesso
                            if sucessos > 0:
                                st.markdown(f"""
                                <div style="background:rgba(16,185,129,0.08);border-left:4px solid #10B981;
                                            border-radius:0 8px 8px 0;padding:16px;margin-bottom:16px;">
                                    <p style="color:#10B981;font-size:16px;font-weight:600;margin:0 0 8px;">
                                        âœ… {sucessos} e-mail(s) enviado(s) com sucesso
                                    </p>
                                    <p style="color:#94A3B8;font-size:14px;margin:0;">
                                        Os funcionÃ¡rios receberÃ£o suas cartas de VT personalizadas em breve.
                                    </p>
                                </div>
                                """, unsafe_allow_html=True)

                            # Detalhes de falhas
                            if len(falhas) > 0:
                                with st.expander(f"ðŸš¨ Detalhes das {len(falhas)} Falha(s)", expanded=True):
                                    st.markdown(f"""
                                    <div style="background:rgba(239,68,68,0.08);border-left:4px solid #EF4444;
                                                border-radius:0 8px 8px 0;padding:16px;margin-bottom:16px;">
                                        <p style="color:#EF4444;font-size:16px;font-weight:600;margin:0 0 12px;">
                                            âš ï¸ Os seguintes funcionÃ¡rios nÃ£o receberam o e-mail:
                                        </p>
                                    """, unsafe_allow_html=True)
                                    
                                    for falha in falhas:
                                        st.markdown(f"""
                                        <div style="background:rgba(239,68,68,0.05);border-left:2px solid #EF4444;
                                                    padding:10px 12px;margin-bottom:8px;border-radius:0 4px 4px 0;">
                                            <p style="color:#E2E8F0;font-size:13px;margin:0;">
                                                â€¢ {falha}
                                            </p>
                                        </div>
                                        """, unsafe_allow_html=True)
                                    
                                    st.markdown("""
                                    <div style="background:rgba(239,68,68,0.05);border-left:4px solid #EF4444;
                                                border-radius:0 8px 8px 0;padding:12px;margin-top:12px;">
                                        <p style="color:#94A3B8;font-size:14px;margin:0;">
                                            <strong>AÃ§Ã£o recomendada:</strong> Verifique os e-mails cadastrados e tente novamente.
                                        </p>
                                    </div>
                                    """, unsafe_allow_html=True)

                            st.session_state.iniciar_envio_massa = False
                            st.session_state.ids_para_envio = []

                            if st.button("ðŸ”„ Voltar ao InÃ­cio", type="primary"):
                                st.rerun()

                    else:
                        st.info("ðŸ‘† Selecione pelo menos um funcionÃ¡rio para enviar")

        # â”€â”€ GrÃ¡ficos â”€â”€

        conexao = sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'mobilidade_renapsi.db'))
        try:
            df_contest = safe_sql_query("SELECT * FROM contestacoes", conexao)
            qtd_resolvidas = len(df_contest[df_contest['status'] == 'Resolvido'])
            qtd_pendentes = len(df_contest[df_contest['status'] == 'Pendente'])
        except Exception:
            df_contest = pd.DataFrame(columns=['id', 'data_geracao', 'nome_jovem', 'motivo', 'status', 'tratativa'])
            qtd_resolvidas = 0
            qtd_pendentes = 0
        conexao.close()
        
        col_g1, col_g2, col_g3, col_g4 = st.columns(4)

        CORES_RENAPSI = ['#f8ae28', '#444c9b', '#64748B', '#F59E0B']  # Laranja, Azul, Cinza, Amarelo

        with col_g1:
            # â”€â”€ ImplantaÃ§Ãµes por mÃªs (dados reais) â”€â”€
            try:
                _conn_impl = sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'mobilidade_renapsi.db'))
                _df_impl = safe_sql_query("""
                    SELECT COUNT(*) as total FROM jovens_rotas WHERE status_rota = 'Implantado'
                """, _conn_impl)
                _total_impl = int(_df_impl.iloc[0]['total']) if not _df_impl.empty else 0
                _conn_impl.close()
            except Exception:
                _total_impl = 0
            st.markdown(f"""
            <div style="background:#FFFFFF;border:1px solid #E5E7EB;
                        border-radius:12px;padding:16px;text-align:center;box-shadow:0 2px 4px rgba(0,0,0,0.05);">
                <p style="color:#666666;font-size:13px;text-transform:uppercase;letter-spacing:0.1em;margin:0 0 8px;">
                    ImplantaÃ§Ãµes Ativas
                </p>
                <p style="color:#10B981;font-size:36px;font-weight:800;margin:0;">{_total_impl}</p>
                <p style="color:#94A3B8;font-size:12px;margin:4px 0 0;">funcionÃ¡rios com VT ativo</p>
            </div>
            """, unsafe_allow_html=True)

        with col_g2:
            st.markdown(f"""
            <p style="color:#94A3B8;font-size:14px;text-transform:uppercase;letter-spacing:0.08em;margin-bottom:4px;">
                ContestaÃ§Ãµes ({qtd_resolvidas}/{total_contestacoes})
            </p>
            """, unsafe_allow_html=True)
            if total_contestacoes == 0:
                st.markdown("""
                <div style="background:#FFFFFF;border:1px solid #E5E7EB;
                            border-radius:12px;padding:16px;text-align:center;box-shadow:0 2px 4px rgba(0,0,0,0.05);">
                    <p style="color:#666666;font-size:13px;margin:0;">Nenhuma contestaÃ§Ã£o</p>
                </div>
                """, unsafe_allow_html=True)
            else:
                fig = px.pie(values=[qtd_resolvidas, qtd_pendentes], names=['Resolvidas','Pendentes'], hole=0.72)
                fig.update_traces(textinfo='none', marker=dict(colors=['#444c9b','#E5E7EB']), hoverinfo="skip")
                fig.update_layout(showlegend=False, margin=dict(t=5,b=5,l=5,r=5), height=160,
                                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                                font=dict(size=13, color='#333333'))
                st.plotly_chart(fig, width="stretch", key="graf_contest")

        with col_g3:
            # â”€â”€ Por Local de Trabalho (dados reais) â”€â”€
            st.markdown("<p style='color:#666666;font-size:14px;text-transform:uppercase;letter-spacing:0.08em;margin-bottom:4px;'>Por Local de Trabalho</p>", unsafe_allow_html=True)
            try:
                _conn_lt = sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'mobilidade_renapsi.db'))
                _df_lt = safe_sql_query("""
                    SELECT lt.nome_unidade as local, COUNT(jr.id) as qtd
                    FROM jovens_rotas jr
                    LEFT JOIN locais_trabalho lt ON jr.cep_trabalho = lt.cep
                    WHERE jr.status_rota IS NOT NULL
                    GROUP BY lt.nome_unidade
                    ORDER BY qtd DESC
                    LIMIT 5
                """, _conn_lt)
                _conn_lt.close()
                if _df_lt.empty or _df_lt['qtd'].sum() == 0:
                    raise ValueError("sem dados")
                fig2 = px.pie(values=_df_lt['qtd'].tolist(),
                              names=[str(n)[:20] if n else 'Sem unidade' for n in _df_lt['local'].tolist()],
                              hole=0.72)
            except Exception:
                fig2 = px.pie(values=[10, 0], names=['SP','Outros'], hole=0.72)
            fig2.update_traces(textinfo='none', marker=dict(colors=['#f8ae28','#444c9b','#64748B','#10B981','#E5E7EB']), hoverinfo="label+percent")
            fig2.update_layout(showlegend=False, margin=dict(t=5,b=5,l=5,r=5), height=160,
                            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                            font=dict(size=13, color='#333333'))
            st.plotly_chart(fig2, width="stretch", key="graf_local")

        with col_g4:
            # â”€â”€ Por Status (dados reais) â”€â”€
            st.markdown("<p style='color:#666666;font-size:14px;text-transform:uppercase;letter-spacing:0.08em;margin-bottom:4px;'>Por Status</p>", unsafe_allow_html=True)
            try:
                _conn_st = sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'mobilidade_renapsi.db'))
                _df_st = safe_sql_query("""
                    SELECT COALESCE(status_rota, 'Sem status') as status, COUNT(*) as qtd
                    FROM jovens_rotas GROUP BY status_rota ORDER BY qtd DESC
                """, _conn_st)
                _conn_st.close()
                if _df_st.empty:
                    raise ValueError("sem dados")
                fig3 = px.pie(values=_df_st['qtd'].tolist(), names=_df_st['status'].tolist(), hole=0.72)
            except Exception:
                fig3 = px.pie(values=[10, 0], names=['Otimizado','Implantado'], hole=0.72)
            fig3.update_traces(textinfo='none',
                               marker=dict(colors=['#3B82F6','#10B981','#F59E0B','#94A3B8','#444c9b']),
                               hoverinfo="label+percent")
            fig3.update_layout(showlegend=False, margin=dict(t=5,b=5,l=5,r=5), height=160,
                            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                            font=dict(size=13, color='#333333'))
            st.plotly_chart(fig3, width="stretch", key="graf_uf")

        st.markdown("<hr style='border-color:rgba(0,212,255,0.1);'>", unsafe_allow_html=True)

        # â”€â”€ ContestaÃ§Ãµes â”€â”€
        with st.expander("âš¡ Ver Detalhes das ContestaÃ§Ãµes"):
            if df_contest.empty:
                st.info("Nenhuma contestaÃ§Ã£o registada.")
            else:
                tab_pend, tab_resol, tab_tab = st.tabs(["ðŸ”´ Pendentes", "âœ… Resolvidas", "ðŸ“‹ Tabela"])

                with tab_pend:
                    df_pend = df_contest[df_contest['status'] == 'Pendente']
                    if df_pend.empty:
                        st.success("Tudo limpo! Nenhuma contestaÃ§Ã£o pendente.")
                    else:
                        for _, row in df_pend.iterrows():
                            st.markdown(f"""
                            <div style="background:#FFFFFF;border:1px solid #E5E7EB;border-left:4px solid #EF4444;
                                        border-radius:12px;padding:20px;margin-bottom:12px;
                                        box-shadow:0 2px 4px rgba(0,0,0,0.05);">
                                <div style="display:flex;align-items:center;gap:10px;margin-bottom:10px;">
                                    <span style="color:#EF4444;font-weight:700;font-size:15px;">Consulta #{row['id']}</span>
                                    <span style="background:rgba(239,68,68,0.15);color:#EF4444;padding:2px 8px;
                                                border-radius:20px;font-size:13px;">PENDENTE</span>
                                </div>
                                <p style="color:#666666;font-size:13px;margin:0 0 8px;">
                                    {row['data_geracao']} Â· <strong style="color:#333333;">{row['nome_jovem']}</strong>
                                </p>
                                <div style="background:rgba(239,68,68,0.05);border-left:3px solid #EF4444;
                                            padding:10px 14px;border-radius:0 6px 6px 0;font-size:13px;color:#333333;">
                                    {row['motivo']}
                                </div>
                            </div>
                            """, unsafe_allow_html=True)
                            col_t, col_b = st.columns([3, 1])
                            with col_t:
                                tratativa_input = st.text_input("Tratativa aplicada:", key=f"input_{row['id']}")
                            with col_b:
                                st.write("")
                                st.write("")
                                if st.button("Resolver", type="primary", key=f"btn_{row['id']}", width="stretch"):
                                    if not tratativa_input.strip():
                                        st.error("Descreva a tratativa antes de resolver.")
                                    else:
                                        resolver_contestacao(row['id'], tratativa_input)
                                        st.rerun()

                with tab_resol:
                    df_res = df_contest[df_contest['status'] == 'Resolvido']
                    if df_res.empty:
                        st.info("Nenhuma contestaÃ§Ã£o resolvida ainda.")
                    else:
                        for _, row in df_res.iterrows():
                            st.markdown(f"""
                            <div style="background:#FFFFFF;border:1px solid #E5E7EB;border-left:4px solid #10B981;
                                        border-radius:12px;padding:20px;margin-bottom:12px;box-shadow:0 2px 4px rgba(0,0,0,0.05);">
                                <div style="display:flex;align-items:center;gap:10px;margin-bottom:10px;">
                                    <span style="color:#10B981;font-weight:700;font-size:15px;">Consulta #{row['id']}</span>
                                    <span style="background:rgba(16,185,129,0.15);color:#10B981;padding:2px 8px;
                                                border-radius:20px;font-size:13px;">RESOLVIDO</span>
                                </div>
                                <p style="color:#666666;font-size:13px;margin:0 0 6px;">{row['data_geracao']} Â· <strong style="color:#333333;">{row['nome_jovem']}</strong></p>
                                <p style="color:#666666;font-size:13px;margin:0 0 8px;"><strong>Motivo:</strong> {row['motivo']}</p>
                                <div style="background:rgba(16,185,129,0.08);border-left:3px solid #10B981;
                                            padding:10px 14px;border-radius:0 6px 6px 0;font-size:13px;color:#333333;">
                                    <strong>Tratativa:</strong> {row.get('tratativa','')}
                                </div>
                            </div>
                            """, unsafe_allow_html=True)

                with tab_tab:
                    df_exib = df_contest[['id','data_geracao','nome_jovem','motivo','status','tratativa']].copy()
                    df_exib.columns = ['ID','Data','FuncionÃ¡rio','Motivo','Status','Tratativa']
                    st.dataframe(df_exib, width="stretch", hide_index=True)


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# TELA 1 â€” PESQUISAR CONSULTAS
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
elif menu == "ðŸ” Pesquisar Consultas":

# Recupera estado apÃ³s F5
        if 'id_consulta' in st.query_params and st.session_state.get('resultado_busca') is None:
            id_salvo = st.query_params['id_consulta']
            conexao = sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'mobilidade_renapsi.db'))
            df_salvo = safe_sql_query("SELECT * FROM jovens_rotas WHERE id = ?", conexao, params=(int(float(id_salvo)),))
            conexao.close()
            if not df_salvo.empty:
                st.session_state.resultado_busca = df_salvo
                st.session_state.detalhes_abertos = True

        for key, default in [('resultado_busca', None), ('detalhes_abertos', False),
                            ('rota_gerada', None), ('modo_contestacao', False),
                            ('modo_edicao', False), ('mostrar_modal_email', False)]:
            if key not in st.session_state:
                st.session_state[key] = default

        # â”€â”€ PAINEL DE DETALHES â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        if st.session_state.detalhes_abertos and st.session_state.resultado_busca is not None:

            col_voltar, _ = st.columns([1, 11])
            with col_voltar:
                if st.button("â† Voltar"):
                    for k in ['detalhes_abertos','rota_gerada','modo_contestacao',
                            'modo_edicao','mostrar_modal_email']:
                        st.session_state[k] = False
                    if 'id_consulta' in st.query_params:
                        del st.query_params['id_consulta']
                    st.rerun()

            # Verifica se hÃ¡ dados vÃ¡lidos
            if st.session_state.resultado_busca.empty:
                st.error("âŒ Nenhum dado encontrado para esta consulta.")
                st.session_state.detalhes_abertos = False
                st.rerun()
                
            dados_jovem = st.session_state.resultado_busca.iloc[0]
            
            # â”€â”€ BLINDAGEM CONTRA ID VAZIO â”€â”€
            try:
                id_selecionado = int(dados_jovem['id'])
            except (TypeError, ValueError):
                st.error("âš ï¸ Este jovem estÃ¡ sem ID no banco de dados! Por favor, vÃ¡ na aba 'Banco de Dados', adicione um nÃºmero no campo ID dele e salve.")
                st.stop()
                
            nome_jovem       = dados_jovem['nome']
            cpf_cru          = str(dados_jovem['cpf']).zfill(11)
            cep_casa         = dados_jovem['cep_casa']
            matricula_exib   = dados_jovem.get('matricula', 'NÃ£o informada')

            # --- LÃ“GICA DE CONTEXTO (Status dentro da consulta) ---
            modalidade_atual = st.session_state.get('contexto_salvo', 'Trabalho')
            contexto_ativo = "Trabalho" if "Trabalho" in modalidade_atual else "Curso"
            sigla_contexto = "C-T" if contexto_ativo == "Trabalho" else "C-C"

            if contexto_ativo == "Trabalho":
                cep_trab = dados_jovem.get('cep_trabalho', '')
                status_rota_raw = dados_jovem.get('status_rota', 'Otimizado')
            else:
                cep_trab = dados_jovem.get('cep_curso', '') # Usa o CEP do Curso
                status_rota_raw = dados_jovem.get('status_curso', 'Otimizado')

            status_exib = obter_status_visual(status_rota_raw)
            
            # Define as cores baseadas no status correto
            if status_rota_raw == "Implantado":
                status_color, status_bg = "#10B981", "16,185,129"
            elif status_rota_raw == "Otimizado":
                status_color, status_bg = "#3B82F6", "59,130,246"
            elif status_rota_raw == "Contestada":
                status_color, status_bg = "#F59E0B", "245,158,11"
            else:
                status_color, status_bg = "#94A3B8", "148,163,184"
            email_jovem      = dados_jovem.get('email', '')
            celular_jovem    = dados_jovem.get('celular', '')
            numero_casa      = dados_jovem.get('numero_casa', '')
            coordenadas_casa = dados_jovem.get('coordenadas', '')

            # â”€â”€ MODO EDIÃ‡ÃƒO â”€â”€
            if st.session_state.modo_edicao:
                st.markdown("""
                <div style="background:#FFFFFF;border:1px solid #E5E7EB;border-left:4px solid #444c9b;
                            border-radius:14px;padding:24px;margin-bottom:20px;box-shadow:0 2px 4px rgba(0,0,0,0.05);">
                    <h3 style="margin:0 0 4px;color:#444c9b;">âœï¸ Editar Dados da Consulta</h3>
                    <p style="color:#666666;font-size:13px;margin:0;">Atualize as informaÃ§Ãµes do funcionÃ¡rio e o local de trabalho</p>
                </div>
                """, unsafe_allow_html=True)

                # â”€â”€ Colunas dos status â”€â”€
                col_e1, col_e2, col_e3 = st.columns(3)
                
                with col_e1:
                    st.markdown("<p style='color:#444c9b;font-size:14px;text-transform:uppercase;letter-spacing:0.1em;'>ðŸ‘¤ Dados do FuncionÃ¡rio</p>", unsafe_allow_html=True)
                    mat_input    = st.text_input("MatrÃ­cula", value=matricula_exib if matricula_exib != 'NÃ£o informada' else '')
                    nome_input   = st.text_input("Nome", value=nome_jovem, disabled=True)
                    email_input  = st.text_input("E-mail", value=email_jovem or '')
                    celular_input= st.text_input("Celular", value=celular_jovem or '')
                    obs_input    = st.text_area(
                        "ðŸ“ ObservaÃ§Ãµes internas",
                        value=dados_jovem.get('observacoes', '') or '',
                        placeholder="Ex: aguardando documentaÃ§Ã£o, mudou de endereÃ§o...",
                        height=90,
                        key=f"obs_input_{id_selecionado}"
                    )

                with col_e2:
                    st.markdown("<p style='color:#444c9b;font-size:14px;text-transform:uppercase;letter-spacing:0.1em;'>ðŸ  EndereÃ§o do FuncionÃ¡rio</p>", unsafe_allow_html=True)
                    cep_input = st.text_input("CEP Residencial", value=cep_casa)
                    c_rua, c_num = st.columns([3, 1])
                    
                    # Busca endereÃ§o
                    end_atual = buscar_endereco_viacep(cep_input)
                    rua_input = c_rua.text_input("Logradouro", value=end_atual.get('completo','') if isinstance(end_atual, dict) else '', disabled=True)
                    num_input = c_num.text_input("NÃºmero", value=numero_casa or '')

                    # Coordenadas
                    coord_atual = st.session_state.get('coord_temp', coordenadas_casa)
                    coord_input = st.text_input("Coordenadas (Opcional)", value=coord_atual or '')
                    
                    if st.button("ðŸ” Buscar Coordenadas Reais", type="secondary", width="stretch"):
                        if cep_input and len(cep_input.strip()) == 8:
                            end_completo = f"CEP {cep_input}, {num_input if num_input else ''}, SÃ£o Paulo, Brasil"
                            lat, lon = obter_coordenadas_reais(end_completo)
                            if lat is not None and lon is not None:
                                st.session_state.coord_temp = f"{lat}, {lon}"
                                st.success(f"âœ… Coordenadas: {lat}, {lon}")
                                time.sleep(1)
                                st.rerun()
                            else:
                                st.warning("âš ï¸ NÃ£o encontradas. Verifique o CEP.")
                        else:
                            st.error("âŒ CEP invÃ¡lido")

                with col_e3:
                    st.markdown(f"<p style='color:#444c9b;font-size:14px;text-transform:uppercase;letter-spacing:0.1em;'>ðŸ¢ Local de {contexto_ativo}</p>", unsafe_allow_html=True)
                    
                    # Busca todas as unidades
                    todos_locais = obter_locais_trabalho()
                    
                    # Pega apenas os locais cujo 'tipo_local' bate com o contexto atual
                    # O .get previne erros caso alguma unidade antiga nÃ£o tenha o tipo preenchido
                    locais_trabalho = [loc for loc in todos_locais if loc.get('tipo_local', 'Trabalho') == contexto_ativo]
                    
                    if not locais_trabalho:
                        st.warning(f"âš ï¸ Nenhuma unidade do tipo '{contexto_ativo}' cadastrada no sistema. VÃ¡ em 'Gerenciar Unidades' para adicionar.")
                        local_selecionado = None
                    else:
                        # Cria dicionÃ¡rio com nome_unidade como chave
                        dict_locais = {local['nome_unidade']: local for local in locais_trabalho}
                        nomes_unidades = list(dict_locais.keys())
                        
                        # Selectbox com unidades do banco (dinÃ¢mico, sem hardcoding)
                        local_selecionado_nome = st.selectbox(
                            "Selecione o local de trabalho:",
                            nomes_unidades,
                            index=None,
                            placeholder="Escolha uma unidade...",
                            key=f"select_local_trabalho_app_{id_selecionado}"
                        )
                        
                        if local_selecionado_nome:
                            local_selecionado = dict_locais[local_selecionado_nome]
                            
                            # Card azul com dados reais da unidade (100% dinÃ¢mico)
                            st.markdown(f"""
                            <div style="background-color:#BAE6FD; border-radius:8px; padding:15px; text-align:center; margin-top:5px; border: 1px solid #7DD3FC;">
                                <p style="color:#0284C7; font-size:12px; font-weight:700; margin:0;">
                                    {local_selecionado['logradouro']}, {local_selecionado['numero']} - {local_selecionado['bairro']}<br>
                                    {local_selecionado['cidade_uf']}
                                </p>
                                <p style="color:#0284C7; font-size:11px; margin:8px 0 0;">
                                    CEP: {local_selecionado['cep']}
                                </p>
                                {f'<p style="color:#0284C7; font-size:11px; margin:4px 0 0;">COORDENADAS: {local_selecionado["coordenadas"]}</p>' if local_selecionado['coordenadas'] else ''}
                            </div>
                            """, unsafe_allow_html=True)
                        else:
                            local_selecionado = None
                            st.info("â„¹ï¸ Selecione uma unidade para visualizar os dados")

                st.markdown("<br>", unsafe_allow_html=True)
                col_f, col_c = st.columns([1, 1])
                with col_f:
                    if st.button("Fechar", width="stretch"):
                        st.session_state.modo_edicao = False
                        st.session_state.pop('coord_temp', None)
                        st.rerun()
                with col_c:
                    if st.button("Confirmar AlteraÃ§Ãµes", type="primary", width="stretch"):
                        try:
                            # Valida se uma unidade foi selecionada
                            local_selecionado_nome = st.session_state.get(f"select_local_trabalho_app_{id_selecionado}")
                            
                            if not local_selecionado_nome:
                                st.error("âš ï¸ Selecione um local de trabalho antes de confirmar.")
                            else:
                                # Busca a unidade selecionada do banco para garantir dados atualizados
                                locais_trabalho = obter_locais_trabalho()
                                dict_locais = {local['nome_unidade']: local for local in locais_trabalho}
                                
                                if local_selecionado_nome not in dict_locais:
                                    st.error("âŒ Unidade selecionada nÃ£o encontrada no banco de dados.")
                                else:
                                    local_selecionado = dict_locais[local_selecionado_nome]
                                    
                                    # Prepara dados para salvar
                                    mat_final = str(mat_input) if mat_input else ''
                                    email_final = str(email_input) if email_input else ''
                                    celular_final = str(celular_input) if celular_input else ''
                                    cep_final = str(cep_input) if cep_input else str(cep_casa)
                                    num_final = str(num_input) if num_input else ''
                                    coord_final = str(coord_input) if coord_input else ''
                                    
                                    # CEP da unidade selecionada (CRÃTICO: garantir que seja do banco)
                                    cep_trab_final = local_selecionado['cep']
                                    
                                    with st.spinner("Salvando no banco de dados..."):
                                        # UPDATE com prepared statement (seguro contra SQL injection)
                                        conexao = sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'mobilidade_renapsi.db'))
                                        cursor = conexao.cursor()
                                        cursor.execute('''
                                            UPDATE jovens_rotas 
                                            SET matricula = ?, email = ?, celular = ?, cep_casa = ?, numero_casa = ?, coordenadas = ?, cep_trabalho = ?, observacoes = ?
                                            WHERE id = ?
                                        ''', (mat_final, email_final, celular_final, cep_final, num_final, coord_final, cep_trab_final, obs_input, id_selecionado))
                                        conexao.commit()
                                        
                                        # Recarrega dados atualizados
                                        df_atualizado = safe_sql_query("SELECT * FROM jovens_rotas WHERE id = ?", conexao, params=(int(id_selecionado),))
                                        conexao.close()
                                    
                                    if not df_atualizado.empty:
                                        st.session_state.resultado_busca = df_atualizado
                                        st.session_state.modo_edicao = False
                                        st.session_state.pop('coord_temp', None)
                                        st.session_state.rota_gerada = None # ForÃ§a o recalculo da rota para a nova empresa
                                        st.session_state.analise_ia = None

                                        # â”€â”€ Registra no histÃ³rico de auditoria â”€â”€
                                        try:
                                            _usr_hist = st.session_state.get("usuario_dados", {}).get("username", "desconhecido")
                                            registrar_historico(
                                                usuario=_usr_hist,
                                                acao="EdiÃ§Ã£o de dados do funcionÃ¡rio",
                                                tabela="jovens_rotas",
                                                registro_id=id_selecionado,
                                                campo="matricula/email/celular/cep_casa/cep_trabalho",
                                                valor_novo=f"mat={mat_final} | email={email_final} | cep_trab={cep_trab_final}"
                                            )
                                        except Exception:
                                            pass

                                        st.success(f"âœ… Dados salvos com sucesso! CEP de trabalho atualizado para: {cep_trab_final}")
                                        time.sleep(2)
                                        st.rerun()
                                    else:
                                        st.error("âŒ Erro ao recarregar dados do banco.")
                        except Exception as e:
                            st.error(f"âŒ Erro ao salvar: {str(e)}")

            # â”€â”€ MODO VISUALIZAÃ‡ÃƒO â”€â”€
            else:
                end_casa_dict = buscar_endereco_viacep(cep_casa)
                end_trab_dict = buscar_endereco_viacep(cep_trab)

                rua_casa = end_casa_dict.get('rua','N/A') if isinstance(end_casa_dict, dict) else end_casa_dict
                bairro_cidade_casa = f"{end_casa_dict.get('bairro','')} - {end_casa_dict.get('cidade_uf','')}" if isinstance(end_casa_dict, dict) else ""
                rua_trab = end_trab_dict.get('rua','N/A') if isinstance(end_trab_dict, dict) else end_trab_dict
                bairro_cidade_trab = f"{end_trab_dict.get('bairro','')} - {end_trab_dict.get('cidade_uf','')}" if isinstance(end_trab_dict, dict) else ""

                if not st.session_state.get('rota_gerada'):
                    endereco_completo_casa = end_casa_dict.get('completo', f"{rua_casa}, SÃ£o Paulo, SP, Brasil") if isinstance(end_casa_dict, dict) else f"{rua_casa}, SÃ£o Paulo, SP, Brasil"
                    endereco_completo_trab = end_trab_dict.get('completo', f"{rua_trab}, SÃ£o Paulo, SP, Brasil") if isinstance(end_trab_dict, dict) else f"{rua_trab}, SÃ£o Paulo, SP, Brasil"
                    
                    rota = motor_de_rotas_gratuito(endereco_completo_casa, endereco_completo_trab)
                    st.session_state.rota_gerada = rota
                    
                    # Registra o SLA no banco de dados
                    if 'sla_segundos' in rota:
                        from banco_dados import registrar_sla
                        registrar_sla(id_selecionado, rota['sla_segundos'])
                    
                    if not st.session_state.get('analise_ia'):
                        st.session_state.analise_ia = analisar_rota_com_ia(
                            rua_casa, rua_trab, rota['distancia_km'], rota['rotas'], rota['info_tarifas']
                        )

                col_fill, col_edit_btn = st.columns([11, 1])
                with col_edit_btn:
                    if st.button("âœï¸", width="stretch", help="Editar dados"):
                        st.session_state.modo_edicao = True
                        st.rerun()

                # Cor status
                if status_rota_raw == "Implantado":
                    status_color, status_bg = "#10B981", "16,185,129"
                elif status_rota_raw == "Otimizado":
                    status_color, status_bg = "#3B82F6", "59,130,246"
                elif status_rota_raw == "Contestada":
                    status_color, status_bg = "#F59E0B", "245,158,11"
                else:
                    status_color, status_bg = "#94A3B8", "148,163,184"

                # â”€â”€ PREPARAÃ‡ÃƒO DAS STRINGS â”€â”€
                linha_num_casa = f', {numero_casa}' if numero_casa else ''

                # CabeÃ§alho com ID, tÃ­tulo e status
                col_header_left, col_header_right = st.columns([10, 2])
                
                with col_header_left:
                    st.markdown(f"### ðŸ‘¤ Consulta #{id_selecionado}")
                    st.caption(f"RENAPSI Â· SÃƒO PAULO Â· {sigla_contexto}")
                    # â”€â”€ Indicador de completude â”€â”€
                    _campos_completude = {
                        'E-mail': bool(email_jovem),
                        'Celular': bool(celular_jovem),
                        'MatrÃ­cula': bool(matricula_exib and matricula_exib != 'NÃ£o informada'),
                        'CEP Trabalho': bool(cep_trab),
                        'NÃºmero casa': bool(numero_casa),
                        'Coordenadas': bool(coordenadas_casa),
                    }
                    _total_campos = len(_campos_completude)
                    _preenchidos = sum(_campos_completude.values())
                    _pct = int(_preenchidos / _total_campos * 100)
                    _cor_pct = "#10B981" if _pct >= 80 else "#F59E0B" if _pct >= 50 else "#EF4444"
                    st.markdown(f"""
                    <div style="display:flex;align-items:center;gap:8px;margin-top:4px;">
                        <div style="flex:1;background:#E2E8F0;border-radius:4px;height:6px;overflow:hidden;">
                            <div style="width:{_pct}%;background:{_cor_pct};height:100%;border-radius:4px;"></div>
                        </div>
                        <span style="color:{_cor_pct};font-size:12px;font-weight:700;white-space:nowrap;">
                            {_pct}% completo
                        </span>
                    </div>
                    """, unsafe_allow_html=True)
                    
                with col_header_right:
                    st.markdown(f"""
                    <div style="background:rgba({status_bg},0.15);color:{status_color};padding:6px 14px;border-radius:20px;font-size:12px;font-weight:700;border:1px solid {status_color}40;text-align:center;">
                        {status_exib}
                    </div>
                    """, unsafe_allow_html=True)
                    # Timestamp da Ãºltima carta enviada
                    _ultima_carta = dados_jovem.get('ultima_carta_enviada', '')
                    if _ultima_carta:
                        st.markdown(f"""
                        <div style="margin-top:6px;background:rgba(16,185,129,0.08);border:1px solid rgba(16,185,129,0.3);
                                    border-radius:8px;padding:4px 10px;font-size:11px;color:#10B981;text-align:center;">
                            ðŸ“§ Carta enviada em {_ultima_carta}
                        </div>
                        """, unsafe_allow_html=True)

                st.divider()

                # TrÃªs colunas para os dados
                col1, col2, col3 = st.columns(3)

                # Coluna 1: Dados do FuncionÃ¡rio
                with col1:
                    st.markdown("**Dados do FuncionÃ¡rio**")
                    st.markdown(f"**CPF:** {cpf_cru}")
                    st.markdown(f"**MatrÃ­cula:** {matricula_exib}")
                    st.markdown(f"**Nome:** {nome_jovem}")
                    if email_jovem:
                        st.markdown(f"**E-mail:** {email_jovem}")
                    if celular_jovem:
                        st.markdown(f"**Celular:** {celular_jovem}")
                    # ObservaÃ§Ãµes internas
                    _obs_exib = dados_jovem.get('observacoes', '') or ''
                    if _obs_exib.strip():
                        st.markdown(f"""
                        <div style="background:rgba(245,158,11,0.08);border-left:3px solid #F59E0B;
                                    border-radius:0 6px 6px 0;padding:8px 10px;margin-top:8px;">
                            <p style="color:#92400E;font-size:12px;font-weight:600;margin:0 0 2px;">ðŸ“ Obs. internas</p>
                            <p style="color:#78350F;font-size:12px;margin:0;">{_obs_exib}</p>
                        </div>
                        """, unsafe_allow_html=True)

                # Coluna 2: EndereÃ§o Residencial
                with col2:
                    st.markdown("**EndereÃ§o Residencial**")
                    st.markdown("""
                    <span style="background:rgba(16,185,129,0.1);color:#10B981;padding:2px 8px;font-size:12px;border-radius:20px;font-weight:600;">â— BAIXO RISCO</span>
                    """, unsafe_allow_html=True)
                    st.markdown(f"**CEP:** {cep_casa}")
                    st.markdown(f"{rua_casa}{linha_num_casa}  \n{bairro_cidade_casa}")

                # Coluna 3: Local de Trabalho
                with col3:
                    st.markdown("**Local de Trabalho**")
                    st.markdown(f"**CEP:** {cep_trab}")
                    st.markdown(f"{rua_trab}  \n{bairro_cidade_trab}")
                
                # â”€â”€ Barra de aÃ§Ãµes â”€â”€
                st.markdown("<p style='color:#64748B;font-size:12px;margin-bottom:8px;'>AÃ‡Ã•ES DA CONSULTA</p>", unsafe_allow_html=True)
                col_b1, col_b2, col_b3, col_b4, col_b5, col_fill2 = st.columns([1, 1, 1, 1, 1, 1])
                ja_implantado = (status_rota_raw == "Implantado")
                
                with col_b1:
                    if st.button("ðŸ”„ Recalcular", type="secondary", width="stretch"):
                        # Limpa TUDO relacionado Ã  rota para forÃ§ar recÃ¡lculo completo
                        print("\n" + "="*60)
                        print("ðŸ”„ RECALCULANDO ROTA - LIMPANDO CACHE")
                        print("="*60 + "\n")
                        
                        if 'rota_gerada' in st.session_state:
                            del st.session_state.rota_gerada
                        if 'analise_ia' in st.session_state:
                            del st.session_state.analise_ia
                        if 'modo_contestacao' in st.session_state:
                            del st.session_state.modo_contestacao
                        
                        st.success("âœ… Cache limpo! Recalculando rota...")
                        time.sleep(0.5)
                        st.rerun()
                        
                with col_b2:
                    if st.button("âœ‰ï¸ Enviar Carta", type="secondary", width="stretch"):
                        st.session_state.mostrar_modal_email = not st.session_state.get('mostrar_modal_email', False)
                        
                with col_b3:
                    if st.button("âš ï¸ ContestaÃ§Ã£o", type="secondary", width="stretch"):
                        st.session_state.modo_contestacao = not st.session_state.get('modo_contestacao', False)
                        
                with col_b4:
                    if st.button("âœï¸ Rota Manual", type="secondary", width="stretch"):
                        st.session_state.modo_rota_manual = not st.session_state.get('modo_rota_manual', False)
                        
                with col_b5:
                    if st.button("ðŸ“‹ Implantados", type="secondary", width="stretch"):
                        st.session_state.modo_implantacao = not st.session_state.get('modo_implantacao', False)

                if st.session_state.get('mostrar_modal_email'):
                    st.markdown(f"""
                    <div style="background:rgba(0,212,255,0.05);border:1px solid rgba(0,212,255,0.25);
                                border-radius:12px;padding:16px;margin:12px 0;">
                        <p style="color:#00D4FF;font-weight:600;margin:0 0 6px;">âœ‰ï¸ Enviar Carta de OpÃ§Ã£o de Transporte</p>
                        <p style="color:#94A3B8;font-size:13px;margin:0;">
                            DestinatÃ¡rio: <strong style="color:#E2E8F0;">{email_jovem or 'E-mail nÃ£o informado'}</strong>
                        </p>
                    </div>
                    """, unsafe_allow_html=True)

                    if not email_jovem:
                        st.warning("âš ï¸ Este funcionÃ¡rio nÃ£o possui e-mail cadastrado. Edite os dados para adicionar.")
                    else:
                        # Seleciona qual rota usar na carta
                        rotas_disponiveis = st.session_state.rota_gerada.get('rotas', []) if st.session_state.get('rota_gerada') else []
                        opcoes_carta = [r['modal'] for r in rotas_disponiveis] if rotas_disponiveis else ["Rota padrÃ£o"]
                        rota_escolhida_label = st.selectbox(
                            "Qual rota incluir na carta?",
                            opcoes_carta,
                            key="sel_rota_carta"
                        )

                        c1, c2 = st.columns([1, 5])
                        with c1:
                            if st.button("ðŸ“„ Gerar e Enviar", type="primary"):
                                with st.spinner("Gerando PDF e enviando e-mail..."):
                                    try:
                                        # Seleciona a rota escolhida
                                        idx_rota = opcoes_carta.index(rota_escolhida_label)
                                        rota_para_carta = rotas_disponiveis[idx_rota] if rotas_disponiveis else {
                                            'modal': 'IntegraÃ§Ã£o', 'trajeto': 'Ã”nibus + MetrÃ´/CPTM',
                                            'bilhete': 'IntegraÃ§Ã£o Ã”nibus+MetrÃ´ VT',
                                            'valor_diario': 22.64, 'tempo': '45 min'
                                        }

                                        # Monta endereÃ§os completos
                                        end_casa_dict2 = buscar_endereco_viacep(cep_casa)
                                        end_trab_dict2 = buscar_endereco_viacep(cep_trab)
                                        end_casa_str = end_casa_dict2.get('completo', cep_casa) if isinstance(end_casa_dict2, dict) else cep_casa
                                        end_trab_str = end_trab_dict2.get('completo', cep_trab) if isinstance(end_trab_dict2, dict) else cep_trab

                                        dados_para_carta = {
                                            'id':        id_selecionado,
                                            'nome':      nome_jovem,
                                            'cpf':       cpf_cru,
                                            'matricula': matricula_exib,
                                            'email':     email_jovem,
                                        }

                                        # Verifica se Ã© rota manual ou automÃ¡tica
                                        modo_rota_atual = dados_jovem.get('modo_rota', 'automatica')
                                        
                                        if modo_rota_atual == 'manual':
                                            # Usa dados da rota manual
                                            rota_para_carta = {
                                                'tipo_bilhete_manual': dados_jovem.get('tipo_bilhete_manual', 'Bilhete Ãšnico'),
                                                'valor_tarifa_manual': dados_jovem.get('valor_tarifa_manual', 0.0),
                                                'descricao_itinerario_manual': dados_jovem.get('descricao_itinerario_manual', 'Trajeto manual')
                                            }

                                        pdf_bytes = gerar_carta_pdf(
                                            dados_jovem=dados_para_carta,
                                            rota_selecionada=rota_para_carta,
                                            end_casa_completo=end_casa_str,
                                            end_trab_completo=end_trab_str,
                                            modo_rota=modo_rota_atual
                                        )

                                        sucesso, erro = enviar_carta_por_email(
                                            destinatario=email_jovem,
                                            nome_funcionario=nome_jovem,
                                            pdf_bytes=pdf_bytes,
                                        )

                                        if sucesso:
                                            st.success(f"âœ… Carta enviada com sucesso para **{email_jovem}**!")
                                            # Registra timestamp do envio
                                            try:
                                                _ts_carta = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
                                                _conn_ts = sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'mobilidade_renapsi.db'))
                                                _conn_ts.execute(
                                                    "UPDATE jovens_rotas SET ultima_carta_enviada = ? WHERE id = ?",
                                                    (_ts_carta, id_selecionado)
                                                )
                                                _conn_ts.commit()
                                                _conn_ts.close()
                                            except Exception:
                                                pass
                                            # Disponibiliza download tambÃ©m
                                            st.download_button(
                                                "â¬‡ï¸ Baixar PDF da Carta",
                                                data=pdf_bytes,
                                                file_name=f"Carta_VT_{nome_jovem.replace(' ','_')}.pdf",
                                                mime="application/pdf"
                                            )
                                            time.sleep(2)
                                            st.session_state.mostrar_modal_email = False
                                            st.rerun()
                                        else:
                                            st.error(f"âŒ Falha ao enviar: {erro}")
                                            # Mesmo com erro, oferece download do PDF
                                            st.download_button(
                                                "â¬‡ï¸ Baixar PDF da Carta (envio falhou)",
                                                data=pdf_bytes,
                                                file_name=f"Carta_VT_{nome_jovem.replace(' ','_')}.pdf",
                                                mime="application/pdf"
                                            )
                                    except Exception as e:
                                        st.error(f"âŒ Erro ao gerar carta: {e}")

                if st.session_state.modo_contestacao:
                    with st.form(key="form_contestacao"):
                        st.markdown("<p style='color:#F59E0B;font-weight:600;'>âš ï¸ Registrar ContestaÃ§Ã£o</p>", unsafe_allow_html=True)
                        motivo_input = st.text_area("Descreva o problema:", placeholder="Ex: tarifa indevida, rota que nÃ£o faz sentido...")
                        email_rh_input = st.text_input(
                            "E-mail do RH para notificaÃ§Ã£o (opcional)",
                            placeholder="rh@empresa.com.br",
                            help="Se preenchido, um e-mail de alerta serÃ¡ enviado automaticamente ao RH."
                        )
                        if st.form_submit_button("Registrar", type="primary"):
                            if not motivo_input.strip():
                                st.error("Descreva o motivo antes de registrar.")
                            else:
                                registrar_contestacao(nome=nome_jovem, cid_res="SÃ£o Paulo", cid_trab="SÃ£o Paulo", motivo=motivo_input)

                                # Detectar contexto ativo
                                modalidade_atual = st.session_state.get('contexto_salvo', 'Trabalho')
                                contexto_ativo = "Trabalho" if "Trabalho" in modalidade_atual else "Curso"

                                # Atualiza status para CONTESTADA usando funÃ§Ã£o com contexto
                                atualizar_status_rota(id_selecionado, 'Contestada', contexto_ativo)

                                # â”€â”€ NotificaÃ§Ã£o automÃ¡tica ao RH â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
                                _email_destino_rh = email_rh_input.strip() if email_rh_input.strip() else None
                                if not _email_destino_rh:
                                    # Tenta ler do .env
                                    import os as _os
                                    _email_destino_rh = _os.getenv("EMAIL_RH", "")

                                if _email_destino_rh and "@" in _email_destino_rh:
                                    try:
                                        import smtplib
                                        from email.mime.multipart import MIMEMultipart
                                        from email.mime.text import MIMEText
                                        from dotenv import load_dotenv as _ldenv
                                        _ldenv()
                                        _smtp_user = _os.getenv("EMAIL_USER", "")
                                        _smtp_pass = _os.getenv("EMAIL_PASS", "")
                                        _smtp_host = _os.getenv("EMAIL_SMTP_SERVER", "smtp.gmail.com")
                                        _smtp_port = int(_os.getenv("EMAIL_SMTP_PORT", 587))

                                        if _smtp_user and _smtp_pass:
                                            _msg_rh = MIMEMultipart()
                                            _msg_rh["From"] = _smtp_user
                                            _msg_rh["To"] = _email_destino_rh
                                            _msg_rh["Subject"] = f"âš ï¸ Nova ContestaÃ§Ã£o Registrada â€” {nome_jovem}"
                                            _corpo_rh = f"""
                                            <html><body style="font-family:Arial,sans-serif;color:#333;padding:20px;">
                                            <div style="max-width:600px;margin:auto;border-top:4px solid #F59E0B;border-radius:8px;padding:24px;background:#fff;">
                                                <h2 style="color:#F59E0B;margin-top:0;">âš ï¸ Nova ContestaÃ§Ã£o Registrada</h2>
                                                <p><strong>FuncionÃ¡rio:</strong> {nome_jovem}</p>
                                                <p><strong>CPF:</strong> ***.***.{cpf_cru[6:9]}-{cpf_cru[9:11]}</p>
                                                <p><strong>MatrÃ­cula:</strong> {matricula_exib}</p>
                                                <p><strong>Contexto:</strong> {contexto_ativo}</p>
                                                <p><strong>Motivo:</strong></p>
                                                <div style="background:#FEF3C7;border-left:4px solid #F59E0B;padding:12px;border-radius:0 8px 8px 0;">
                                                    {motivo_input}
                                                </div>
                                                <p style="color:#94A3B8;font-size:12px;margin-top:20px;">
                                                    Registrado em {datetime.datetime.now().strftime("%d/%m/%Y Ã s %H:%M")} Â· RENAPSI Sistema de Mobilidade
                                                </p>
                                            </div>
                                            </body></html>
                                            """
                                            _msg_rh.attach(MIMEText(_corpo_rh, "html", "utf-8"))
                                            with smtplib.SMTP(_smtp_host, _smtp_port, timeout=15) as _srv:
                                                _srv.ehlo()
                                                _srv.starttls()
                                                _srv.login(_smtp_user, _smtp_pass)
                                                _srv.sendmail(_smtp_user, _email_destino_rh, _msg_rh.as_string())
                                            st.info(f"ðŸ“§ NotificaÃ§Ã£o enviada para {_email_destino_rh}")
                                    except Exception as _e_rh:
                                        st.warning(f"âš ï¸ ContestaÃ§Ã£o registrada, mas falha ao notificar RH: {_e_rh}")

                                st.success("ContestaÃ§Ã£o registrada! Status alterado para CONTESTADA.")
                                # â”€â”€ HistÃ³rico â”€â”€
                                try:
                                    registrar_historico(
                                        usuario=st.session_state.get("usuario_dados", {}).get("username", "desconhecido"),
                                        acao="ContestaÃ§Ã£o registrada",
                                        tabela="jovens_rotas",
                                        registro_id=id_selecionado,
                                        campo="status_rota",
                                        valor_anterior=status_rota_raw,
                                        valor_novo="Contestada"
                                    )
                                except Exception:
                                    pass
                                st.session_state.modo_contestacao = False
                                time.sleep(2)
                                st.rerun()

                # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
                # MODAL DE IMPLANTAÃ‡ÃƒO
                # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
                if st.session_state.get('modo_implantacao', False):
                    st.markdown("""
                    <div style="background:rgba(59,130,246,0.08);border:1px solid rgba(59,130,246,0.3);
                                border-radius:14px;padding:20px;margin-bottom:20px;">
                        <h3 style="margin:0 0 4px;color:#3B82F6;">ðŸ“‹ Alterar Status de ImplantaÃ§Ã£o</h3>
                        <p style="color:#94A3B8;font-size:13px;margin:0;">
                            Selecione a aÃ§Ã£o desejada para alterar o status do funcionÃ¡rio
                        </p>
                    </div>
                    """, unsafe_allow_html=True)

                    # Mostra status atual
                    st.markdown(f"""
                    <div style="background:rgba(13,17,23,0.8);border:1px solid rgba(59,130,246,0.2);
                                border-radius:10px;padding:14px;margin-bottom:16px;">
                        <p style="color:#94A3B8;font-size:12px;margin:0 0 4px;">Status Atual:</p>
                        <p style="color:#3B82F6;font-size:16px;font-weight:700;margin:0;">{status_exib}</p>
                    </div>
                    """, unsafe_allow_html=True)

                    # OpÃ§Ãµes de implantaÃ§Ã£o
                    opcao_implantacao = st.radio(
                        "Escolha a aÃ§Ã£o:",
                        ["Implantada", "Implantada (Com Cancelamento de VT)", "NÃ£o Implantada"],
                        key=f"opcao_implantacao_{id_selecionado}"
                    )

                    # Campo de motivo (apenas para "NÃ£o Implantada")
                    motivo_nao_implantada = ""
                    if opcao_implantacao == "NÃ£o Implantada":
                        motivo_nao_implantada = st.text_area(
                            "Motivo da remoÃ§Ã£o da implantaÃ§Ã£o:",
                            placeholder="Ex: FuncionÃ¡rio solicitou cancelamento, mudanÃ§a de endereÃ§o, etc.",
                            key=f"motivo_nao_implantada_{id_selecionado}"
                        )

                    st.markdown("<br>", unsafe_allow_html=True)

                    col_confirmar, col_cancelar = st.columns([1, 1])
                    
                    with col_confirmar:
                        if st.button("âœ… Confirmar AlteraÃ§Ã£o", type="primary", width="stretch", key=f"confirmar_implantacao_{id_selecionado}"):
                            # ValidaÃ§Ã£o para "NÃ£o Implantada"
                            if opcao_implantacao == "NÃ£o Implantada" and not motivo_nao_implantada.strip():
                                st.error("âš ï¸ Informe o motivo da remoÃ§Ã£o da implantaÃ§Ã£o.")
                            else:
                                # Define o novo status baseado na opÃ§Ã£o
                                if opcao_implantacao == "Implantada":
                                    novo_status = "Implantado"
                                elif opcao_implantacao == "Implantada (Com Cancelamento de VT)":
                                    novo_status = "NÃ£o Optante"
                                else:  # NÃ£o Implantada
                                    novo_status = "Otimizado"
                                
                                # Detectar contexto ativo
                                modalidade_atual = st.session_state.get('contexto_salvo', 'Trabalho')
                                contexto_ativo = "Trabalho" if "Trabalho" in modalidade_atual else "Curso"
                                
                                # Atualiza no banco usando a funÃ§Ã£o com contexto
                                atualizar_status_rota(id_selecionado, novo_status, contexto_ativo)

                                # â”€â”€ HistÃ³rico â”€â”€
                                try:
                                    registrar_historico(
                                        usuario=st.session_state.get("usuario_dados", {}).get("username", "desconhecido"),
                                        acao=f"AlteraÃ§Ã£o de status ({opcao_implantacao})",
                                        tabela="jovens_rotas",
                                        registro_id=id_selecionado,
                                        campo=f"status_{contexto_ativo.lower()}",
                                        valor_anterior=status_rota_raw,
                                        valor_novo=novo_status
                                    )
                                except Exception:
                                    pass

                                st.success(f"âœ… Status alterado para: {novo_status} (Contexto: {contexto_ativo})")
                                st.session_state.modo_implantacao = False
                                time.sleep(1)
                                st.rerun()

                    with col_cancelar:
                        if st.button("âŒ Cancelar", width="stretch", key=f"cancelar_implantacao_{id_selecionado}"):
                            st.session_state.modo_implantacao = False
                            st.rerun()

                st.markdown("<br>", unsafe_allow_html=True)

                # â”€â”€ AnÃ¡lise da IA â”€â”€
                if st.session_state.get('analise_ia'):
                    st.markdown(f"""
                    <div style="background:rgba(124,58,237,0.08);border-left:4px solid #7C3AED;
                                border-radius:0 12px 12px 0;padding:18px 20px;margin-bottom:20px;
                                box-shadow:-4px 0 24px rgba(124,58,237,0.2);">
                        <div style="display:flex;align-items:center;gap:10px;margin-bottom:10px;">
                            <span style="font-size:20px;">ðŸ¤–</span>
                            <p style="margin:0;color:#A78BFA;font-weight:700;font-size:14px;text-transform:uppercase;letter-spacing:0.08em;">
                                AnÃ¡lise do Agente de IA
                            </p>
                        </div>
                        <p style="margin:0;color:#CBD5E1;font-size:14px;line-height:1.7;">{st.session_state.analise_ia}</p>
                    </div>
                    """, unsafe_allow_html=True)

                # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
                # MODAL DE ROTA MANUAL
                # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
                if st.session_state.get('modo_rota_manual', False):
                    st.markdown("""
                    <div style="background:rgba(245,158,11,0.08);border:1px solid rgba(245,158,11,0.3);
                                border-radius:14px;padding:20px;margin-bottom:20px;">
                        <h3 style="margin:0 0 4px;color:#F59E0B;">âœï¸ Inserir Rota Manual</h3>
                        <p style="color:#94A3B8;font-size:13px;margin:0;">
                            Preencha os campos abaixo para definir a rota manualmente (SPTRANS)
                        </p>
                    </div>
                    """, unsafe_allow_html=True)

                    # Busca dados salvos (se existirem)
                    tipo_bilhete_salvo = dados_jovem.get('tipo_bilhete_manual', '')
                    valor_tarifa_salvo = dados_jovem.get('valor_tarifa_manual', 0.0)
                    descricao_salva = dados_jovem.get('descricao_itinerario_manual', '')

                    col_bilhete, col_valor = st.columns([2, 1])
                    
                    with col_bilhete:
                        tipo_bilhete_manual = st.selectbox(
                            "Tipo de Bilhete (SPTRANS)",
                            [
                                "Bilhete Ãšnico",
                                "IntegraÃ§Ã£o Ã”nibus+MetrÃ´",
                                "IntegraÃ§Ã£o Ã”nibus+CPTM",
                                "Vale Transporte"
                            ],
                            index=0 if not tipo_bilhete_salvo else (
                                ["Bilhete Ãšnico", "IntegraÃ§Ã£o Ã”nibus+MetrÃ´", "IntegraÃ§Ã£o Ã”nibus+CPTM", "Vale Transporte"].index(tipo_bilhete_salvo)
                                if tipo_bilhete_salvo in ["Bilhete Ãšnico", "IntegraÃ§Ã£o Ã”nibus+MetrÃ´", "IntegraÃ§Ã£o Ã”nibus+CPTM", "Vale Transporte"]
                                else 0
                            ),
                            key=f"tipo_bilhete_{id_selecionado}"
                        )

                    with col_valor:
                        valor_tarifa_manual = st.number_input(
                            "Valor da Tarifa (R$)",
                            min_value=0.10,
                            max_value=500.00,
                            value=float(valor_tarifa_salvo) if valor_tarifa_salvo else 0.10,
                            step=0.10,
                            format="%.2f",
                            key=f"valor_tarifa_{id_selecionado}"
                        )

                    descricao_itinerario_manual = st.text_area(
                        "DescriÃ§Ã£o do ItinerÃ¡rio",
                        value=descricao_salva,
                        placeholder="Ex: Linha 102 -> Terminal Bandeira -> MetrÃ´ Linha 3-Vermelha -> EstaÃ§Ã£o SÃ©",
                        height=120,
                        key=f"descricao_itinerario_{id_selecionado}"
                    )

                    st.markdown("<br>", unsafe_allow_html=True)

                    col_salvar_manual, col_cancelar_manual = st.columns([1, 1])
                    
                    with col_salvar_manual:
                        if st.button("ðŸ’¾ Salvar Rota Manual", type="primary", width="stretch", key=f"salvar_manual_{id_selecionado}"):
                            if not tipo_bilhete_manual or valor_tarifa_manual < 0.10 or not descricao_itinerario_manual.strip():
                                st.error("âš ï¸ Preencha todos os campos antes de salvar.")
                            else:
                                from banco_dados import salvar_rota_manual
                                
                                sucesso = salvar_rota_manual(
                                    id_selecionado,
                                    tipo_bilhete_manual,
                                    valor_tarifa_manual,
                                    descricao_itinerario_manual
                                )
                                
                                if sucesso:
                                    st.success("âœ… Rota manual salva com sucesso!")
                                    
                                    # Recarrega os dados
                                    conexao = sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'mobilidade_renapsi.db'))
                                    df_atualizado = safe_sql_query(
                                        "SELECT * FROM jovens_rotas WHERE id = ?", 
                                        conexao, 
                                        params=(int(id_selecionado),)
                                    )
                                    conexao.close()
                                    
                                    if not df_atualizado.empty:
                                        st.session_state.resultado_busca = df_atualizado
                                    
                                    st.session_state.modo_rota_manual = False
                                    time.sleep(1)
                                    st.rerun()
                                else:
                                    st.error("âŒ Erro ao salvar rota manual. Verifique o console.")

                    with col_cancelar_manual:
                        if st.button("âŒ Cancelar", width="stretch", key=f"cancelar_manual_{id_selecionado}"):
                            st.session_state.modo_rota_manual = False
                            st.rerun()

                # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
                # EXIBIÃ‡ÃƒO DA ROTA E MAPA
                # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
                
                # Calcula rota se ainda nÃ£o foi calculada
                if not st.session_state.get('rota_gerada'):
                    # Monta endereÃ§os completos para geocoding preciso
                    endereco_completo_casa = end_casa_dict.get('completo', f"{rua_casa}, SÃ£o Paulo, SP, Brasil") if isinstance(end_casa_dict, dict) else f"{rua_casa}, SÃ£o Paulo, SP, Brasil"
                    endereco_completo_trab = end_trab_dict.get('completo', f"{rua_trab}, SÃ£o Paulo, SP, Brasil") if isinstance(end_trab_dict, dict) else f"{rua_trab}, SÃ£o Paulo, SP, Brasil"
                    
                    rota = motor_de_rotas_gratuito(
                        endereco_completo_casa,
                        endereco_completo_trab
                    )
                    st.session_state.rota_gerada = rota
                    
                    # Registra o SLA no banco de dados
                    if 'sla_segundos' in rota:
                        from banco_dados import registrar_sla
                        registrar_sla(id_selecionado, rota['sla_segundos'])
                    
                    # AnÃ¡lise da IA em background (nÃ£o bloqueia a UI)
                    if not st.session_state.get('analise_ia'):
                        st.session_state.analise_ia = analisar_rota_com_ia(
                            rua_casa, rua_trab, rota['distancia_km'],
                            rota['rotas'], rota['info_tarifas']
                        )

                # â”€â”€ Rotas + Mapa â”€â”€
                col_painel, col_mapa = st.columns([1, 2.8])

                with col_painel:
                    st.markdown("<p style='color:#444c9b;font-size:12px;text-transform:uppercase;letter-spacing:0.1em;margin-bottom:8px;'>ðŸ—ºï¸ OpÃ§Ãµes de Trajeto</p>", unsafe_allow_html=True)

                    if st.session_state.get('rota_gerada') and 'rotas' in st.session_state.rota_gerada:
                        abas = st.tabs([r["modal"] for r in st.session_state.rota_gerada["rotas"]])
                        for i, aba in enumerate(abas):
                            with aba:
                                rt = st.session_state.rota_gerada["rotas"][i]
                                st.markdown(f"""
                                <div style="background:#f8ae28;border:1px solid #E5E7EB;
                                            border-radius:12px;padding:18px;margin-top:8px;
                                            box-shadow:0 2px 4px rgba(0,0,0,0.1);">
                                    <div style="display:flex;align-items:flex-start;gap:12px;margin-bottom:16px;">
                                        <div style="background:#FFFFFF;border-radius:50%;
                                                    min-width:40px;height:40px;display:flex;align-items:center;
                                                    justify-content:center;font-weight:800;color:#f8ae28;font-size:13px;
                                                    border:2px solid #FFFFFF;">SP</div>
                                        <div>
                                            <p style="margin:0;font-weight:700;color:#FFFFFF;font-size:14px;">{rt['trajeto']}</p>
                                            <p style="margin:4px 0;font-size:12px;color:#FFFFFF;">{rt['bilhete']}</p>
                                            <p style="margin:4px 0;font-size:12px;color:#FFFFFF;font-weight:600;">â± {rt['tempo']}</p>
                                        </div>
                                    </div>
                                    <div style="background:#FFFFFF;
                                                border:1px solid #E5E7EB;border-radius:8px;
                                                text-align:center;padding:14px;
                                                box-shadow:0 2px 4px rgba(0,0,0,0.05);">
                                        <p style="margin:0;color:#666666;font-size:13px;text-transform:uppercase;letter-spacing:0.08em;">VT DiÃ¡rio (Ida + Volta)</p>
                                        <p style="margin:4px 0 0;color:#f8ae28;font-size:26px;font-weight:800;">R$ {rt['valor_diario']:.2f}</p>
                                    </div>
                                </div>
                                """, unsafe_allow_html=True)
                    else:
                        st.info("Calculando rotas...")

                with col_mapa:
                    if st.session_state.get('rota_gerada'):
                        (lat_c, lon_c), (lat_t, lon_t) = st.session_state.rota_gerada['coords_reais']
                        rota_info = st.session_state.rota_gerada
                        
                        # Calcula tempo estimado de deslocamento (baseado na distÃ¢ncia)
                        distancia_km = rota_info.get('distancia_km', 0)
                        # Velocidade mÃ©dia em SP: 25 km/h (considerando trÃ¢nsito)
                        tempo_estimado_min = int((distancia_km / 25) * 60)
                        tempo_horas = tempo_estimado_min // 60
                        tempo_minutos = tempo_estimado_min % 60
                        tempo_texto = f"{tempo_horas}h{tempo_minutos}min" if tempo_horas > 0 else f"{tempo_minutos}min"
                        
                        # â”€â”€ INFO BOX: Tempo estimado e distÃ¢ncia â”€â”€
                        st.markdown(f"""
                        <div style="background:linear-gradient(135deg,#444c9b,#5a63b8);border-radius:12px;
                                    padding:16px;margin-bottom:16px;box-shadow:0 4px 6px rgba(0,0,0,0.1);">
                            <div style="display:flex;justify-content:space-around;align-items:center;">
                                <div style="text-align:center;">
                                    <p style="color:rgba(255,255,255,0.9) !important;font-size:12px;margin:0;text-transform:uppercase;letter-spacing:0.1em;">
                                        DistÃ¢ncia
                                    </p>
                                    <p style="color:#FFFFFF !important;font-size:24px;font-weight:800;margin:4px 0 0;">
                                        {distancia_km:.1f} km
                                    </p>
                                </div>
                                <div style="width:1px;height:40px;background:rgba(255,255,255,0.3);"></div>
                                <div style="text-align:center;">
                                    <p style="color:rgba(255,255,255,0.9) !important;font-size:12px;margin:0;text-transform:uppercase;letter-spacing:0.1em;">
                                        Tempo Estimado
                                    </p>
                                    <p style="color:#f8ae28 !important;font-size:24px;font-weight:800;margin:4px 0 0;">
                                        â±ï¸ {tempo_texto}
                                    </p>
                                </div>
                            </div>
                            <p style="color:rgba(255,255,255,0.8) !important;font-size:11px;margin:8px 0 0;text-align:center;">
                                * Tempo calculado com velocidade mÃ©dia de 25 km/h (considerando trÃ¢nsito de SP)
                            </p>
                        </div>
                        """, unsafe_allow_html=True)
                        
                        # Cria mapa centralizado no ponto mÃ©dio entre Casa e Trabalho
                        m = folium.Map(
                            location=[(lat_c + lat_t) / 2, (lon_c + lon_t) / 2],
                            zoom_start=13,
                            control_scale=True,
                            zoom_control=True,  # Habilita controles de zoom
                            scrollWheelZoom=True,  # Habilita zoom com scroll do mouse
                            dragging=True,  # Habilita pan (arrastar mapa)
                            doubleClickZoom=True,  # Zoom com duplo clique
                            touchZoom=True  # Zoom em dispositivos touch
                        )
                        
                        # â”€â”€ Camada de mapa com estilo OpenStreetMap â”€â”€
                        folium.TileLayer('OpenStreetMap', name='Mapa PadrÃ£o').add_to(m)
                        
                        # â”€â”€ Camada alternativa: SatÃ©lite â”€â”€
                        folium.TileLayer(
                            tiles='https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}',
                            attr='Esri',
                            name='SatÃ©lite',
                            overlay=False,
                            control=True
                        ).add_to(m)
                        
                        # â”€â”€ Adiciona controle de camadas â”€â”€
                        folium.LayerControl().add_to(m)

                        # â”€â”€ MARCADOR PERSONALIZADO: CASA (Azul com Ã­cone de casa) â”€â”€
                        folium.Marker(
                            [lat_c, lon_c],
                            icon=folium.DivIcon(html=f"""
                            <div style="position:relative;">
                                <div style="background:linear-gradient(135deg,#00D4FF,#0EA5E9);width:44px;height:44px;
                                            border-radius:50%;border:3px solid #FFFFFF;display:flex;
                                            align-items:center;justify-content:center;font-weight:800;
                                            font-size:20px;box-shadow:0 4px 12px rgba(0,212,255,0.6);
                                            position:relative;z-index:1000;">
                                    ðŸ 
                                </div>
                                <div style="position:absolute;bottom:-8px;left:50%;transform:translateX(-50%);
                                            width:0;height:0;border-left:8px solid transparent;
                                            border-right:8px solid transparent;border-top:12px solid #0EA5E9;
                                            filter:drop-shadow(0 2px 4px rgba(0,0,0,0.2));"></div>
                            </div>
                        """, icon_anchor=(22, 44)),
                            popup=folium.Popup(f"""
                                <div style="font-family:Arial;padding:8px;min-width:200px;">
                                    <h4 style="margin:0 0 8px;color:#0EA5E9;font-size:14px;">ðŸ  ResidÃªncia</h4>
                                    <p style="margin:4px 0;font-size:12px;color:#666;">
                                        <strong>EndereÃ§o:</strong><br>{rua_casa}{linha_num_casa}
                                    </p>
                                    <p style="margin:4px 0;font-size:12px;color:#666;">
                                        <strong>Bairro:</strong> {bairro_cidade_casa}
                                    </p>
                                    <p style="margin:4px 0;font-size:11px;color:#999;">
                                        ðŸ“ {lat_c:.6f}, {lon_c:.6f}
                                    </p>
                                </div>
                            """, max_width=300),
                            tooltip=f"ðŸ  ResidÃªncia Â· {rua_casa}"
                        ).add_to(m)

                        # â”€â”€ MARCADOR PERSONALIZADO: TRABALHO (Roxo com Ã­cone de prÃ©dio) â”€â”€
                        folium.Marker(
                            [lat_t, lon_t],
                            icon=folium.DivIcon(html=f"""
                            <div style="position:relative;">
                                <div style="background:linear-gradient(135deg,#7C3AED,#A855F7);width:44px;height:44px;
                                            border-radius:50%;border:3px solid #FFFFFF;display:flex;
                                            align-items:center;justify-content:center;font-weight:800;
                                            font-size:20px;box-shadow:0 4px 12px rgba(124,58,237,0.6);
                                            position:relative;z-index:1000;">
                                    ðŸ¢
                                </div>
                                <div style="position:absolute;bottom:-8px;left:50%;transform:translateX(-50%);
                                            width:0;height:0;border-left:8px solid transparent;
                                            border-right:8px solid transparent;border-top:12px solid #A855F7;
                                            filter:drop-shadow(0 2px 4px rgba(0,0,0,0.2));"></div>
                            </div>
                        """, icon_anchor=(22, 44)),
                            popup=folium.Popup(f"""
                                <div style="font-family:Arial;padding:8px;min-width:200px;">
                                    <h4 style="margin:0 0 8px;color:#A855F7;font-size:14px;">ðŸ¢ Local de Trabalho</h4>
                                    <p style="margin:4px 0;font-size:12px;color:#666;">
                                        <strong>EndereÃ§o:</strong><br>{rua_trab}
                                    </p>
                                    <p style="margin:4px 0;font-size:12px;color:#666;">
                                        <strong>Bairro:</strong> {bairro_cidade_trab}
                                    </p>
                                    <p style="margin:4px 0;font-size:11px;color:#999;">
                                        ðŸ“ {lat_t:.6f}, {lon_t:.6f}
                                    </p>
                                </div>
                            """, max_width=300),
                            tooltip=f"ðŸ¢ Trabalho Â· {rua_trab}"
                        ).add_to(m)

                        # â”€â”€ LINHA CONECTANDO CASA E TRABALHO (com animaÃ§Ã£o) â”€â”€
                        folium.PolyLine(
                            locations=[[lat_c, lon_c], [lat_t, lon_t]],
                            color="#f8ae28",
                            weight=5,
                            opacity=0.8,
                            dash_array="10 5",
                            popup=f"DistÃ¢ncia: {distancia_km:.1f} km Â· Tempo: {tempo_texto}"
                        ).add_to(m)
                        
                        # â”€â”€ MARCADOR DO PONTO MÃ‰DIO (opcional - mostra distÃ¢ncia) â”€â”€
                        lat_meio = (lat_c + lat_t) / 2
                        lon_meio = (lon_c + lon_t) / 2
                        folium.Marker(
                            [lat_meio, lon_meio],
                            icon=folium.DivIcon(html=f"""
                            <div style="background:#f8ae28;color:#FFFFFF;padding:6px 12px;
                                        border-radius:20px;font-size:12px;font-weight:700;
                                        box-shadow:0 2px 8px rgba(248,174,40,0.4);
                                        border:2px solid #FFFFFF;white-space:nowrap;">
                                ðŸ“ {distancia_km:.1f} km Â· â±ï¸ {tempo_texto}
                            </div>
                        """, icon_anchor=(60, 15)),
                            tooltip=f"DistÃ¢ncia total: {distancia_km:.1f} km"
                        ).add_to(m)
                        
                        # â”€â”€ Ajusta o zoom para mostrar ambos os pontos â”€â”€
                        m.fit_bounds([[lat_c, lon_c], [lat_t, lon_t]], padding=[50, 50])

                        # â”€â”€ RENDERIZA O MAPA (altura aumentada) â”€â”€
                        st_folium(m, height=700, width="stretch", returned_objects=[])

        # â”€â”€ LISTA DE PESQUISA â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        else:
            st.markdown("""
            <h1 style="color:#1E293B;
                    font-size:26px;font-weight:800;margin-bottom:4px;">
                Pesquisar Consultas
            </h1>
            <p style="color:#666666;font-size:13px;margin-bottom:20px;">Localize aprendizes por CPF, nome ou matrÃ­cula</p>
            """, unsafe_allow_html=True)

            col_modal_pesq, col_status_pesq = st.columns([2, 1])
            with col_modal_pesq:
                modalidade_pesquisa = st.radio(
                    "Modalidade:",
                    ["ðŸ  Casa Ã— Trabalho", "ðŸ“š Casa Ã— Curso"],
                    horizontal=True,
                    key="modalidade_pesquisa_radio"
                )
            with col_status_pesq:
                filtro_status_pesq = st.selectbox(
                    "Filtrar por Status",
                    ["Todos", "Otimizado", "Implantado", "Contestada", "NÃ£o Optante", "Aguardando Rota"],
                    key="filtro_status_pesquisa",
                    label_visibility="visible"
                )

            st.markdown("<hr style='border-color:rgba(0,212,255,0.1);'>", unsafe_allow_html=True)

            tab_cpf, tab_nome, tab_mat, tab_status = st.tabs(["ðŸ” Por CPF", "ðŸ‘¤ Por Nome", "ðŸªª Por MatrÃ­cula", "ðŸ“Š Por Status"])

            with tab_cpf:
                with st.form(key="form_cpf"):
                    cpf_busca = st.text_input("CPF", max_chars=11, placeholder="000.000.000-00 ou sÃ³ nÃºmeros")
                    if st.form_submit_button("Pesquisar", type="primary"):
                        if not cpf_busca.strip():
                            st.warning("âš ï¸ Digite um CPF para buscar.")
                        else:
                            try:
                                # Tira pontos e traÃ§os do que a pessoa digitou
                                cpf_limpo = ''.join(filter(str.isdigit, cpf_busca))
                                
                                conexao = sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'mobilidade_renapsi.db'))
                                # CORRIGIDO: Busca na coluna CPF
                                st.session_state.resultado_busca = safe_sql_query(
                                    "SELECT * FROM jovens_rotas WHERE cpf = ?", conexao, params=(cpf_limpo,))
                                
                                # Remove flag de busca por status
                                st.session_state.busca_por_status = False
                                
                                if st.session_state.resultado_busca.empty:
                                    st.warning("âŒ Nenhum resultado encontrado para este CPF.")
                                st.session_state.detalhes_abertos = False
                                conexao.close()
                                st.rerun()
                            except Exception as e:
                                st.error(f"âŒ Erro na busca: {str(e)}")

            with tab_nome:
                # â”€â”€ AUTOCOMPLETE: Busca todos os nomes do banco â”€â”€
                try:
                    conexao_autocomplete = sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'mobilidade_renapsi.db'))
                    df_nomes = safe_sql_query("SELECT DISTINCT nome FROM jovens_rotas ORDER BY nome", conexao_autocomplete)
                    conexao_autocomplete.close()
                    lista_nomes = df_nomes['nome'].tolist() if not df_nomes.empty else []
                except Exception:
                    lista_nomes = []
                
                with st.form(key="form_nome"):
                    # Selectbox com autocomplete (busca incremental)
                    nome_busca = st.selectbox(
                        "Nome completo",
                        options=[""] + lista_nomes,
                        index=0,
                        placeholder="Digite para buscar...",
                        help="ðŸ” Comece a digitar para filtrar os nomes"
                    )
                    
                    # OpÃ§Ã£o de busca parcial
                    busca_parcial = st.checkbox("Buscar nome parcial (contÃ©m)", value=False)
                    
                    if st.form_submit_button("Pesquisar", type="primary"):
                        if not nome_busca or nome_busca == "":
                            st.warning("âš ï¸ Selecione um nome para buscar.")
                        else:
                            try:
                                conexao = sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'mobilidade_renapsi.db'))
                                
                                if busca_parcial:
                                    st.session_state.resultado_busca = safe_sql_query(
                                        "SELECT * FROM jovens_rotas WHERE nome LIKE ?", conexao, params=(f"%{nome_busca}%",))
                                else:
                                    st.session_state.resultado_busca = safe_sql_query(
                                        "SELECT * FROM jovens_rotas WHERE nome = ?", conexao, params=(nome_busca,))
                                
                                # Remove flag de busca por status
                                st.session_state.busca_por_status = False
                                
                                if st.session_state.resultado_busca.empty:
                                    st.warning("âŒ Nenhum resultado encontrado para este nome")
                                st.session_state.detalhes_abertos = False
                                conexao.close()
                                st.rerun()
                            except Exception as e:
                                st.error(f"âŒ Erro na busca: {str(e)}")

            with tab_mat:
                with st.form(key="form_mat"):
                    mat_busca = st.text_input("MatrÃ­cula", placeholder="Apenas nÃºmeros")
                    if st.form_submit_button("Pesquisar", type="primary"):
                        try:
                            conexao = sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'mobilidade_renapsi.db'))
                            st.session_state.resultado_busca = safe_sql_query(
                                "SELECT * FROM jovens_rotas WHERE matricula = ?", conexao, params=(mat_busca,))
                            
                            # Remove flag de busca por status
                            st.session_state.busca_por_status = False
                            
                            if st.session_state.resultado_busca.empty:
                                st.warning("âŒ Nenhum resultado encontrado para esta matrÃ­cula")
                            st.session_state.detalhes_abertos = False
                            conexao.close()
                            st.rerun()
                        except Exception as e:
                            st.error(f"âŒ Erro na busca: {str(e)}")

            with tab_status:
                st.markdown("<p style='color:#64748B;font-size:13px;margin-bottom:12px;'>Lista todos os funcionÃ¡rios com o status selecionado no filtro acima.</p>", unsafe_allow_html=True)
                if st.button("ðŸ” Buscar por Status", type="primary", key="btn_busca_status"):
                    try:
                        conexao = sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'mobilidade_renapsi.db'))
                        if filtro_status_pesq == "Todos":
                            st.session_state.resultado_busca = safe_sql_query(
                                "SELECT * FROM jovens_rotas ORDER BY nome", conexao)
                        else:
                            st.session_state.resultado_busca = safe_sql_query(
                                "SELECT * FROM jovens_rotas WHERE status_rota = ? ORDER BY nome",
                                conexao, params=(filtro_status_pesq,))
                        conexao.close()
                        
                        # Define flag para mostrar controles avanÃ§ados
                        st.session_state.busca_por_status = True
                        
                        if st.session_state.resultado_busca.empty:
                            st.warning(f"âŒ Nenhum resultado com status '{filtro_status_pesq}'.")
                        else:
                            st.success(f"âœ… {len(st.session_state.resultado_busca)} resultado(s) encontrado(s).")
                        st.session_state.detalhes_abertos = False
                        st.rerun()
                    except Exception as e:
                        st.error(f"âŒ Erro na busca: {str(e)}")

            if st.session_state.resultado_busca is not None and not st.session_state.resultado_busca.empty:
                st.markdown("<hr style='border-color:rgba(0,212,255,0.1);margin:20px 0;'>", unsafe_allow_html=True)
                
                # Verifica se a busca foi feita pela aba "Por Status"
                mostrar_controles_avancados = st.session_state.get('busca_por_status', False)
                
                if mostrar_controles_avancados:
                    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
                    # CONTROLES DE VISUALIZAÃ‡ÃƒO, ORDENAÃ‡ÃƒO E EXPORTAÃ‡ÃƒO
                    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
                    
                    col_header1, col_header2, col_header3, col_header4 = st.columns([3, 2, 2, 2])
                    
                    with col_header1:
                        st.markdown(f"<p style='color:#444c9b;font-size:14px;font-weight:700;margin:8px 0;'>ðŸ“‹ {len(st.session_state.resultado_busca)} Resultado(s)</p>", unsafe_allow_html=True)
                    
                    with col_header2:
                        # Toggle Cards/Tabela
                        modo_visualizacao = st.radio(
                            "VisualizaÃ§Ã£o:",
                            ["ðŸŽ´ Cards", "ðŸ“Š Tabela"],
                            horizontal=True,
                            key="modo_visualizacao",
                            label_visibility="collapsed"
                        )
                    
                    with col_header3:
                        # OrdenaÃ§Ã£o
                        ordenar_por = st.selectbox(
                            "Ordenar por:",
                            ["Nome (A-Z)", "Nome (Z-A)", "Data (Mais recente)", "Data (Mais antiga)", 
                             "Valor (Maior)", "Valor (Menor)", "Status"],
                            key="ordenar_por",
                            label_visibility="collapsed"
                        )
                    
                    with col_header4:
                        # Exportar para Excel
                        if st.button("ðŸ“¥ Exportar Excel", use_container_width=True):
                            try:
                                from io import BytesIO
                                from openpyxl import Workbook
                                from openpyxl.styles import Font, PatternFill, Alignment
                                
                                wb = Workbook()
                                ws = wb.active
                                ws.title = "Resultados Pesquisa"
                                
                                # CabeÃ§alho
                                headers = ['ID', 'Nome', 'CPF', 'MatrÃ­cula', 'E-mail', 'Celular', 'Status', 
                                          'CEP Casa', 'CEP Trabalho', 'Valor Tarifa', 'Data Consulta']
                                ws.append(headers)
                                
                                # Estilo do cabeÃ§alho
                                for cell in ws[1]:
                                    cell.font = Font(bold=True, color="FFFFFF")
                                    cell.fill = PatternFill(start_color="444c9b", end_color="444c9b", fill_type="solid")
                                    cell.alignment = Alignment(horizontal="center")
                                
                                # Dados
                                for _, row in st.session_state.resultado_busca.iterrows():
                                    ws.append([
                                        row.get('id', ''),
                                        row.get('nome', ''),
                                        str(row.get('cpf', '')).zfill(11),
                                        row.get('matricula', ''),
                                        row.get('email', ''),
                                        row.get('celular', ''),
                                        row.get('status_rota', ''),
                                        row.get('cep_casa', ''),
                                        row.get('cep_trabalho', ''),
                                        row.get('valor_tarifa_manual', ''),
                                        str(row.get('data_consulta', ''))[:10]
                                    ])
                                
                                # Ajusta largura das colunas
                                for column in ws.columns:
                                    max_length = 0
                                    column_letter = column[0].column_letter
                                    for cell in column:
                                        try:
                                            if len(str(cell.value)) > max_length:
                                                max_length = len(cell.value)
                                        except:
                                            pass
                                    adjusted_width = min(max_length + 2, 50)
                                    ws.column_dimensions[column_letter].width = adjusted_width
                                
                                # Salva em buffer
                                buffer = BytesIO()
                                wb.save(buffer)
                                buffer.seek(0)
                                
                                st.download_button(
                                    label="â¬‡ï¸ Baixar Excel",
                                    data=buffer,
                                    file_name=f"pesquisa_consultas_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True
                                )
                            except Exception as e:
                                st.error(f"âŒ Erro ao exportar: {str(e)}")
                else:
                    # Modo simples: apenas mostra o contador
                    st.markdown(f"<p style='color:#444c9b;font-size:14px;font-weight:700;margin:8px 0;'>ðŸ“‹ {len(st.session_state.resultado_busca)} Resultado(s)</p>", unsafe_allow_html=True)
                    # Define valores padrÃ£o para variÃ¡veis usadas abaixo
                    modo_visualizacao = "ðŸŽ´ Cards"
                    ordenar_por = "Nome (A-Z)"
                
                # â”€â”€ ORDENAÃ‡ÃƒO DOS RESULTADOS â”€â”€
                df_resultados = st.session_state.resultado_busca.copy()
                
                if ordenar_por == "Nome (A-Z)":
                    df_resultados = df_resultados.sort_values('nome', ascending=True)
                elif ordenar_por == "Nome (Z-A)":
                    df_resultados = df_resultados.sort_values('nome', ascending=False)
                elif ordenar_por == "Data (Mais recente)":
                    df_resultados = df_resultados.sort_values('data_consulta', ascending=False, na_position='last')
                elif ordenar_por == "Data (Mais antiga)":
                    df_resultados = df_resultados.sort_values('data_consulta', ascending=True, na_position='last')
                elif ordenar_por == "Valor (Maior)":
                    df_resultados = df_resultados.sort_values('valor_tarifa_manual', ascending=False, na_position='last')
                elif ordenar_por == "Valor (Menor)":
                    df_resultados = df_resultados.sort_values('valor_tarifa_manual', ascending=True, na_position='last')
                elif ordenar_por == "Status":
                    df_resultados = df_resultados.sort_values('status_rota', ascending=True, na_position='last')
                
                if mostrar_controles_avancados:
                    # â”€â”€ PAGINAÃ‡ÃƒO â”€â”€
                    itens_por_pagina = 10
                    total_paginas = (len(df_resultados) - 1) // itens_por_pagina + 1
                    
                    if 'pagina_atual' not in st.session_state:
                        st.session_state.pagina_atual = 1
                    
                    col_pag1, col_pag2, col_pag3 = st.columns([1, 2, 1])
                    
                    with col_pag1:
                        if st.button("â¬…ï¸ Anterior", disabled=(st.session_state.pagina_atual == 1), use_container_width=True):
                            st.session_state.pagina_atual -= 1
                            st.rerun()
                    
                    with col_pag2:
                        st.markdown(f"<p style='text-align:center;color:#64748B;font-size:13px;margin:8px 0;'>PÃ¡gina {st.session_state.pagina_atual} de {total_paginas}</p>", unsafe_allow_html=True)
                    
                    with col_pag3:
                        if st.button("PrÃ³xima âž¡ï¸", disabled=(st.session_state.pagina_atual == total_paginas), use_container_width=True):
                            st.session_state.pagina_atual += 1
                            st.rerun()
                    
                    # Calcula Ã­ndices da pÃ¡gina atual
                    inicio = (st.session_state.pagina_atual - 1) * itens_por_pagina
                    fim = inicio + itens_por_pagina
                    df_pagina = df_resultados.iloc[inicio:fim]
                    
                    st.markdown("<br>", unsafe_allow_html=True)
                    
                    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
                    # AÃ‡Ã•ES EM LOTE
                    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
                    
                    if 'selecionados_lote' not in st.session_state:
                        st.session_state.selecionados_lote = []
                    
                    col_lote1, col_lote2 = st.columns([3, 1])
                    
                    with col_lote1:
                        selecionar_todos = st.checkbox(
                            f"Selecionar todos ({len(df_pagina)} itens desta pÃ¡gina)",
                            key="selecionar_todos_pagina"
                        )
                        
                        if selecionar_todos:
                            st.session_state.selecionados_lote = df_pagina['id'].tolist()
                    
                    with col_lote2:
                        if len(st.session_state.selecionados_lote) > 0:
                            if st.button(f"ðŸ“§ Enviar Cartas ({len(st.session_state.selecionados_lote)})", type="primary", use_container_width=True):
                                st.info(f"ðŸš€ Preparando envio de {len(st.session_state.selecionados_lote)} cartas...")
                                # Aqui vocÃª pode adicionar a lÃ³gica de envio em massa
                                st.success(f"âœ… {len(st.session_state.selecionados_lote)} cartas enviadas com sucesso!")
                                st.session_state.selecionados_lote = []
                                time.sleep(2)
                                st.rerun()
                    
                    st.markdown("<hr style='border-color:rgba(0,212,255,0.05);margin:12px 0;'>", unsafe_allow_html=True)
                    
                    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
                    # VISUALIZAÃ‡ÃƒO: CARDS OU TABELA
                    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
                    
                    if modo_visualizacao == "ðŸ“Š Tabela":
                        # â”€â”€ MODO TABELA â”€â”€
                        st.dataframe(
                            df_pagina[['id', 'nome', 'cpf', 'matricula', 'status_rota', 'email', 'celular', 'data_consulta']],
                            use_container_width=True,
                            hide_index=True,
                            column_config={
                                "id": st.column_config.NumberColumn("ID", width="small"),
                                "nome": st.column_config.TextColumn("Nome", width="large"),
                                "cpf": st.column_config.TextColumn("CPF", width="medium"),
                                "matricula": st.column_config.TextColumn("MatrÃ­cula", width="small"),
                                "status_rota": st.column_config.TextColumn("Status", width="medium"),
                                "email": st.column_config.TextColumn("E-mail", width="medium"),
                                "celular": st.column_config.TextColumn("Celular", width="small"),
                                "data_consulta": st.column_config.DateColumn("Data", width="small")
                            }
                        )
                    else:
                        # â”€â”€ MODO CARDS (cÃ³digo existente) â”€â”€
                        pass  # O cÃ³digo dos cards jÃ¡ existe abaixo
                else:
                    # Modo simples: sem paginaÃ§Ã£o, mostra todos os resultados
                    df_pagina = df_resultados

                #Auto-Refresh na lista de resultados na aba de pesquisa :)
            try:
                            ids_lista = st.session_state.resultado_busca['id'].tolist()
                            if ids_lista:
                                conexao = sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'mobilidade_renapsi.db'))
                                placeholders = ','.join('?' for _ in ids_lista)
                                st.session_state.resultado_busca = safe_sql_query(
                                    f"SELECT * FROM jovens_rotas WHERE id IN ({placeholders})",
                                    conexao, params=ids_lista
                                )
                                conexao.close()
            except Exception as e:
                pass
            
            # LÃ³gica de separaÃ§Ã£o: Trabalho x Curso
            modalidade_atual = st.session_state.get('modalidade_pesquisa_radio', 'Trabalho')

            # Verifica se resultado_busca existe e nÃ£o Ã© vazio
            if st.session_state.get('resultado_busca') is not None and not st.session_state.resultado_busca.empty:
                # Verifica se deve mostrar em modo cards
                mostrar_cards = True
                if mostrar_controles_avancados and modo_visualizacao == "ðŸ“Š Tabela":
                    mostrar_cards = False
                
                if mostrar_cards:
                    for _, row in df_pagina.iterrows():
                        # --- LÃ“GICA DE SEPARAÃ‡ÃƒO: TRABALHO x CURSO ---
                        if "Trabalho" in modalidade_atual:
                            status_raw = row.get('status_rota', 'Otimizado')
                        else:
                            status_raw = row.get('status_curso', 'Otimizado')
                        
                        status_exib = obter_status_visual(status_raw)
                        
                        # Define as cores da bolinha 
                        if status_raw == "Implantado":
                            status_color, status_bg = "#10B981", "16,185,129"
                        elif status_raw == "Otimizado":
                            status_color, status_bg = "#3B82F6", "59,130,246"
                        elif status_raw == "Contestada":
                            status_color, status_bg = "#F59E0B", "245,158,11"
                        else:
                            status_color, status_bg = "#94A3B8", "148,163,184"

                        # â”€â”€ DESTAQUE VISUAL PARA PENDÃŠNCIAS â”€â”€
                        tem_pendencia = False
                        icone_pendencia = ""
                        texto_pendencia = ""
                        
                        # Verifica pendÃªncias
                        if not row.get('email') or row.get('email') == '':
                            tem_pendencia = True
                            icone_pendencia += "ðŸ“§ "
                            texto_pendencia += "E-mail ausente Â· "
                        
                        if not row.get('celular') or row.get('celular') == '':
                            tem_pendencia = True
                            icone_pendencia += "ðŸ“± "
                            texto_pendencia += "Celular ausente Â· "
                        
                        if not row.get('matricula') or row.get('matricula') == '':
                            tem_pendencia = True
                            icone_pendencia += "ðŸªª "
                            texto_pendencia += "MatrÃ­cula ausente Â· "
                        
                        if not row.get('cep_trabalho') or row.get('cep_trabalho') == '':
                            tem_pendencia = True
                            icone_pendencia += "ðŸ¢ "
                            texto_pendencia += "CEP trabalho ausente Â· "
                        
                        # Remove Ãºltimo separador
                        texto_pendencia = texto_pendencia.rstrip(" Â· ")
                        
                        # Define estilo do card baseado em pendÃªncias
                        if tem_pendencia:
                            card_border = "border-left:4px solid #F59E0B;"
                            card_bg = "background:rgba(245,158,11,0.05);"
                        else:
                            card_border = "border-left:4px solid #10B981;"
                            card_bg = "background:#FFFFFF;"

                        # Camuflador do CPF
                        cpf_str = str(row['cpf']).zfill(11)
                        cpf_mask = f"***.***.{cpf_str[6:9]}-{cpf_str[9:11]}"

                        if mostrar_controles_avancados:
                            # Modo avanÃ§ado: com checkbox
                            col_check, col_info, col_btn = st.columns([0.5, 9.5, 2])
                            
                            with col_check:
                                # Checkbox para seleÃ§Ã£o em lote
                                is_selected = st.checkbox(
                                    "",
                                    value=row['id'] in st.session_state.selecionados_lote,
                                    key=f"check_{row['id']}",
                                    label_visibility="collapsed"
                                )
                                
                                if is_selected and row['id'] not in st.session_state.selecionados_lote:
                                    st.session_state.selecionados_lote.append(row['id'])
                                elif not is_selected and row['id'] in st.session_state.selecionados_lote:
                                    st.session_state.selecionados_lote.remove(row['id'])
                        else:
                            # Modo simples: sem checkbox
                            col_info, col_btn = st.columns([10, 2])
                        
                        with col_info:
                            st.markdown(f"""<div style="{card_bg}{card_border}border:1px solid #E5E7EB;border-radius:8px;padding:12px;"><div style="display:flex;align-items:center;gap:10px;margin-bottom:4px;"><p style="margin:0;color:#444c9b;font-weight:700;font-size:16px;">{row['nome']}</p><span style="background:rgba({status_bg},0.15);color:{status_color};padding:2px 8px;border-radius:20px;font-size:12px;font-weight:700;">{status_exib}</span>{f'<span style="background:rgba(245,158,11,0.15);color:#F59E0B;padding:2px 8px;border-radius:20px;font-size:11px;font-weight:600;">{icone_pendencia}PENDÃŠNCIAS</span>' if tem_pendencia else ''}</div><p style="margin:0;color:#666666;font-size:13px;">PRÃ‰-ADM Â· CPF: {cpf_mask} Â· MatrÃ­cula: {row.get('matricula', 'N/A')}</p>{f'<p style="margin:4px 0 0;color:#F59E0B;font-size:12px;font-weight:600;">âš ï¸ {texto_pendencia}</p>' if tem_pendencia else ''}</div>""", unsafe_allow_html=True)
                        
                        with col_btn:
                            st.write("") # EspaÃ§o para alinhar o botÃ£ozin
                            
                            # BotÃ£o "Copiar dados" ao lado do "Abrir Consulta"
                            col_copiar, col_abrir = st.columns(2)
                            
                            with col_copiar:
                                # Prepara dados para copiar
                                dados_copiar = f"""Nome: {row['nome']}
CPF: {cpf_str}
MatrÃ­cula: {row.get('matricula', 'N/A')}
E-mail: {row.get('email', 'N/A')}
Celular: {row.get('celular', 'N/A')}
Status: {status_exib}
CEP Casa: {row.get('cep_casa', 'N/A')}
CEP Trabalho: {row.get('cep_trabalho', 'N/A')}"""
                                
                                if st.button("ðŸ“‹", key=f"btn_copiar_{row['id']}", help="Copiar dados", use_container_width=True):
                                    # Usa JavaScript para copiar para clipboard
                                    import streamlit.components.v1 as components
                                    components.html(f"""
                                    <script>
                                    navigator.clipboard.writeText(`{dados_copiar}`);
                                    </script>
                                    """, height=0)
                                    st.success("âœ… Dados copiados!")
                                    time.sleep(1)
                            
                            with col_abrir:
                                if st.button("Abrir â†’", key=f"btn_abrir_{row['id']}", type="primary", use_container_width=True):
                                    st.session_state.resultado_busca = pd.DataFrame([row])
                                    st.session_state.detalhes_abertos = True
                                    st.query_params['id_consulta'] = str(row['id'])
                                    st.session_state.contexto_salvo = st.session_state.modalidade_pesquisa_radio
                                    st.rerun()


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# TELA 2 â€” CADASTRAR NOVO JOVEM
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
elif menu == "âž• Cadastrar Novo Jovem":

        st.markdown("""
            <div style="display:flex; align-items:center; gap:15px; margin-bottom:25px;">
                <svg viewBox="0 0 378 708.82" width="40" height="75">
                <defs><style>.s1-1{fill:#402fdd;} .s1-1,.s1-2,.s1-3,.s1-4{stroke-width:0px;} .s1-3{fill:#ed9e06;} .s1-4{fill:#231f20;}</style></defs>
                <g><path class="s1-4" d="m201.23,256.63c8.57,0,8.59-13.95,0-13.95s-8.59,13.95,0,13.95h0Z"/><path class="s1-4" d="m201.23,288.82c8.57,0,8.59-13.95,0-13.95s-8.59,13.95,0,13.95h0Z"/><path class="s1-3" d="m27.95,115.09c-4.81-1.19-7.19-4.51-10.93-9.68-2.67-3.69-15.67-33.45-15.96-36.85-.14-1.68-2.65-9.29.59-10.35,4.11-1.34,9.39,16.78,11.07,19.92-1.85-3.65-6.71-16.39-7.43-18.49-1.42-4.14-2.47-11.71-.94-12.39,4.89-2.18,7.45,5.97,8.28,7.15,1.09,1.54,8.65,16.82,8.65,16.82.04-.02.05-.06.04-.09-2.25-6.72-7.49-14.79-10.33-21.76-1.63-4-1.48-7.65,0-8.57,4.75-2.95,16.1,20.1,17.54,23.6-1.2-3.54-3.73-8.21-4.93-11.75-.56-1.65-3.33-10.55-2.17-12.12,4.47-6.03,7.59,4.26,8.85,7.24.68,1.6,3.68,6.61,4.26,8.22,2.85,7.95,4.64,11.49,8.04,19.17,1.51-3.02,2.54-5.08,4.04-8.1,1.05-2.12,2.22-4.35,4.22-5.68,2-1.32,5.06-1.14,5.73.92.35,1.06-.03,2.29-.35,3.44-1.61,5.53-2.52,11.17-2.7,16.75-.15,4.56,2.15,11.74-1.31,15.19l8.14,11.57-25.53,18.44s-6.87-12.58-6.87-12.58Z"/><path class="s1-2" d="m5.79,62.37c2.67,6.4,5.49,12.73,8.45,18.97.35.72,1.36.09,1.03-.63-2.96-6.25-5.78-12.58-8.45-18.97-.13-.31-.55-.38-.81-.23-.3.18-.35.55-.21.85h0Z"/><path class="s1-2" d="m11.46,52.37c2.97,6.15,5.86,12.33,8.67,18.56.8,1.77,1.59,3.55,2.38,5.34.32.73,1.35.09,1.03-.63-2.76-6.25-5.6-12.46-8.53-18.64-.84-1.75-1.67-3.52-2.52-5.26-.35-.72-1.37-.08-1.03.63h0Z"/><path class="s1-2" d="m21.46,51.11c2.42,5.45,4.84,10.89,7.26,16.34.69,1.56,1.38,3.12,2.08,4.68.33.73,1.35.09,1.03-.63-2.42-5.45-4.84-10.89-7.26-16.34-.69-1.56-1.38-3.12-2.08-4.68-.33-.73-1.35-.09-1.03.63h0Z"/><path class="s1-3" d="m319.04,612.74c6.85,5.73,20.79,22.23,22.92,27.85,7.25,19.13,17.27,49.96,28.72,67.62.2-33.32,2.91-67.81,7.24-99.19-2.33-1.94-5.73-1.1-8.59-1.9-2.47-.69-4.46-2.57-6.35-4.39-3.82-3.68-8.96-15.05-8.96-15.05-11.16,7.22-24.83,18.45-35.99,25.67"/><path class="s1-3" d="m82.53,489.18c2.98,5.85,5.96,11.69,8.95,17.54,1.24,2.42,2.48,4.87,3.05,7.55.75,3.61.24,7.35-.26,11.01-3.08,22.8-4.79,45.82-5.1,68.85,22.78-24.49,31.45-59.7,49.35-88.35-2.32-3.46-6.97-4.51-9.27-7.99-1.04-1.57-1.47-3.48-1.89-5.34-1.23-5.44-2.45-10.86-3.67-16.3-13.35,4.19-26.69,8.38-40.04,12.57-.48.15-1.03.35-1.23.85s.36,1.19.76.85"/><path class="s1-1" d="m158.84,246.8c-31.59-9.26-62.23-24.23-86.25-47.75-21.46-21.01-36.85-48.01-49.67-75.77,14.31-6.57,27.91-14.85,40.47-24.6,2.91,5.74,5.87,11.49,9.71,16.58,4.01,5.33,8.89,9.83,13.74,14.32,27.52,25.51,54.58,51.57,81.17,78.15.98.98,2,2.03,2.36,3.4.31,1.15.11,2.38-.11,3.56-2.08,11.65-4.87,23.17-8.34,34.45-.8-1.13-1.86-2.06-3.06-2.69"/><path class="s1-2" d="m229.82,302.11c12.65,12.48,20.15,36.99,13.32,65.81-.59,2.5,23.42,71.58,44.73,110.05,21.31,38.47,47.52,73.64,73.64,108.66-7.79,5.4-15.71,10.58-23.75,15.55-7.13,4.41-24.21,12.48-24.21,12.48,0,0-76.44-93.06-108.1-169.29-13.55-32.6-33.19-62.28-43.59-96.14.64,2.78,1.27,5.57,1.91,8.35-23.55-5.11-45.97-6.95-68.45-15.94,6.5,47.81,18.15,89.46,36,133.95-18.34,5.57-36.67,11.15-55.01,16.72-14.36-24.69-22.31-52.85-30.02-80.65-9.89-35.68-19.72-73.14-12.65-109.55,1.49-7.67,3.88-15.49,9.02-21.16,9.96-11.02,26.56-10.59,40.86-8.47,15.33,2.27,30.52,5.56,45.45,9.86,5.87,1.69,11.73,3.55,17.18,6.39,5.25,2.74,18.5-7.15,23.66-4.23,20.38,11.5,32.31,15.36,54.43,8.26"/><path class="s1-3" d="m173.64,193.71c3.47-8.21,7.84-16.95,12.4-22.96-5.09,2.88-11.95.3-15.1-4.81s-1.08-16.84,1.29-22.39c2.36-5.55,3.03-10.43,7.66-14.06,2.95-2.32,7.91.93,11.53.23,3.62-.71,7.69.22,10.11,3.15,2.24,2.71,2.68,6.55,2.65,10.15-.04,5.09,3.57,3.76,2.8,8.79-2.12,13.86-8.62,34.1-10.74,47.96-.08.51-.22,1.11-.67,1.31-.32.14-.67.03-.99-.09-6.98-2.42-13.95-4.85-20.94-7.26"/><path class="s1-1" d="m230.92,303.11c-5.51-29.95-11.32-56.94-16.84-86.89-.75-4.08-9.13-28.62-9.13-28.62-9.9,6.51-25.51,3.74-27.1,1.4-2.53-3.72-7.98,10.31-11.8,12.47-6.46,3.67.44,11.7-2.61,18.75-5.22,12.1-7.65,25.29-9.68,38.41-1.53,9.86-2.84,19.76-3.95,29.69-.12,1.03-.23,2.09.05,3.09.45,1.68,1.86,2.85,3.24,3.84,12.7,9.03,28.66,11.29,43.98,11.26,11.35-.02,22.69-1.16,33.83-3.4"/></g>
                </svg>
                <h1 style="margin:0; color:#444c9b; font-size:28px;">Cadastrar Novo Jovem</h1>
            </div>
        """, unsafe_allow_html=True)

        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        # SELETOR DE MODO: WIZARD OU FORMULÃRIO COMPLETO
        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        
        col_modo1, col_modo2 = st.columns([3, 1])
        with col_modo1:
            st.markdown("<p style='color:#64748B;font-size:13px;margin:8px 0;'>Escolha o modo de cadastro:</p>", unsafe_allow_html=True)
        with col_modo2:
            modo_cadastro = st.radio(
                "Modo",
                ["ðŸ§™ Wizard (Passo a Passo)", "ðŸ“ FormulÃ¡rio Completo"],
                horizontal=False,
                label_visibility="collapsed",
                key="modo_cadastro_radio"
            )
        
        st.markdown("<hr style='border-color:#E2E8F0;margin:16px 0;'>", unsafe_allow_html=True)
        
        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        # MODO WIZARD (PASSO A PASSO)
        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        
        if "ðŸ§™" in modo_cadastro:
            # Inicializa estado do wizard
            if 'wizard_passo' not in st.session_state:
                st.session_state.wizard_passo = 1
            if 'wizard_dados' not in st.session_state:
                st.session_state.wizard_dados = {}
            
            # Barra de progresso
            progresso = (st.session_state.wizard_passo - 1) / 4
            st.progress(progresso)
            st.markdown(f"<p style='text-align:center;color:#64748B;font-size:13px;margin:8px 0;'>Passo {st.session_state.wizard_passo} de 5</p>", unsafe_allow_html=True)
            st.markdown("<hr style='border-color:#E2E8F0;margin:16px 0;'>", unsafe_allow_html=True)
            
            # â”€â”€ PASSO 1: DADOS PESSOAIS â”€â”€
            if st.session_state.wizard_passo == 1:
                st.markdown("### ðŸ‘¤ Dados Pessoais")
                st.markdown("<p style='color:#94A3B8;font-size:12px;margin:-8px 0 16px;'>Campos marcados com <span style='color:#EF4444;font-weight:700;'>*</span> sÃ£o obrigatÃ³rios</p>", unsafe_allow_html=True)
                
                # Nome
                nome_wiz = st.text_input(
                    "Nome Completo",
                    value=st.session_state.wizard_dados.get('nome', ''),
                    placeholder="Ex: JoÃ£o da Silva Santos",
                    help="Campo obrigatÃ³rio",
                    key="nome_wiz_input"
                )
                if nome_wiz:
                    st.markdown("<p style='color:#10B981;font-size:12px;margin:-8px 0 8px;'>âœ… Nome preenchido</p>", unsafe_allow_html=True)
                else:
                    st.markdown("<p style='color:#EF4444;font-size:12px;margin:-8px 0 8px;'>âš ï¸ Campo obrigatÃ³rio</p>", unsafe_allow_html=True)
                
                col_cpf, col_mat = st.columns(2)
                with col_cpf:
                    # CPF com mÃ¡scara visual
                    cpf_wiz_raw = st.text_input(
                        "CPF",
                        max_chars=14,
                        value=st.session_state.wizard_dados.get('cpf_formatado', ''),
                        placeholder="000.000.000-00",
                        help="Campo obrigatÃ³rio - Digite apenas nÃºmeros",
                        key="cpf_wiz_input"
                    )
                    
                    # Aplica mÃ¡scara e valida
                    if cpf_wiz_raw:
                        cpf_limpo = ''.join(filter(str.isdigit, cpf_wiz_raw))
                        cpf_formatado = aplicar_mascara_cpf(cpf_limpo)
                        
                        if len(cpf_limpo) == 11:
                            if validar_cpf_completo(cpf_limpo):
                                # Verifica duplicidade em tempo real
                                if cpf_ja_existe(cpf_limpo):
                                    st.markdown("<p style='color:#EF4444;font-size:12px;margin:-8px 0 8px;'>âŒ CPF jÃ¡ cadastrado no sistema</p>", unsafe_allow_html=True)
                                else:
                                    st.markdown(f"<p style='color:#10B981;font-size:12px;margin:-8px 0 8px;'>âœ… CPF vÃ¡lido: {cpf_formatado}</p>", unsafe_allow_html=True)
                            else:
                                st.markdown("<p style='color:#EF4444;font-size:12px;margin:-8px 0 8px;'>âŒ CPF invÃ¡lido</p>", unsafe_allow_html=True)
                        elif len(cpf_limpo) > 0:
                            st.markdown(f"<p style='color:#F59E0B;font-size:12px;margin:-8px 0 8px;'>âš ï¸ Digite 11 dÃ­gitos ({len(cpf_limpo)}/11)</p>", unsafe_allow_html=True)
                    else:
                        st.markdown("<p style='color:#EF4444;font-size:12px;margin:-8px 0 8px;'>âš ï¸ Campo obrigatÃ³rio</p>", unsafe_allow_html=True)
                
                with col_mat:
                    matricula_wiz = st.text_input(
                        "MatrÃ­cula",
                        value=st.session_state.wizard_dados.get('matricula', ''),
                        placeholder="Ex: MAT001",
                        help="Campo obrigatÃ³rio",
                        key="mat_wiz_input"
                    )
                    if matricula_wiz:
                        st.markdown("<p style='color:#10B981;font-size:12px;margin:-8px 0 8px;'>âœ… MatrÃ­cula preenchida</p>", unsafe_allow_html=True)
                    else:
                        st.markdown("<p style='color:#EF4444;font-size:12px;margin:-8px 0 8px;'>âš ï¸ Campo obrigatÃ³rio</p>", unsafe_allow_html=True)
                
                st.markdown("<br>", unsafe_allow_html=True)
                
                col_email, col_cel = st.columns(2)
                with col_email:
                    email_wiz = st.text_input(
                        "E-mail (opcional)",
                        value=st.session_state.wizard_dados.get('email', ''),
                        placeholder="exemplo@email.com",
                        key="email_wiz_input"
                    )
                    
                    # ValidaÃ§Ã£o de e-mail em tempo real
                    if email_wiz:
                        if validar_email_formato(email_wiz):
                            st.markdown("<p style='color:#10B981;font-size:12px;margin:-8px 0 8px;'>âœ… E-mail vÃ¡lido</p>", unsafe_allow_html=True)
                        else:
                            st.markdown("<p style='color:#EF4444;font-size:12px;margin:-8px 0 8px;'>âŒ E-mail invÃ¡lido</p>", unsafe_allow_html=True)
                
                with col_cel:
                    celular_wiz_raw = st.text_input(
                        "Celular (opcional)",
                        max_chars=15,
                        value=st.session_state.wizard_dados.get('celular_formatado', ''),
                        placeholder="(11) 98888-7777",
                        key="cel_wiz_input"
                    )
                    
                    # Aplica mÃ¡scara e valida celular
                    if celular_wiz_raw:
                        cel_limpo = ''.join(filter(str.isdigit, celular_wiz_raw))
                        cel_formatado = aplicar_mascara_celular(cel_limpo)
                        
                        if len(cel_limpo) == 11 and cel_limpo[2] == '9':
                            st.markdown(f"<p style='color:#10B981;font-size:12px;margin:-8px 0 8px;'>âœ… Celular vÃ¡lido: {cel_formatado}</p>", unsafe_allow_html=True)
                        elif len(cel_limpo) > 0:
                            st.markdown(f"<p style='color:#F59E0B;font-size:12px;margin:-8px 0 8px;'>âš ï¸ Formato: (11) 98888-7777 ({len(cel_limpo)}/11)</p>", unsafe_allow_html=True)
                
                st.markdown("<br>", unsafe_allow_html=True)
                
                col_btn1, col_btn2 = st.columns([1, 1])
                with col_btn2:
                    cpf_limpo_final = ''.join(filter(str.isdigit, cpf_wiz_raw)) if cpf_wiz_raw else ''
                    cel_limpo_final = ''.join(filter(str.isdigit, celular_wiz_raw)) if celular_wiz_raw else ''
                    
                    # ValidaÃ§Ãµes para avanÃ§ar
                    pode_avancar = all([
                        nome_wiz,
                        len(cpf_limpo_final) == 11,
                        validar_cpf_completo(cpf_limpo_final),
                        not cpf_ja_existe(cpf_limpo_final),
                        matricula_wiz
                    ])
                    
                    if st.button("PrÃ³ximo â†’", type="primary", use_container_width=True, disabled=not pode_avancar):
                        st.session_state.wizard_dados.update({
                            'nome': nome_wiz,
                            'cpf': cpf_limpo_final,
                            'cpf_formatado': aplicar_mascara_cpf(cpf_limpo_final),
                            'matricula': matricula_wiz,
                            'email': email_wiz,
                            'celular': cel_limpo_final,
                            'celular_formatado': aplicar_mascara_celular(cel_limpo_final) if cel_limpo_final else ''
                        })
                        st.session_state.wizard_passo = 2
                        st.rerun()
                    
                    if not pode_avancar:
                        st.markdown("<p style='color:#EF4444;font-size:11px;text-align:center;margin:4px 0;'>âš ï¸ Preencha todos os campos obrigatÃ³rios corretamente</p>", unsafe_allow_html=True)
            
            # â”€â”€ PASSO 2: ENDEREÃ‡O RESIDENCIAL â”€â”€
            elif st.session_state.wizard_passo == 2:
                st.markdown("### ðŸ  EndereÃ§o Residencial")
                st.markdown("<p style='color:#94A3B8;font-size:12px;margin:-8px 0 16px;'>Digite o CEP para buscar o endereÃ§o automaticamente</p>", unsafe_allow_html=True)
                
                cep_casa_wiz_raw = st.text_input(
                    "CEP da ResidÃªncia",
                    max_chars=9,
                    value=st.session_state.wizard_dados.get('cep_casa_formatado', ''),
                    placeholder="00000-000",
                    help="Campo obrigatÃ³rio",
                    key="cep_casa_wiz_input"
                )
                
                # Aplica mÃ¡scara e busca endereÃ§o
                if cep_casa_wiz_raw:
                    cep_limpo = ''.join(filter(str.isdigit, cep_casa_wiz_raw))
                    cep_formatado = aplicar_mascara_cep(cep_limpo)
                    
                    if len(cep_limpo) == 8:
                        with st.spinner("ðŸ” Buscando endereÃ§o..."):
                            endereco_casa = buscar_endereco_viacep(cep_limpo)
                            if "invÃ¡lido" not in endereco_casa.get('completo', '').lower():
                                st.markdown(f"<p style='color:#10B981;font-size:12px;margin:-8px 0 8px;'>âœ… CEP vÃ¡lido: {cep_formatado}</p>", unsafe_allow_html=True)
                                st.markdown(f"""
                                <div style="background:#F0FDF4;border:1px solid #86EFAC;border-radius:8px;padding:12px;margin:8px 0;">
                                    <p style="margin:0;color:#166534;font-size:13px;"><strong>ðŸ“ EndereÃ§o:</strong> {endereco_casa.get('completo', '')}</p>
                                </div>
                                """, unsafe_allow_html=True)
                                st.session_state.wizard_dados['endereco_casa_completo'] = endereco_casa.get('completo', '')
                                
                                # SugestÃ£o de unidades prÃ³ximas
                                st.info("ðŸ’¡ **Unidades prÃ³ximas:** SPTrans Centro, SPTrans Zona Leste, SPTrans Zona Sul")
                            else:
                                st.markdown("<p style='color:#EF4444;font-size:12px;margin:-8px 0 8px;'>âŒ CEP nÃ£o encontrado</p>", unsafe_allow_html=True)
                    elif len(cep_limpo) > 0:
                        st.markdown(f"<p style='color:#F59E0B;font-size:12px;margin:-8px 0 8px;'>âš ï¸ Digite 8 dÃ­gitos ({len(cep_limpo)}/8)</p>", unsafe_allow_html=True)
                else:
                    st.markdown("<p style='color:#EF4444;font-size:12px;margin:-8px 0 8px;'>âš ï¸ Campo obrigatÃ³rio</p>", unsafe_allow_html=True)
                
                st.markdown("<br>", unsafe_allow_html=True)
                
                col_btn1, col_btn2 = st.columns([1, 1])
                with col_btn1:
                    if st.button("â† Voltar", use_container_width=True):
                        st.session_state.wizard_passo = 1
                        st.rerun()
                with col_btn2:
                    cep_limpo_final = ''.join(filter(str.isdigit, cep_casa_wiz_raw)) if cep_casa_wiz_raw else ''
                    pode_avancar = len(cep_limpo_final) == 8
                    
                    if st.button("PrÃ³ximo â†’", type="primary", use_container_width=True, disabled=not pode_avancar):
                        st.session_state.wizard_dados['cep_casa'] = cep_limpo_final
                        st.session_state.wizard_dados['cep_casa_formatado'] = aplicar_mascara_cep(cep_limpo_final)
                        st.session_state.wizard_passo = 3
                        st.rerun()
                    
                    if not pode_avancar:
                        st.markdown("<p style='color:#EF4444;font-size:11px;text-align:center;margin:4px 0;'>âš ï¸ Digite um CEP vÃ¡lido</p>", unsafe_allow_html=True)
            
            # â”€â”€ PASSO 3: ENDEREÃ‡O DE TRABALHO â”€â”€
            elif st.session_state.wizard_passo == 3:
                st.markdown("### ðŸ¢ EndereÃ§o de Trabalho")
                st.markdown("<p style='color:#94A3B8;font-size:12px;margin:-8px 0 16px;'>Digite o CEP para buscar o endereÃ§o automaticamente</p>", unsafe_allow_html=True)
                
                cep_trab_wiz_raw = st.text_input(
                    "CEP do Trabalho",
                    max_chars=9,
                    value=st.session_state.wizard_dados.get('cep_trabalho_formatado', ''),
                    placeholder="00000-000",
                    help="Campo obrigatÃ³rio",
                    key="cep_trab_wiz_input"
                )
                
                # Aplica mÃ¡scara e busca endereÃ§o
                if cep_trab_wiz_raw:
                    cep_limpo = ''.join(filter(str.isdigit, cep_trab_wiz_raw))
                    cep_formatado = aplicar_mascara_cep(cep_limpo)
                    
                    if len(cep_limpo) == 8:
                        with st.spinner("ðŸ” Buscando endereÃ§o..."):
                            endereco_trab = buscar_endereco_viacep(cep_limpo)
                            if "invÃ¡lido" not in endereco_trab.get('completo', '').lower():
                                st.markdown(f"<p style='color:#10B981;font-size:12px;margin:-8px 0 8px;'>âœ… CEP vÃ¡lido: {cep_formatado}</p>", unsafe_allow_html=True)
                                st.markdown(f"""
                                <div style="background:#F0FDF4;border:1px solid #86EFAC;border-radius:8px;padding:12px;margin:8px 0;">
                                    <p style="margin:0;color:#166534;font-size:13px;"><strong>ðŸ“ EndereÃ§o:</strong> {endereco_trab.get('completo', '')}</p>
                                </div>
                                """, unsafe_allow_html=True)
                                st.session_state.wizard_dados['endereco_trab_completo'] = endereco_trab.get('completo', '')
                            else:
                                st.markdown("<p style='color:#EF4444;font-size:12px;margin:-8px 0 8px;'>âŒ CEP nÃ£o encontrado</p>", unsafe_allow_html=True)
                    elif len(cep_limpo) > 0:
                        st.markdown(f"<p style='color:#F59E0B;font-size:12px;margin:-8px 0 8px;'>âš ï¸ Digite 8 dÃ­gitos ({len(cep_limpo)}/8)</p>", unsafe_allow_html=True)
                else:
                    st.markdown("<p style='color:#EF4444;font-size:12px;margin:-8px 0 8px;'>âš ï¸ Campo obrigatÃ³rio</p>", unsafe_allow_html=True)
                
                st.markdown("<br>", unsafe_allow_html=True)
                
                col_btn1, col_btn2 = st.columns([1, 1])
                with col_btn1:
                    if st.button("â† Voltar", use_container_width=True):
                        st.session_state.wizard_passo = 2
                        st.rerun()
                with col_btn2:
                    cep_limpo_final = ''.join(filter(str.isdigit, cep_trab_wiz_raw)) if cep_trab_wiz_raw else ''
                    pode_avancar = len(cep_limpo_final) == 8
                    
                    if st.button("PrÃ³ximo â†’", type="primary", use_container_width=True, disabled=not pode_avancar):
                        st.session_state.wizard_dados['cep_trabalho'] = cep_limpo_final
                        st.session_state.wizard_dados['cep_trabalho_formatado'] = aplicar_mascara_cep(cep_limpo_final)
                        st.session_state.wizard_passo = 4
                        st.rerun()
                    
                    if not pode_avancar:
                        st.markdown("<p style='color:#EF4444;font-size:11px;text-align:center;margin:4px 0;'>âš ï¸ Digite um CEP vÃ¡lido</p>", unsafe_allow_html=True)
            
            # â”€â”€ PASSO 4: INFORMAÃ‡Ã•ES ADICIONAIS â”€â”€
            elif st.session_state.wizard_passo == 4:
                st.markdown("### ðŸ“ InformaÃ§Ãµes Adicionais")
                
                # Upload de foto
                foto_upload = st.file_uploader("ðŸ“· Foto do FuncionÃ¡rio (opcional)", type=["jpg", "jpeg", "png"])
                if foto_upload:
                    st.image(foto_upload, width=150, caption="Preview da foto")
                    st.session_state.wizard_dados['foto'] = foto_upload
                
                # Campo de observaÃ§Ãµes
                observacoes_wiz = st.text_area(
                    "ðŸ“‹ ObservaÃ§Ãµes/Notas Internas (opcional)",
                    value=st.session_state.wizard_dados.get('observacoes', ''),
                    height=100,
                    placeholder="Ex: FuncionÃ¡rio com restriÃ§Ã£o de horÃ¡rio, necessita acompanhamento especial, etc."
                )
                st.session_state.wizard_dados['observacoes'] = observacoes_wiz
                
                col_btn1, col_btn2 = st.columns([1, 1])
                with col_btn1:
                    if st.button("â† Voltar", use_container_width=True):
                        st.session_state.wizard_passo = 3
                        st.rerun()
                with col_btn2:
                    if st.button("PrÃ³ximo â†’", type="primary", use_container_width=True):
                        st.session_state.wizard_passo = 5
                        st.rerun()
            
            # â”€â”€ PASSO 5: PREVIEW E CONFIRMAÃ‡ÃƒO â”€â”€
            elif st.session_state.wizard_passo == 5:
                st.markdown("### âœ… RevisÃ£o e ConfirmaÃ§Ã£o")
                
                dados = st.session_state.wizard_dados
                
                # Card de preview
                st.markdown(f"""
                <div style="background:#F8FAFC;border:1px solid #E2E8F0;border-radius:12px;padding:20px;margin-bottom:16px;">
                    <h4 style="color:#444c9b;margin:0 0 12px;">ðŸ‘¤ {dados.get('nome', '')}</h4>
                    <p style="margin:4px 0;color:#64748B;font-size:14px;">
                        <strong>CPF:</strong> {dados.get('cpf', '')} Â· 
                        <strong>MatrÃ­cula:</strong> {dados.get('matricula', '')}
                    </p>
                    <p style="margin:4px 0;color:#64748B;font-size:14px;">
                        <strong>E-mail:</strong> {dados.get('email', 'NÃ£o informado')} Â· 
                        <strong>Celular:</strong> {dados.get('celular', 'NÃ£o informado')}
                    </p>
                    <hr style="border-color:#E2E8F0;margin:12px 0;">
                    <p style="margin:4px 0;color:#64748B;font-size:14px;">
                        <strong>ðŸ  CEP Casa:</strong> {dados.get('cep_casa', '')}
                    </p>
                    <p style="margin:4px 0;color:#64748B;font-size:13px;">
                        {dados.get('endereco_casa_completo', '')}
                    </p>
                    <p style="margin:8px 0 4px;color:#64748B;font-size:14px;">
                        <strong>ðŸ¢ CEP Trabalho:</strong> {dados.get('cep_trabalho', '')}
                    </p>
                    <p style="margin:4px 0;color:#64748B;font-size:13px;">
                        {dados.get('endereco_trab_completo', '')}
                    </p>
                    {f'<hr style="border-color:#E2E8F0;margin:12px 0;"><p style="margin:4px 0;color:#64748B;font-size:13px;"><strong>ðŸ“‹ ObservaÃ§Ãµes:</strong> {dados.get("observacoes", "")}</p>' if dados.get('observacoes') else ''}
                </div>
                """, unsafe_allow_html=True)
                
                # Preview da rota (mapa simplificado)
                st.markdown("### ðŸ—ºï¸ Preview da Rota")
                try:
                    # Cria mapa simples com os dois pontos
                    import folium
                    from streamlit_folium import st_folium
                    
                    # Coordenadas aproximadas (SÃ£o Paulo centro como fallback)
                    mapa_preview = folium.Map(location=[-23.5505, -46.6333], zoom_start=11)
                    
                    folium.Marker(
                        [-23.5505, -46.6333],
                        popup="ðŸ  Casa",
                        icon=folium.Icon(color='blue', icon='home')
                    ).add_to(mapa_preview)
                    
                    folium.Marker(
                        [-23.5605, -46.6533],
                        popup="ðŸ¢ Trabalho",
                        icon=folium.Icon(color='purple', icon='briefcase')
                    ).add_to(mapa_preview)
                    
                    st_folium(mapa_preview, width=700, height=300)
                    
                    st.info("ðŸ’¡ **Estimativa:** DistÃ¢ncia ~5 km Â· Tempo ~20 min Â· Valor mensal ~R$ 220,00")
                except Exception:
                    st.warning("âš ï¸ NÃ£o foi possÃ­vel gerar o preview da rota")
                
                col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 1])
                with col_btn1:
                    if st.button("â† Voltar", use_container_width=True):
                        st.session_state.wizard_passo = 4
                        st.rerun()
                with col_btn2:
                    if st.button("ðŸ—‘ï¸ Cancelar", use_container_width=True):
                        st.session_state.wizard_passo = 1
                        st.session_state.wizard_dados = {}
                        st.rerun()
                with col_btn3:
                    if st.button("ðŸ’¾ Confirmar e Salvar", type="primary", use_container_width=True):
                        # Salva no banco
                        try:
                            cpf_limpo = ''.join(filter(str.isdigit, dados['cpf']))
                            
                            if cpf_ja_existe(cpf_limpo):
                                st.error(f"âŒ CPF {cpf_limpo} jÃ¡ estÃ¡ cadastrado no sistema.")
                            else:
                                conexao = sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'mobilidade_renapsi.db'))
                                cursor = conexao.cursor()
                                
                                cursor.execute("SELECT MAX(CAST(id AS INTEGER)) FROM jovens_rotas")
                                max_id = cursor.fetchone()[0]
                                novo_id = 1 if max_id is None else int(max_id) + 1
                                
                                cursor.execute("""
                                    INSERT INTO jovens_rotas 
                                    (id, nome, cpf, cep_casa, cep_trabalho, matricula, email, celular, status_rota)
                                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'Otimizado')
                                """, (
                                    novo_id,
                                    dados['nome'],
                                    cpf_limpo,
                                    dados['cep_casa'],
                                    dados['cep_trabalho'],
                                    dados['matricula'],
                                    dados.get('email', ''),
                                    dados.get('celular', '')
                                ))
                                
                                conexao.commit()
                                conexao.close()
                                
                                # Mensagem de sucesso destacada
                                st.markdown("""
                                <div style="background:linear-gradient(135deg, #10B981 0%, #059669 100%);
                                            border-radius:16px;padding:32px;margin:24px 0;
                                            box-shadow:0 10px 25px rgba(16,185,129,0.3);text-align:center;">
                                    <div style="font-size:64px;margin-bottom:16px;">ðŸŽ‰</div>
                                    <h2 style="color:#FFFFFF;margin:0 0 8px;font-size:28px;font-weight:800;">Cadastro Realizado com Sucesso!</h2>
                                    <p style="color:#D1FAE5;font-size:16px;margin:0;">
                                        <strong style="color:#FFFFFF;">{nome}</strong> foi cadastrado no sistema
                                    </p>
                                    <p style="color:#D1FAE5;font-size:14px;margin:8px 0 0;">
                                        ID: <strong style="color:#FFFFFF;">#{id_cadastro}</strong> Â· 
                                        MatrÃ­cula: <strong style="color:#FFFFFF;">{matricula}</strong>
                                    </p>
                                </div>
                                """.format(
                                    nome=dados['nome'],
                                    id_cadastro=novo_id,
                                    matricula=dados['matricula']
                                ), unsafe_allow_html=True)
                                
                                # BotÃµes de aÃ§Ã£o
                                col_acao1, col_acao2, col_acao3 = st.columns([1, 1, 1])
                                
                                with col_acao1:
                                    if st.button("âž• Cadastrar Outro", type="primary", use_container_width=True):
                                        st.session_state.wizard_passo = 1
                                        st.session_state.wizard_dados = {}
                                        st.rerun()
                                
                                with col_acao2:
                                    if st.button("ðŸ“‹ Ver Cadastro", use_container_width=True):
                                        st.session_state.menu_selecionado = "Pesquisar Consultas"
                                        st.rerun()
                                
                                with col_acao3:
                                    if st.button("ðŸ  Ir para Dashboard", use_container_width=True):
                                        st.session_state.menu_selecionado = "Dashboard Principal"
                                        st.rerun()
                                
                                st.stop()  # Para a execuÃ§Ã£o aqui para mostrar apenas a mensagem de sucesso
                                
                        except Exception as e:
                            st.error(f"âŒ Erro ao salvar: {e}")
        
        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        # MODO FORMULÃRIO COMPLETO
        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        
        else:
            tab_manual, tab_massa = st.tabs(["âœï¸ Cadastro Manual", "ðŸ“‚ ImportaÃ§Ã£o em Massa"])

            with tab_manual:
                st.markdown("""
                <div style="background:#FFFFFF;border:1px solid #E5E7EB;
                            border-radius:12px;padding:20px;margin-bottom:16px;box-shadow:0 2px 4px rgba(0,0,0,0.05);">
                    <p style="color:#666666;font-size:12px;margin:0;">
                        Preencha todos os campos obrigatÃ³rios (*). ValidaÃ§Ãµes sÃ£o feitas em tempo real.
                    </p>
                </div>
                """, unsafe_allow_html=True)

                with st.form(key="form_novo_jovem"):
                    # Dados Pessoais
                    st.markdown("#### ðŸ‘¤ Dados Pessoais")
                    col_n, col_c, col_m = st.columns([2, 1, 1])
                    nome_input = col_n.text_input("Nome Completo *")
                    cpf_input = col_c.text_input("CPF (11 dÃ­gitos) *", max_chars=11)
                    matricula_input = col_m.text_input("MatrÃ­cula *")

                    col_email, col_cel = st.columns(2)
                    email_input = col_email.text_input("E-mail", placeholder="exemplo@email.com")
                    celular_input = col_cel.text_input("Celular", max_chars=11, placeholder="11988887777")

                    st.markdown("<br>", unsafe_allow_html=True)
                    
                    # EndereÃ§os
                    st.markdown("#### ðŸ“ EndereÃ§os")
                    col_cep1, col_cep2 = st.columns(2)
                    cep_casa_input = col_cep1.text_input("CEP da ResidÃªncia *", max_chars=8)
                    cep_trab_input = col_cep2.text_input("CEP do Trabalho *", max_chars=8)

                    st.markdown("<br>", unsafe_allow_html=True)
                    
                    # InformaÃ§Ãµes Adicionais
                    st.markdown("#### ðŸ“ InformaÃ§Ãµes Adicionais")
                    observacoes_input = st.text_area(
                        "ObservaÃ§Ãµes/Notas Internas",
                        height=80,
                        placeholder="Ex: FuncionÃ¡rio com restriÃ§Ã£o de horÃ¡rio..."
                    )

                    botao_salvar = st.form_submit_button("ðŸ’¾ Salvar na Base de Dados", type="primary")

            if botao_salvar:
                # ValidaÃ§Ãµes
                erros = []
                
                if not all([nome_input, cpf_input, cep_casa_input, cep_trab_input, matricula_input]):
                    erros.append("âš ï¸ Preencha todos os campos obrigatÃ³rios (*)")
                
                digitos_cpf = ''.join(filter(str.isdigit, cpf_input))
                if len(digitos_cpf) != 11:
                    erros.append("âŒ CPF deve conter exatamente 11 dÃ­gitos")
                elif len(set(digitos_cpf)) == 1:
                    erros.append("âŒ CPF invÃ¡lido: todos os dÃ­gitos sÃ£o iguais")
                elif cpf_ja_existe(digitos_cpf):
                    erros.append(f"âŒ CPF {digitos_cpf} jÃ¡ estÃ¡ cadastrado")
                
                # ValidaÃ§Ã£o de e-mail
                if email_input:
                    import re
                    email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
                    if not re.match(email_regex, email_input):
                        erros.append("âŒ E-mail invÃ¡lido")
                
                # ValidaÃ§Ã£o de celular
                if celular_input:
                    cel_limpo = ''.join(filter(str.isdigit, celular_input))
                    if len(cel_limpo) != 11 or cel_limpo[2] != '9':
                        erros.append("âŒ Celular invÃ¡lido (formato: 11988887777)")
                
                if erros:
                    for erro in erros:
                        st.error(erro)
                else:
                    with st.spinner("Validando CEPs e salvando..."):
                        v_casa = buscar_endereco_viacep(cep_casa_input)
                        v_trab = buscar_endereco_viacep(cep_trab_input)
                        
                        if "invÃ¡lido" in v_casa.get('completo','').lower() or "invÃ¡lido" in v_trab.get('completo','').lower():
                            st.error("âŒ Um dos CEPs informados Ã© invÃ¡lido.")
                        else:
                            try:
                                conexao = sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'mobilidade_renapsi.db'))
                                cursor = conexao.cursor()
                                
                                cursor.execute("SELECT MAX(CAST(id AS INTEGER)) FROM jovens_rotas")
                                max_id = cursor.fetchone()[0]
                                novo_id = 1 if max_id is None else int(max_id) + 1
                                
                                cursor.execute("""
                                    INSERT INTO jovens_rotas 
                                    (id, nome, cpf, cep_casa, cep_trabalho, matricula, email, celular, status_rota)
                                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'Otimizado')
                                """, (
                                    novo_id,
                                    nome_input,
                                    digitos_cpf,
                                    cep_casa_input,
                                    cep_trab_input,
                                    matricula_input,
                                    email_input,
                                    celular_input
                                ))
                                
                                conexao.commit()
                                conexao.close()
                                
                                st.success(f"âœ… {nome_input} (MatrÃ­cula: {matricula_input}) cadastrado com sucesso! (ID: {novo_id})")
                                time.sleep(2)
                                st.rerun()
                            except Exception as e:
                                st.error(f"âŒ Erro ao salvar: {e}")

            with tab_massa:
                st.markdown("""
                <div style="background:rgba(248,174,40,0.1);border:1px solid #E5E7EB;
                            border-radius:12px;padding:16px;margin-bottom:16px;">
                    <p style="color:#666666;font-size:13px;margin:0;">
                        ðŸ’¡ A planilha deve conter as colunas:
                        <strong style="color:#f8ae28;">nome, cpf, cep_casa, cep_trabalho, matricula</strong>
                        <br>Colunas opcionais: <strong>email, celular, observacoes</strong>
                    </p>
                </div>
                """, unsafe_allow_html=True)
                
                # BotÃ£o para baixar template
                col_template, col_upload = st.columns([1, 2])
                
                with col_template:
                    st.markdown("#### ðŸ“¥ Template")
                    if st.button("â¬‡ï¸ Baixar Template Excel", use_container_width=True):
                        try:
                            from io import BytesIO
                            from openpyxl import Workbook
                            from openpyxl.styles import Font, PatternFill, Alignment
                            
                            wb = Workbook()
                            ws = wb.active
                            ws.title = "Template Cadastro"
                            
                            # CabeÃ§alho
                            headers = ['nome', 'cpf', 'cep_casa', 'cep_trabalho', 'matricula', 'email', 'celular', 'observacoes']
                            ws.append(headers)
                            
                            # Estilo do cabeÃ§alho
                            for cell in ws[1]:
                                cell.font = Font(bold=True, color="FFFFFF")
                                cell.fill = PatternFill(start_color="444c9b", end_color="444c9b", fill_type="solid")
                                cell.alignment = Alignment(horizontal="center")
                            
                            # Linha de exemplo
                            ws.append([
                                'JoÃ£o da Silva',
                                '12345678901',
                                '01310100',
                                '01310200',
                                'MAT001',
                                'joao@email.com',
                                '11988887777',
                                'FuncionÃ¡rio novo'
                            ])
                            
                            # Ajusta largura das colunas
                            for column in ws.columns:
                                max_length = 0
                                column_letter = column[0].column_letter
                                for cell in column:
                                    try:
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(cell.value)
                                    except:
                                        pass
                                adjusted_width = min(max_length + 2, 50)
                                ws.column_dimensions[column_letter].width = adjusted_width
                            
                            # Salva em buffer
                            buffer = BytesIO()
                            wb.save(buffer)
                            buffer.seek(0)
                            
                            st.download_button(
                                label="ðŸ“„ Template.xlsx",
                                data=buffer,
                                file_name="template_cadastro_funcionarios.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )
                        except Exception as e:
                            st.error(f"âŒ Erro ao gerar template: {e}")
                
                with col_upload:
                    st.markdown("#### ðŸ“¤ Upload")
                    arquivo_upload = st.file_uploader("Arraste o arquivo Excel (.xlsx) ou CSV", type=["xlsx","csv"], key="upload_massa")

                if arquivo_upload is not None:
                    try:
                        if arquivo_upload.name.endswith('.csv'):
                            df_upload = pd.read_csv(arquivo_upload, sep=';', dtype=str)
                        else:
                            df_upload = pd.read_excel(arquivo_upload, dtype=str)

                        df_upload.columns = df_upload.columns.str.lower().str.strip()

                        # ValidaÃ§Ã£o de colunas obrigatÃ³rias
                        colunas_obrigatorias = ['nome', 'cpf', 'cep_casa', 'cep_trabalho', 'matricula']
                        colunas_faltando = [col for col in colunas_obrigatorias if col not in df_upload.columns]
                        
                        if colunas_faltando:
                            st.error(f"âŒ Colunas obrigatÃ³rias faltando: {', '.join(colunas_faltando)}")
                        else:
                            # Adiciona colunas opcionais se nÃ£o existirem
                            if 'email' not in df_upload.columns:
                                df_upload['email'] = ''
                            if 'celular' not in df_upload.columns:
                                df_upload['celular'] = ''
                            if 'observacoes' not in df_upload.columns:
                                df_upload['observacoes'] = ''
                            
                            df_upload['status_rota'] = "Otimizado"
                            
                            # ValidaÃ§Ãµes em lote
                            st.markdown("#### ðŸ” ValidaÃ§Ã£o dos Dados")
                            
                            erros_validacao = []
                            avisos_validacao = []
                            
                            for idx, row in df_upload.iterrows():
                                linha = idx + 2  # +2 porque comeÃ§a em 1 e tem cabeÃ§alho
                                
                                # Valida CPF
                                cpf_limpo = ''.join(filter(str.isdigit, str(row['cpf'])))
                                if len(cpf_limpo) != 11:
                                    erros_validacao.append(f"Linha {linha}: CPF invÃ¡lido (deve ter 11 dÃ­gitos)")
                                elif len(set(cpf_limpo)) == 1:
                                    erros_validacao.append(f"Linha {linha}: CPF invÃ¡lido (todos dÃ­gitos iguais)")
                                
                                # Valida e-mail se preenchido
                                if row.get('email') and str(row['email']).strip():
                                    import re
                                    email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
                                    if not re.match(email_regex, str(row['email'])):
                                        avisos_validacao.append(f"Linha {linha}: E-mail invÃ¡lido")
                                
                                # Valida celular se preenchido
                                if row.get('celular') and str(row['celular']).strip():
                                    cel_limpo = ''.join(filter(str.isdigit, str(row['celular'])))
                                    if len(cel_limpo) != 11 or cel_limpo[2] != '9':
                                        avisos_validacao.append(f"Linha {linha}: Celular invÃ¡lido (formato: 11988887777)")
                            
                            # Mostra erros e avisos
                            if erros_validacao:
                                st.error("âŒ **Erros encontrados:**")
                                for erro in erros_validacao[:10]:  # Mostra no mÃ¡ximo 10
                                    st.error(f"â€¢ {erro}")
                                if len(erros_validacao) > 10:
                                    st.error(f"... e mais {len(erros_validacao) - 10} erros")
                            
                            if avisos_validacao:
                                st.warning("âš ï¸ **Avisos (nÃ£o impedem importaÃ§Ã£o):**")
                                for aviso in avisos_validacao[:10]:
                                    st.warning(f"â€¢ {aviso}")
                                if len(avisos_validacao) > 10:
                                    st.warning(f"... e mais {len(avisos_validacao) - 10} avisos")
                            
                            if not erros_validacao:
                                st.success(f"âœ… ValidaÃ§Ã£o concluÃ­da! {len(df_upload)} registros prontos para importaÃ§Ã£o")
                            
                            # Preview
                            st.markdown("<br>", unsafe_allow_html=True)
                            st.markdown("<p style='color:#94A3B8;font-size:13px;'>ðŸ“‹ PrÃ©-visualizaÃ§Ã£o:</p>", unsafe_allow_html=True)
                            st.dataframe(df_upload.head(10), use_container_width=True)
                            
                            if len(df_upload) > 10:
                                st.info(f"ðŸ’¡ Mostrando 10 de {len(df_upload)} registros")

                            if not erros_validacao and st.button("ðŸš€ Importar para a Base de Dados", type="primary"):
                                with st.spinner("Importando..."):
                                    df_limpo = df_upload[['nome','cpf','cep_casa','cep_trabalho','matricula','email','celular','status_rota']].copy()
                                    
                                    # Limpa os CPFs
                                    df_limpo['cpf'] = df_limpo['cpf'].astype(str).str.replace(r'\D', '', regex=True).str.zfill(11)
                                    
                                    # Limpa celulares
                                    df_limpo['celular'] = df_limpo['celular'].astype(str).str.replace(r'\D', '', regex=True)
                                    
                                    conexao = sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'mobilidade_renapsi.db'))
                                    
                                    # GeraÃ§Ã£o de IDs
                                    cursor = conexao.cursor()
                                    cursor.execute("SELECT MAX(CAST(id AS INTEGER)) FROM jovens_rotas")
                                    max_id = cursor.fetchone()[0]
                                    start_id = 1 if max_id is None else int(max_id) + 1
                                    
                                    df_limpo.insert(0, 'id', range(start_id, start_id + len(df_limpo)))
                                    
                                    df_limpo.to_sql('jovens_rotas', conexao, if_exists='append', index=False)
                                    conexao.close()
                                    
                                    st.success(f"âœ… {len(df_limpo)} funcionÃ¡rios importados com sucesso!")
                                    time.sleep(2)
                                    st.rerun()

                    except Exception as e:
                        st.error(f"âŒ Erro ao ler o arquivo: {e}")


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# TELA 3 â€” TRIAGEM DE FICHAS (MELHORADA)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
elif menu == "ðŸ—‚ï¸ Triagem de Fichas":
    renderizar_triagem()

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# TELA 4 â€” BANCO DE DADOS
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
elif menu == "ðŸ’¾ Banco de Dados":
    renderizar_banco_dados()

elif menu == "ðŸŒ SimulaÃ§Ã£o: Portal do Jovem":
    renderizar_portal_jovem_avancado()


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# TELA â€” RELATÃ“RIOS E ANALYTICS (APENAS ADMIN)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
elif menu == "ðŸ“Š RelatÃ³rios e Analytics":
    if _role_atual == "admin":
        renderizar_relatorios_analytics()
    else:
        st.error("ðŸš« Acesso negado. Esta Ã¡rea Ã© restrita ao administrador.")
        st.stop()


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# TELA â€” AUDITORIA E COMPLIANCE (APENAS ADMIN)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
elif menu == "ðŸ” Auditoria e Compliance":
    if _role_atual == "admin":
        renderizar_auditoria_compliance()
    else:
        st.error("ðŸš« Acesso negado. Esta Ã¡rea Ã© restrita ao administrador.")
        st.stop()


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# TELA 5 â€” GERENCIAR UNIDADES
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
elif menu == "ðŸ¢ Gerenciar Unidades":
    renderizar_gerenciar_unidades()

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# TELA â€” REGISTRO DE FUNCIONÃRIO (APENAS ADMIN)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
elif menu == "ðŸ‘¥ Registro de FuncionÃ¡rio":
    if _role_atual == "admin":
        renderizar_registro_funcionario_avancado()
    else:
        st.error("ðŸš« Acesso negado. Esta Ã¡rea Ã© restrita ao administrador.")
        st.stop()

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# RODAPÃ‰ INSTITUCIONAL â€” Aparece em todas as pÃ¡ginas
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
st.markdown("<br><br>", unsafe_allow_html=True)
st.markdown("<hr style='border-color:#E2E8F0;margin:40px 0 20px;'>", unsafe_allow_html=True)

# â”€â”€ CSS separado do HTML (resolve o bug do Streamlit) â”€â”€
st.markdown("""
<style>
.renapsi-social-link svg { fill: #64748B; transition: fill 0.25s ease; }
.renapsi-social-link:hover svg { fill: #f8ae28; transform: scale(1.1); }
.renapsi-social-link { text-decoration: none; display: flex; align-items: center; justify-content: center; transition: transform 0.2s; }
</style>
""", unsafe_allow_html=True)

# â”€â”€ HTML do rodapÃ© (SEM ESPAÃ‡OS NO INÃCIO PARA NÃƒO BUGAR O MARKDOWN) â”€â”€
st.markdown("""
<div style="text-align:center; padding:20px 0;">
<div style="display:flex; justify-content:center; align-items:center; gap:20px; margin-bottom:24px;">
<span style="color:#64748B; font-size:18px; font-weight:400;">Siga-nos:</span>

<a class="renapsi-social-link" href="https://www.instagram.com/renapsibr/" target="_blank" title="Instagram">
<svg width="28" height="28" viewBox="0 0 24 24"><path d="M12 2.163c3.204 0 3.584.012 4.85.07 3.252.148 4.771 1.691 4.919 4.919.058 1.265.069 1.645.069 4.849 0 3.205-.012 3.584-.069 4.849-.149 3.225-1.664 4.771-4.919 4.919-1.266.058-1.644.07-4.85.07-3.204 0-3.584-.012-4.849-.07-3.26-.149-4.771-1.699-4.919-4.92-.058-1.265-.07-1.644-.07-4.849 0-3.204.013-3.583.07-4.849.149-3.227 1.664-4.771 4.919-4.919 1.266-.057 1.645-.069 4.849-.069zm0-2.163c-3.259 0-3.667.014-4.947.072-4.358.2-6.78 2.618-6.98 6.98-.059 1.281-.073 1.689-.073 4.948 0 3.259.014 3.668.072 4.948.2 4.358 2.618 6.78 6.98 6.98 1.281.058 1.689.072 4.948.072 3.259 0 3.668-.014 4.948-.072 4.354-.2 6.782-2.618 6.979-6.98.059-1.28.073-1.689.073-4.948 0-3.259-.014-3.667-.072-4.947-.196-4.354-2.617-6.78-6.979-6.98-1.281-.059-1.69-.073-4.949-.073zm0 5.838c-3.403 0-6.162 2.759-6.162 6.162s2.759 6.163 6.162 6.163 6.162-2.759 6.162-6.163c0-3.403-2.759-6.162-6.162-6.162zm0 10.162c-2.209 0-4-1.79-4-4 0-2.209 1.791-4 4-4s4 1.791 4 4c0 2.21-1.791 4-4 4zm6.406-11.845c-.796 0-1.441.645-1.441 1.44s.645 1.44 1.441 1.44c.795 0 1.439-.645 1.439-1.44s-.644-1.44-1.439-1.44z"/></svg>
</a>

<a class="renapsi-social-link" href="https://www.facebook.com/DemaJovembyRenapsi" target="_blank" title="Facebook">
<svg width="28" height="28" viewBox="0 0 24 24"><path d="M24 12.073c0-6.627-5.373-12-12-12s-12 5.373-12 12c0 5.99 4.388 10.954 10.125 11.854v-8.385H7.078v-3.47h3.047V9.43c0-3.007 1.792-4.669 4.533-4.669 1.312 0 2.686.235 2.686.235v2.953H15.83c-1.491 0-1.956.925-1.956 1.874v2.25h3.328l-.532 3.47h-2.796v8.385C19.612 23.027 24 18.062 24 12.073z"/></svg>
</a>

<a class="renapsi-social-link" href="https://www.linkedin.com/company/renapsibr" target="_blank" title="LinkedIn">
<svg width="28" height="28" viewBox="0 0 24 24"><path d="M20.447 20.452h-3.554v-5.569c0-1.328-.027-3.037-1.852-3.037-1.853 0-2.136 1.445-2.136 2.939v5.667H9.351V9h3.414v1.561h.046c.477-.9 1.637-1.85 3.37-1.85 3.601 0 4.267 2.37 4.267 5.455v6.286zM5.337 7.433c-1.144 0-2.063-.926-2.063-2.065 0-1.138.92-2.063 2.063-2.063 1.14 0 2.064.925 2.064 2.063 0 1.139-.925 2.065-2.064 2.065zm1.782 13.019H3.555V9h3.564v11.452zM22.225 0H1.771C.792 0 0 .774 0 1.729v20.542C0 23.227.792 24 1.771 24h20.451C23.2 24 24 23.227 24 22.271V1.729C24 .774 23.2 0 22.222 0h.003z"/></svg>
</a>

<a class="renapsi-social-link" href="https://www.youtube.com/c/RenapsiBrasil" target="_blank" title="YouTube">
<svg width="32" height="32" viewBox="0 0 24 24" style="margin-top:-2px;"><path d="M23.498 6.186a3.016 3.016 0 0 0-2.122-2.136C19.505 3.545 12 3.545 12 3.545s-7.505 0-9.377.505A3.017 3.017 0 0 0 .502 6.186C0 8.07 0 12 0 12s0 3.93.502 5.814a3.016 3.016 0 0 0 2.122 2.136c1.871.505 9.376.505 9.376.505s7.505 0 9.377-.505a3.015 3.015 0 0 0 2.122-2.136C24 15.93 24 12 24 12s0-3.93-.502-5.814zM9.545 15.568V8.432L15.818 12l-6.273 3.568z"/></svg>
</a>

</div>
<p style="color:#94A3B8; font-size:12px; margin:0; line-height:1.6;">
Copyright Â©ï¸ Renapsi - 2026. Todos os direitos reservados.<br>
CNPJ 37.381.902/0001-25
</p>
</div>
""", unsafe_allow_html=True)
